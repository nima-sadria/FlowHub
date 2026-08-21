"""Application services for source mappings and internal FlowHub Sheets."""

from __future__ import annotations

import base64
import csv
import io
import json
import re
import uuid
from collections import defaultdict
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import tuple_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.flowhub.auth.models import FlowHubUser
from app.flowhub.business_observability.models import BusinessEvent
from app.flowhub.data_layer.models import (
    DlConnectorHealth,
    DlConnectorTelemetry,
    DlInvalidationEvent,
    DlInventoryCache,
    DlProductCache,
    DlRefreshJob,
    DlSourceDiscoveryLock,
    DlSourceDiscoveryReservation,
    DlSourceReadLock,
    DlSourceReadReservation,
    DlSourceSnapshot,
    DlWorksheetDiscoveryCache,
    DlWorkspacePreview,
)
from app.flowhub.integration_platform.models import (
    IntegrationConnectorDiagnostic,
    IntegrationConnectorEvent,
    IntegrationConnectorHealthSnapshot,
    IntegrationConnectorInstance,
    IntegrationConnectorSetting,
    IntegrationConnectorTelemetry,
    IntegrationPollingPolicy,
    IntegrationWebhookEvent,
)
from app.flowhub.pricing_matrix.service import PricingMatrixService
from app.flowhub.setup.models import FlowHubAppConfig
from app.flowhub.setup.service import AppConfigService
from app.flowhub.source_acquisition.models import (
    ACTIVE_RUN_STATUSES,
    AcquisitionRun,
    SourceMappingSchemaExpectation,
    SourceObservation,
    SourceObservationDataset,
    SourceObservationVersionHead,
    SourceObservationWorksheetDataset,
    SourceSchemaAssessment,
)
from app.flowhub.source_workspace.formula import (
    FORMULA_ENGINE_VERSION,
    FormulaResult,
    calculate_sheet,
    column_name,
)
from app.flowhub.source_workspace.identity import canonical_nextcloud_connector_id
from app.flowhub.source_workspace.models import (
    FlowHubSheet,
    SheetCell,
    SheetColumn,
    SheetImportJob,
    SheetRevision,
    SheetRow,
    SourceChannelFieldMapping,
    SourceChannelMapping,
    SourceDataQualityIssue,
    SourceDataQualityScan,
    SourceDataQualityScanSource,
    SourceFieldMapping,
    SourceMappingIdentityAssessment,
    SourceMappingRevision,
    SourceProductIdentity,
    SourceProfile,
    SourceWorksheetChannelFieldMapping,
    SourceWorksheetChannelMapping,
    SourceWorksheetFieldMapping,
    SourceWorksheetRule,
    SourceWorksheetRuleSet,
)
from app.flowhub.source_workspace.repositories import (
    DataQualityRepository,
    SheetRepository,
    SourceRepository,
)
from app.flowhub.sources.spreadsheet_source import (
    SOURCE_ID as LEGACY_EXTERNAL_SOURCE_ID,
)
from app.flowhub.sources.spreadsheet_source import (
    SpreadsheetSourceReadService,
    normalize_source_mapping,
)
from app.flowhub.unified_workspace.domain import (
    ApplyState,
    AvailabilitySignal,
    ReviewState,
    SourceInstruction,
    checksum,
    normalize_direct_price,
    normalize_quantity,
    normalize_stock_status,
    resolve_availability,
    utcnow,
)
from app.flowhub.unified_workspace.events import (
    DomainEvent,
    DomainEventBus,
    PersistenceAuditSubscriber,
)
from app.flowhub.unified_workspace.models import (
    ApplyJob,
    CanonicalProduct,
    ChannelCache,
    CurrencyProfile,
    Listing,
    Review,
    UnifiedAuditEntry,
    UnifiedWorkspace,
    WorkspaceChannel,
    WorkspaceSnapshot,
    WorkspaceSourceBinding,
)

MAX_SHEET_ROWS = 10_000
MAX_SHEET_COLUMNS = 200
MAX_IMPORT_BYTES = 20 * 1024 * 1024
SOURCE_FIELDS = {"name", "source_key", "category", "brand", "cost"}
CHANNEL_FIELDS = {"external_id", "price", "stock", "status"}
REFERENCE_TYPES = {"column_letter", "header_name", "column_id", "disabled"}
DEFAULT_VALUE_POLICY = {
    "blank": "no_change",
    "x": "unavailable",
    "dash": "no_change",
    "zero": "explicit_zero",
    "formula": "calculated_value",
    "invalid": "blocked",
}
IDENTITY_ASSESSMENT_ALGORITHM_VERSION = "source-product-key-v1"
SOURCE_KEY_NORMALIZATION_VERSION = "trim-casefold-v1"
IDENTITY_CONFLICT_GROUP_LIMIT = 100
IDENTITY_CONFLICT_ROWS_PER_GROUP_LIMIT = 100
UNSPECIFIED_IDENTITY_AUTHORITY = {
    "type": "unspecified",
    "systemIdentifier": None,
    "displayLabel": None,
}


def _normalize_source_product_key(value: Any) -> str:
    return str(value or "").strip().casefold()


def _source_product_key_hash(value: Any) -> str:
    return checksum(
        {
            "normalizationVersion": SOURCE_KEY_NORMALIZATION_VERSION,
            "normalizedSourceProductKey": _normalize_source_product_key(value),
        }
    )


def _utc_timestamp(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def _iso_utc_timestamp(value: datetime | None) -> str | None:
    normalized = _utc_timestamp(value)
    return normalized.isoformat() if normalized is not None else None


def _id() -> str:
    return str(uuid.uuid4())


def _clean_name(value: object, fallback: str) -> str:
    text = " ".join(str(value or "").strip().split())
    return text[:240] or fallback


def _unprocessable(code: str, message: str, details: dict[str, Any] | None = None) -> HTTPException:
    return HTTPException(
        status.HTTP_422_UNPROCESSABLE_ENTITY,
        {"code": code, "message": message, "details": details or {}},
    )


def _worksheet_read_quota(quota: dict[str, object]) -> dict[str, object]:
    """Translate the canonical read-policy state to this camel-case API."""
    return {
        "enabled": bool(quota["enabled"]),
        "limit": int(quota["limit"]),
        "usage": int(quota["usage"]),
        "remaining": int(quota["remaining"]),
        "resetAt": quota["reset_at"],
        "exhausted": bool(quota["exhausted"]),
    }


def _worksheet_discovery_quota(quota: dict[str, object]) -> dict[str, object]:
    return {
        "enabled": bool(quota["enabled"]),
        "limit": int(quota["limit"]),
        "usage": int(quota["usage"]),
        "remaining": int(quota["remaining"]),
        "resetAt": quota["reset_at"],
        "exhausted": bool(quota["exhausted"]),
    }


class SourceWorkspaceService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.sources = SourceRepository(db)
        self.sheets = SheetRepository(db)
        self.issues = DataQualityRepository(db)

    # -- Source and Mapping -------------------------------------------------

    def list_sources(self, user: FlowHubUser) -> dict[str, Any]:
        return {"items": [self._source_shape(item) for item in self.sources.list_for_user(user.id)]}

    def available_channels(self) -> dict[str, Any]:
        self._ensure_channels()
        from app.flowhub.commerce.service import CommerceHubService

        commerce_channels = {
            item["id"]: item for item in CommerceHubService(self.db).list_channels()["items"]
        }
        channels = (
            self.db.query(WorkspaceChannel)
            .order_by(WorkspaceChannel.name, WorkspaceChannel.id)
            .all()
        )
        return {
            "items": [
                {
                    "channelId": item.id,
                    "name": commerce_channels.get(item.id, {}).get("name") or item.name,
                    "displayNameCustom": bool(
                        commerce_channels.get(item.id, {}).get("display_name_custom")
                    ),
                    "connectorType": item.connector_type,
                    "capabilityVersion": item.capability_version,
                    "capabilities": {
                        **dict(item.capabilities_json or {}),
                        "mappingRequiredFields": list(
                            self._channel_mapping_required_fields(item)
                        ),
                    },
                    "enabled": item.enabled,
                    "implementationState": item.implementation_state,
                    "available": item.enabled and item.implementation_state == "implemented",
                    "configured": bool(
                        commerce_channels.get(item.id, {}).get("credential_status")
                        == "configured"
                    ),
                }
                for item in channels
            ]
        }

    @staticmethod
    def _channel_mapping_required_fields(channel: WorkspaceChannel) -> tuple[str, ...]:
        """Read mapping requirements from the connector capability contract."""
        capabilities = dict(channel.capabilities_json or {})
        advertised = capabilities.get("mappingRequiredFields")
        if advertised is None:
            advertised = capabilities.get("mapping_required_fields")
        if isinstance(advertised, list):
            values = tuple(
                str(field)
                for field in advertised
                if str(field) in CHANNEL_FIELDS
            )
            if len(values) == len(advertised):
                return values
        # Legacy rows receive the provider-neutral identifier minimum. The
        # channel registry refreshes implemented connectors with their exact
        # capability contract before Mapping validation.
        return ("external_id",)

    def create_source(
        self,
        *,
        name: str,
        source_kind: str,
        external_source_id: str | None,
        worksheet_mode: str,
        worksheet_name: str | None,
        data_start_row: int,
        user: FlowHubUser,
        currency: str | None = None,
        currency_unit: str | None = None,
    ) -> dict[str, Any]:
        if source_kind not in {"flowhub_sheet", "imported_sheet", "external"}:
            raise _unprocessable("SOURCE_KIND_INVALID", "Unsupported Source kind.")
        if source_kind == "external" and not external_source_id:
            raise _unprocessable(
                "EXTERNAL_SOURCE_REQUIRED", "External Sources require an existing Source identity."
            )
        if source_kind == "external" and external_source_id:
            requested_binding = (
                self.db.query(SourceProfile)
                .filter(SourceProfile.external_source_id == str(external_source_id).strip())
                .one_or_none()
            )
            if requested_binding is not None:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    {
                        "code": "SOURCE_CONNECTOR_ALREADY_BOUND",
                        "message": (
                            "This connector already belongs to a Source. "
                            "Create a fresh connector for a replacement Source."
                        ),
                        "lifecycle_status": requested_binding.status,
                    },
                )
            if str(external_source_id).strip() in {"nextcloud", "nextcloud:primary"}:
                external_source_id = canonical_nextcloud_connector_id(
                    self.db, str(external_source_id), allow_unresolved_legacy=True
                )
            existing_binding = (
                self.db.query(SourceProfile)
                .filter(SourceProfile.external_source_id == external_source_id)
                .one_or_none()
            )
            if existing_binding is not None:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    {
                        "code": "SOURCE_CONNECTOR_ALREADY_BOUND",
                        "message": (
                            "This connector already belongs to a Source. "
                            "Create a fresh connector for a replacement Source."
                        ),
                        "lifecycle_status": existing_binding.status,
                    },
                )
        self._validate_worksheet(worksheet_mode, worksheet_name, data_start_row)
        source = SourceProfile(
            id=_id(),
            name=_clean_name(name, "FlowHub Source"),
            source_kind=source_kind,
            external_source_id=external_source_id,
            worksheet_mode=worksheet_mode,
            worksheet_name=worksheet_name,
            data_start_row=data_start_row,
            owner_user_id=user.id,
            status="active",
            version=1,
        )
        self.db.add(source)
        self.db.flush()
        if source_kind in {"flowhub_sheet", "imported_sheet"}:
            self.db.add(
                FlowHubSheet(
                    id=_id(),
                    source_id=source.id,
                    name=source.name,
                    owner_user_id=user.id,
                    current_version=0,
                )
            )
        if currency or currency_unit:
            if not currency or not currency_unit:
                raise _unprocessable(
                    "SOURCE_CURRENCY_INCOMPLETE",
                    "Both Source currency and currency unit are required.",
                )
            PricingMatrixService(self.db).declare_unit(
                scope="source",
                scope_reference=source.id,
                currency=currency,
                unit=currency_unit,
                user=user,
                connector_config_version="source-profile-v1",
                commit=False,
            )
        self.db.commit()
        return self._source_shape(source)

    def get_source(self, source_id: str, user: FlowHubUser) -> dict[str, Any]:
        source = self._owned_source(source_id, user)
        result = self._source_shape(source)
        mapping = self.sources.latest_mapping(source.id)
        result["mapping"] = self._mapping_shape(mapping) if mapping else None
        result["legacyMapping"] = self._legacy_mapping_shape(source) if mapping is None else None
        sheet = self.sheets.for_source(source.id)
        result["sheetId"] = sheet.id if sheet else None
        if mapping is not None:
            _, _, rules = self._worksheet_rule_configs(mapping)
            result["configuredWorksheets"] = sorted(
                item["worksheetName"]
                for item in rules
                if item["worksheetName"] != "*"
            )
        else:
            result["configuredWorksheets"] = []
        if source.source_kind == "external" and self._is_nextcloud_connector(source.external_source_id):
            reader = SpreadsheetSourceReadService(
                self.db, connector_id=str(source.external_source_id)
            )
            result["readQuota"] = reader.read_quota_contract(source_id=source.id)
            result["worksheetDiscovery"] = reader.worksheet_discovery_state(source_id=source.id)
            result["discoveryQuota"] = reader.discovery_quota_contract(source_id=source.id)
        return result

    def source_lifecycle(self, source_id: str, user: FlowHubUser) -> dict[str, Any]:
        """Describe the safe lifecycle action without mutating the Source."""
        source = self._owned_source(source_id, user)
        return self._source_lifecycle_impact(source)

    def _validate_lifecycle_confirmation(
        self,
        *,
        source: SourceProfile,
        expected_source_version: int,
        confirmation_name: str,
    ) -> None:
        if source.version != expected_source_version:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_VERSION_CONFLICT",
                    "message": "Source configuration changed before confirmation.",
                },
            )
        if confirmation_name.strip() != source.name:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_CONFIRMATION_MISMATCH",
                    "message": "Enter the current Source name to confirm this action.",
                },
            )
        if source.status == "deleted":
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_ALREADY_DELETED",
                    "message": "This Source has already been permanently deleted.",
                },
            )

    @staticmethod
    def _raise_if_lifecycle_blocked(impact: dict[str, Any], *, operation: str) -> None:
        if not impact["blockers"]:
            return
        active_acquisition_count = impact["blockers"].get("activeAcquisitionRuns", 0)
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            {
                "code": (
                    "SOURCE_ACTIVE_ACQUISITION"
                    if active_acquisition_count
                    else "SOURCE_ACTIVE_WORKSPACE"
                ),
                "message": (
                    "Wait for active Source reads to finish before this lifecycle operation."
                    if active_acquisition_count
                    else (
                        "Rebind or archive the active Workspace before permanently deleting this Source."
                        if operation == "delete"
                        else "Archive the active Workspace before archiving this Source."
                    )
                ),
                "details": impact,
            },
        )

    def archive_source(
        self,
        *,
        source_id: str,
        expected_source_version: int,
        confirmation_name: str,
        user: FlowHubUser,
    ) -> dict[str, Any]:
        """Archive only; this operation never permanently deletes a Source."""
        source = self._owned_source(source_id, user, lock=True)
        self._validate_lifecycle_confirmation(
            source=source,
            expected_source_version=expected_source_version,
            confirmation_name=confirmation_name,
        )
        if source.status == "archived":
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "SOURCE_ALREADY_ARCHIVED", "message": "This Source is already archived."},
            )
        if source.status != "active":
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "SOURCE_DISABLED", "message": "Only an active Source can be archived."},
            )
        impact = self._source_lifecycle_impact(source)
        self._raise_if_lifecycle_blocked(impact, operation="archive")
        connector_disabled = self._disable_external_connector(source)
        source_metadata = {
            "sourceId": source.id,
            "sourceName": source.name,
            "sourceKind": source.source_kind,
            "sourceVersion": source.version,
            "protectedHistory": impact["protectedHistory"],
            "connectorDisabled": connector_disabled,
        }
        source.status = "archived"
        source.archived_at = utcnow()
        source.version += 1
        source.updated_at = source.archived_at
        self._append_source_lifecycle_audit(
            event_type="source_archived",
            user=user,
            reason="explicit_archive_requested",
            metadata=source_metadata,
        )
        self.db.commit()
        return {
            "outcome": "archived",
            "sourceId": source.id,
            "sourceName": source.name,
            "source": self._source_shape(source),
            "impact": impact,
        }

    def permanently_delete_source(
        self,
        *,
        source_id: str,
        expected_source_version: int,
        confirmation_name: str,
        confirm_permanent_delete: bool,
        confirm_history_policy: bool,
        user: FlowHubUser,
    ) -> dict[str, Any]:
        """Permanently remove operational Source state in one transaction.

        Immutable acquisition, identity, and Workspace provenance is retained
        behind a hidden deleted tombstone because those records are append-only
        and use RESTRICT foreign keys. Sources with no such history are removed
        physically. This method never silently falls back to Archive.
        """
        source = self._owned_source(source_id, user, lock=True)
        self._validate_lifecycle_confirmation(
            source=source,
            expected_source_version=expected_source_version,
            confirmation_name=confirmation_name,
        )
        if not confirm_permanent_delete or not confirm_history_policy:
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                {
                    "code": "SOURCE_DESTRUCTIVE_CONFIRMATION_REQUIRED",
                    "message": (
                        "Permanent deletion requires explicit confirmation that "
                        "operational data is removed and immutable history follows the stated policy."
                    ),
                },
            )
        impact = self._source_lifecycle_impact(source)
        self._raise_if_lifecycle_blocked(impact, operation="delete")
        source_id_value = source.id
        source_name = source.name
        connector_id = source.external_source_id
        metadata = {
            "sourceId": source_id_value,
            "sourceName": source_name,
            "sourceKind": source.source_kind,
            "sourceVersion": source.version,
            "protectedHistory": impact["protectedHistory"],
            "historyPolicy": "immutable_history_tombstone",
        }
        self.db.query(WorkspaceSourceBinding).filter(
            WorkspaceSourceBinding.source_id == source.id
        ).delete(synchronize_session=False)
        self._delete_operational_source_state(source, connector_id=connector_id)
        if impact["protectedHistory"]:
            source.status = "deleted"
            source.deleted_at = utcnow()
            source.archived_at = None
            source.external_source_id = None
            source.version += 1
            source.updated_at = source.deleted_at
            self._append_source_lifecycle_audit(
                event_type="source_deleted",
                user=user,
                reason="permanent_delete_tombstone_required",
                metadata={**metadata, "tombstone": True},
            )
            self.db.commit()
            return {
                "outcome": "deleted",
                "sourceId": source_id_value,
                "sourceName": source_name,
                "source": None,
                "tombstone": True,
                "impact": impact,
            }
        sheet = self.sheets.for_source(source.id)
        if sheet is not None:
            self.db.delete(sheet)
            self.db.flush()
        self.db.delete(source)
        self._append_source_lifecycle_audit(
            event_type="source_deleted",
            user=user,
            reason="permanent_delete_no_history",
            metadata={**metadata, "tombstone": False},
        )
        self.db.commit()
        return {
            "outcome": "deleted",
            "sourceId": source_id_value,
            "sourceName": source_name,
            "source": None,
            "tombstone": False,
            "impact": impact,
        }

    def _delete_operational_source_state(
        self, source: SourceProfile, *, connector_id: str | None
    ) -> None:
        """Delete classified operational projections, never immutable evidence."""
        self.db.query(DlWorkspacePreview).filter(
            DlWorkspacePreview.source_id == source.id
        ).delete(synchronize_session=False)
        for model in (
            DlProductCache,
            DlInventoryCache,
            DlConnectorHealth,
            DlConnectorTelemetry,
            DlRefreshJob,
            DlInvalidationEvent,
            DlSourceReadLock,
            DlSourceReadReservation,
            DlSourceDiscoveryLock,
            DlSourceDiscoveryReservation,
            DlWorksheetDiscoveryCache,
        ):
            if hasattr(model, "source_id"):
                self.db.query(model).filter(model.source_id == source.id).delete(
                    synchronize_session=False
                )
            if connector_id and hasattr(model, "connector_id"):
                self.db.query(model).filter(model.connector_id == connector_id).delete(
                    synchronize_session=False
                )
        if not connector_id:
            return
        for model in (
            IntegrationConnectorDiagnostic,
            IntegrationConnectorEvent,
            IntegrationConnectorHealthSnapshot,
            IntegrationConnectorTelemetry,
            IntegrationWebhookEvent,
            IntegrationPollingPolicy,
            IntegrationConnectorSetting,
        ):
            self.db.query(model).filter(model.connector_id == connector_id).delete(
                synchronize_session=False
            )
        self.db.query(FlowHubAppConfig).filter(
            FlowHubAppConfig.key.like(f"connector_secret.{connector_id}.%")
        ).delete(synchronize_session=False)
        connector = self.db.get(IntegrationConnectorInstance, connector_id)
        if connector is not None:
            self.db.delete(connector)

    def lock_source_for_workspace(
        self,
        source_id: str,
        user: FlowHubUser,
        *,
        expected_source_version: int,
    ) -> SourceProfile:
        """Fence Source lifecycle changes while a Workspace Snapshot is committed."""
        source = self._owned_source(source_id, user, require_active=True, lock=True)
        if source.version != expected_source_version:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_VERSION_CONFLICT",
                    "message": "Source configuration changed during Workspace preparation.",
                },
            )
        return source

    def stage_source_product_identity_bindings(
        self,
        *,
        source_id: str,
        mapping_revision_id: str,
        source_revision_kind: str,
        source_revision_id: str,
        dataset_id: str | None,
        sheet_revision_id: str | None,
        proposals: list[dict[str, Any]],
        user: FlowHubUser,
    ) -> list[dict[str, Any]]:
        """Stage conflict-free bindings inside the Workspace commit transaction."""

        self._owned_source(source_id, user, require_active=True, lock=True)
        mapping = self.db.get(SourceMappingRevision, mapping_revision_id)
        if mapping is None or mapping.source_id != source_id:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_MAPPING_CHANGED",
                    "message": "Source Mapping changed before identity binding.",
                },
            )
        if source_revision_kind == "source_observation":
            dataset = (
                self.db.get(SourceObservationDataset, dataset_id)
                if dataset_id
                else None
            )
            if (
                dataset is None
                or dataset.source_id != source_id
                or dataset.observation_id != source_revision_id
                or sheet_revision_id is not None
            ):
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    {
                        "code": "SOURCE_IDENTITY_EVIDENCE_CHANGED",
                        "message": "Source identity evidence changed before binding.",
                    },
                )
        elif source_revision_kind == "flowhub_sheet_revision":
            revision = (
                self.db.get(SheetRevision, sheet_revision_id)
                if sheet_revision_id
                else None
            )
            sheet = self.db.get(FlowHubSheet, revision.sheet_id) if revision else None
            if (
                revision is None
                or sheet is None
                or sheet.source_id != source_id
                or revision.id != source_revision_id
                or dataset_id is not None
            ):
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    {
                        "code": "SOURCE_IDENTITY_EVIDENCE_CHANGED",
                        "message": "Source identity evidence changed before binding.",
                    },
                )
        else:
            raise RuntimeError("Unsupported Source identity evidence kind.")
        by_hash: dict[str, dict[str, Any]] = {}
        for proposal in proposals:
            normalized_key = _normalize_source_product_key(
                proposal.get("normalizedSourceKey")
            )
            source_key_hash = str(proposal.get("sourceKeyHash") or "")
            if (
                not normalized_key
                or proposal.get("normalizationVersion")
                != SOURCE_KEY_NORMALIZATION_VERSION
                or source_key_hash != _source_product_key_hash(normalized_key)
            ):
                raise RuntimeError("Invalid Source Product identity binding proposal.")
            canonical_product_id = str(proposal.get("canonicalProductId") or "")
            listing_evidence: list[dict[str, Any]] = []
            for raw_listing in list(proposal.get("listingEvidence") or []):
                listing_id = str(raw_listing.get("listingId") or "")
                listing_canonical_id = str(
                    raw_listing.get("canonicalProductId") or ""
                )
                mapping_version = int(raw_listing.get("mappingVersion") or 0)
                if (
                    not listing_id
                    or not canonical_product_id
                    or listing_canonical_id != canonical_product_id
                    or mapping_version < 1
                ):
                    raise RuntimeError(
                        "Invalid Listing evidence in Source identity proposal."
                    )
                listing_evidence.append(
                    {
                        "listingId": listing_id,
                        "canonicalProductId": listing_canonical_id,
                        "mappingVersion": mapping_version,
                    }
                )
            if not listing_evidence:
                raise RuntimeError(
                    "Source identity binding proposal requires Listing evidence."
                )
            listing_evidence = sorted(
                {item["listingId"]: item for item in listing_evidence}.values(),
                key=lambda item: item["listingId"],
            )
            prior = by_hash.get(source_key_hash)
            if prior is not None and prior["canonicalProductId"] != canonical_product_id:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    {
                        "code": "SOURCE_PRODUCT_IDENTITY_CONFLICT",
                        "message": "Source Product Key resolves to different Canonical Products.",
                    },
                )
            by_hash[source_key_hash] = {
                "normalizedSourceKey": normalized_key,
                "canonicalProductId": canonical_product_id,
                "listingEvidence": listing_evidence,
            }
        expected_listings: dict[str, dict[str, Any]] = {}
        for proposal in by_hash.values():
            for listing_evidence in proposal["listingEvidence"]:
                listing_id = listing_evidence["listingId"]
                prior = expected_listings.get(listing_id)
                if prior is not None and prior != listing_evidence:
                    raise HTTPException(
                        status.HTTP_409_CONFLICT,
                        {
                            "code": "SOURCE_LISTING_IDENTITY_CHANGED",
                            "message": "A Listing changed during Source identity resolution.",
                        },
                    )
                expected_listings[listing_id] = listing_evidence
        current_listings = {
            item.id: item
            for item in (
                self.db.query(Listing)
                .filter(Listing.id.in_(set(expected_listings)))
                .order_by(Listing.id)
                .populate_existing()
                .with_for_update()
                .all()
                if expected_listings
                else []
            )
        }
        if any(
            listing_id not in current_listings
            or current_listings[listing_id].canonical_product_id
            != evidence["canonicalProductId"]
            or current_listings[listing_id].mapping_version
            != evidence["mappingVersion"]
            or not current_listings[listing_id].enabled
            or current_listings[listing_id].mapping_state != "resolved"
            for listing_id, evidence in expected_listings.items()
        ):
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_LISTING_IDENTITY_CHANGED",
                    "message": "A Listing changed during Source identity resolution. Try again.",
                },
            )
        existing = {
            item.source_key_hash: item
            for item in (
                self.db.query(SourceProductIdentity)
                .filter(
                    SourceProductIdentity.source_id == source_id,
                    SourceProductIdentity.normalization_version
                    == SOURCE_KEY_NORMALIZATION_VERSION,
                    SourceProductIdentity.source_key_hash.in_(set(by_hash)),
                )
                .all()
                if by_hash
                else []
            )
        }
        accepted_bindings: list[SourceProductIdentity] = []
        for source_key_hash, proposal in by_hash.items():
            binding = existing.get(source_key_hash)
            if binding is not None:
                if binding.canonical_product_id != proposal["canonicalProductId"]:
                    raise HTTPException(
                        status.HTTP_409_CONFLICT,
                        {
                            "code": "SOURCE_PRODUCT_IDENTITY_CONFLICT",
                            "message": "Source Product Key is already bound to another Canonical Product.",
                        },
                    )
                accepted_bindings.append(binding)
                continue
            binding = SourceProductIdentity(
                id=_id(),
                source_id=source_id,
                source_key_hash=source_key_hash,
                normalized_source_key=proposal["normalizedSourceKey"],
                normalization_version=SOURCE_KEY_NORMALIZATION_VERSION,
                canonical_product_id=proposal["canonicalProductId"],
                first_mapping_revision_id=mapping.id,
                first_source_revision_kind=source_revision_kind,
                first_source_revision_id=source_revision_id,
                first_dataset_id=dataset_id,
                first_sheet_revision_id=sheet_revision_id,
                identity_authority_json=self._identity_authority_shape(mapping),
            )
            self.db.add(binding)
            accepted_bindings.append(binding)
        self.db.flush()
        return [
            {
                "id": binding.id,
                "sourceKeyHash": binding.source_key_hash,
                "normalizationVersion": binding.normalization_version,
                "canonicalProductId": binding.canonical_product_id,
                "firstSourceRevisionKind": binding.first_source_revision_kind,
                "firstSourceRevisionId": binding.first_source_revision_id,
                "datasetId": binding.first_dataset_id,
                "sheetRevisionId": binding.first_sheet_revision_id,
            }
            for binding in sorted(
                accepted_bindings, key=lambda item: item.source_key_hash
            )
        ]

    async def list_source_worksheets(
        self, source_id: str, user: FlowHubUser, *, refresh: bool = False
    ) -> dict[str, Any]:
        """Return local metadata or explicitly refresh bounded remote metadata."""
        source = self._owned_source(source_id, user, require_active=True)
        reader = SpreadsheetSourceReadService(
            self.db,
            connector_id=source.external_source_id or LEGACY_EXTERNAL_SOURCE_ID,
        )
        if source.external_source_id:
            connector = self.db.get(IntegrationConnectorInstance, source.external_source_id)
            if connector is not None and not connector.enabled:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    {"code": "SOURCE_DISABLED", "message": "Nextcloud Source is disabled. Enable it before reading source data."},
                )
        sheet = self.sheets.for_source(source.id)
        if sheet is not None:
            revision = self.sheets.latest_revision(sheet.id)
            return {
                "sourceId": source.id,
                "items": [
                    {
                        "name": "Sheet1",
                        "rowCount": revision.row_count if revision else 0,
                    }
                ],
                "sourceRevisionId": revision.id if revision else None,
                "readQuota": _worksheet_read_quota(
                    reader.read_quota_contract(source_id=source.id)
                ),
                "discoveryQuota": _worksheet_discovery_quota(reader.discovery_quota_contract(source_id=source.id)),
                "worksheetDiscovery": {
                    "requiresRemoteRead": False,
                    "metadataSource": "flowhub_sheet",
                    "remoteReadUsed": False,
                    "snapshotId": None,
                    "snapshotVersion": None,
                    "snapshotAt": None,
                },
            }
        if refresh:
            if not self._is_nextcloud_connector(source.external_source_id):
                raise _unprocessable("WORKSHEET_DISCOVERY_UNAVAILABLE", "Remote discovery is unavailable for this Source.")
            refreshed = await reader.refresh_worksheet_discovery(source_profile_id=source.id, user_id=user.id)
            return {
                "sourceId": source.id,
                "items": refreshed["worksheets"],
                "sourceRevisionId": None,
                "readQuota": _worksheet_read_quota(reader.read_quota_contract(source_id=source.id)),
                "discoveryQuota": _worksheet_discovery_quota(refreshed["quota"]),
                "worksheetDiscovery": {
                    "requiresRemoteRead": False, "metadataSource": "remote_metadata", "remoteReadUsed": True,
                    "snapshotId": None, "snapshotVersion": None, "snapshotAt": None,
                    "discoveredAt": refreshed["discovered_at"],
                },
            }
        discovery = (
            reader.worksheet_discovery_state(source_id=source.id)
            if self._is_nextcloud_connector(source.external_source_id)
            else None
        )
        if discovery is not None and discovery["metadata_source"] in {"snapshot", "discovery_cache"}:
            return {
                "sourceId": source.id,
                "items": discovery["worksheets"],
                "sourceRevisionId": f"external:{discovery['snapshot_id']}:{discovery['snapshot_version']}" if discovery["snapshot_id"] is not None else None,
                "readQuota": _worksheet_read_quota(
                    reader.read_quota_contract(source_id=source.id)
                ),
                "discoveryQuota": _worksheet_discovery_quota(reader.discovery_quota_contract(source_id=source.id)),
                "worksheetDiscovery": {
                    "requiresRemoteRead": False,
                    "metadataSource": discovery["metadata_source"],
                    "remoteReadUsed": False,
                    "snapshotId": discovery["snapshot_id"],
                    "snapshotVersion": discovery["snapshot_version"],
                    "snapshotAt": discovery["snapshot_at"],
                    **({"discoveredAt": discovery["discovered_at"]} if discovery.get("discovered_at") else {}),
                },
            }
        return {
            "sourceId": source.id,
            "items": [],
            "sourceRevisionId": None,
            "readQuota": _worksheet_read_quota(
                reader.read_quota_contract(source_id=source.id)
            ),
            "discoveryQuota": _worksheet_discovery_quota(reader.discovery_quota_contract(source_id=source.id)),
            "worksheetDiscovery": {
                "requiresRemoteRead": True,
                "metadataSource": "unavailable",
                "remoteReadUsed": False,
                "snapshotId": None,
                "snapshotVersion": None,
                "snapshotAt": None,
                "discoveredAt": None,
            },
        }

    def save_mapping(
        self,
        *,
        source_id: str,
        expected_source_version: int,
        worksheet_mode: str,
        worksheet_name: str | None,
        data_start_row: int,
        source_fields: list[dict[str, Any]],
        channel_mappings: list[dict[str, Any]],
        value_policy: dict[str, Any],
        worksheet_rule_mode: str = "shared",
        selected_worksheet_names: list[str] | None = None,
        duplicate_product_policy: str = "block",
        worksheet_rules: list[dict[str, Any]] | None = None,
        identity_policy_version: int = 2,
        identity_authority: dict[str, Any] | None = None,
        user: FlowHubUser,
        _commit: bool = True,
    ) -> dict[str, Any]:
        self._ensure_channels()
        source = self._owned_source(source_id, user, require_active=True, lock=True)
        if source.version != expected_source_version:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "SOURCE_VERSION_CONFLICT", "message": "Source configuration changed."},
            )
        if worksheet_rule_mode not in {"shared", "per_worksheet"}:
            raise _unprocessable(
                "WORKSHEET_RULE_MODE_INVALID",
                "Use shared or per-worksheet Source rules.",
            )
        if identity_policy_version != 2:
            raise _unprocessable(
                "SOURCE_IDENTITY_POLICY_UPGRADE_REQUIRED",
                "New Mapping revisions require Source Product Key identity policy v2.",
            )
        normalized_identity_authority = self._normalize_identity_authority(
            identity_authority,
            required=identity_policy_version >= 2,
        )
        normalized_selected_worksheets = self._normalize_selected_worksheet_names(
            selected_worksheet_names or []
        )
        if worksheet_mode == "all":
            if normalized_selected_worksheets:
                raise _unprocessable(
                    "WORKSHEET_SELECTION_INVALID",
                    "Selected worksheet names cannot be combined with all-worksheets participation.",
                )
            self._validate_worksheet("all", None, data_start_row)
            effective_worksheet_name = None
        else:
            if not normalized_selected_worksheets and str(worksheet_name or "").strip():
                normalized_selected_worksheets = [str(worksheet_name).strip()]
            if not normalized_selected_worksheets:
                raise _unprocessable("WORKSHEET_REQUIRED", "Select at least one participating worksheet.")
            if worksheet_name and str(worksheet_name).strip() not in normalized_selected_worksheets:
                raise _unprocessable(
                    "WORKSHEET_SELECTION_INVALID",
                    "The compatibility worksheet must participate in the Source.",
                )
            for selected_name in normalized_selected_worksheets:
                self._validate_worksheet("selected", selected_name, data_start_row)
            effective_worksheet_name = normalized_selected_worksheets[0] if len(normalized_selected_worksheets) == 1 else None
        if worksheet_rule_mode == "per_worksheet" and worksheet_mode == "selected" and len(normalized_selected_worksheets) < 2:
            raise _unprocessable(
                "WORKSHEET_STRATEGY_INVALID",
                "Separate worksheet mappings require more than one participating worksheet.",
            )
        if duplicate_product_policy not in {"block", "last_sheet_wins"}:
            raise _unprocessable(
                "DUPLICATE_PRODUCT_POLICY_INVALID",
                "Choose block or last-sheet-wins duplicate handling.",
            )
        normalized_source_fields = self._normalize_field_mappings(
            source_fields,
            SOURCE_FIELDS,
            required_fields=({"name", "source_key"} if identity_policy_version >= 2 else {"name"}) if worksheet_rule_mode == "shared" else set(),
        )
        normalized_channels = self._normalize_channel_mappings(
            channel_mappings,
            require_enabled=worksheet_rule_mode == "shared",
        )
        normalized_policy = self._normalize_value_policy(value_policy)
        normalized_worksheet_rules: list[dict[str, Any]]
        if worksheet_rule_mode == "shared":
            shared_rule_names = normalized_selected_worksheets or ["*"]
            normalized_worksheet_rules = [
                {
                    "worksheetName": shared_rule_name,
                    "enabled": True,
                    "dataStartRow": data_start_row,
                    "sourceFields": normalized_source_fields,
                    "channels": normalized_channels,
                    "valuePolicy": normalized_policy,
                }
                for shared_rule_name in shared_rule_names
            ]
        else:
            normalized_worksheet_rules = self._normalize_worksheet_rules(
                worksheet_rules or [], require_source_key=identity_policy_version >= 2
            )
            enabled_rule_names = {
                rule["worksheetName"] for rule in normalized_worksheet_rules if rule["enabled"]
            }
            if worksheet_mode == "selected" and enabled_rule_names != set(normalized_selected_worksheets):
                raise _unprocessable(
                    "WORKSHEET_SCOPE_RULE_MISMATCH",
                    "Enabled worksheet rules must exactly match the participating worksheets.",
                    {"participating": normalized_selected_worksheets, "enabledRules": sorted(enabled_rule_names)},
                )
            if worksheet_mode == "all" and len(enabled_rule_names) < 2:
                raise _unprocessable(
                    "WORKSHEET_STRATEGY_INVALID",
                    "Separate worksheet mappings require more than one participating worksheet.",
                )
        if source.source_kind == "external":
            external_references = [
                *[
                    field
                    for rule in normalized_worksheet_rules
                    for field in rule["sourceFields"]
                ],
                *[
                    field
                    for rule in normalized_worksheet_rules
                    for channel in rule["channels"]
                    for field in channel["fields"]
                ],
            ]
            if any(item["referenceType"] == "column_id" for item in external_references):
                raise _unprocessable(
                    "COLUMN_REFERENCE_UNAVAILABLE",
                    "Internal FlowHub column IDs cannot be used for an external Source.",
                )
        latest = self.sources.latest_mapping(source.id)
        version = (latest.version if latest else 0) + 1
        document = {
            "sourceId": source.id,
            "version": version,
            "worksheetMode": worksheet_mode,
            "worksheetName": effective_worksheet_name,
            "selectedWorksheetNames": normalized_selected_worksheets,
            "dataStartRow": data_start_row,
            "sourceFields": normalized_source_fields,
            "channels": normalized_channels,
            "valuePolicy": normalized_policy,
            "worksheetRuleMode": worksheet_rule_mode,
            "duplicateProductPolicy": duplicate_product_policy,
            "worksheetRules": normalized_worksheet_rules,
            "identityPolicyVersion": identity_policy_version,
            "identityAuthority": normalized_identity_authority,
        }
        revision = SourceMappingRevision(
            id=_id(),
            source_id=source.id,
            version=version,
            checksum=checksum(document),
            worksheet_mode=worksheet_mode,
            worksheet_name=effective_worksheet_name,
            data_start_row=data_start_row,
            value_policy_json=normalized_policy,
            identity_authority_json=normalized_identity_authority,
            identity_policy_version=identity_policy_version,
            created_by_user_id=user.id,
        )
        self.db.add(revision)
        self.db.flush()
        for item in normalized_source_fields:
            self.db.add(
                SourceFieldMapping(
                    id=_id(),
                    mapping_revision_id=revision.id,
                    field=item["field"],
                    reference_type=item["referenceType"],
                    reference_value=item["referenceValue"],
                    required=bool(item["required"]),
                )
            )
        for channel in normalized_channels:
            channel_mapping = SourceChannelMapping(
                id=_id(),
                mapping_revision_id=revision.id,
                channel_id=channel["channelId"],
                worksheet_name=channel["worksheetName"],
                enabled=bool(channel["enabled"]),
            )
            self.db.add(channel_mapping)
            self.db.flush()
            for field in channel["fields"]:
                self.db.add(
                    SourceChannelFieldMapping(
                        id=_id(),
                        channel_mapping_id=channel_mapping.id,
                        field=field["field"],
                        reference_type=field["referenceType"],
                        reference_value=field["referenceValue"],
                    )
                )
        self._persist_worksheet_rule_set(
            revision=revision,
            mode=worksheet_rule_mode,
            duplicate_product_policy=duplicate_product_policy,
            rules=normalized_worksheet_rules,
        )
        source.version += 1
        source.worksheet_mode = worksheet_mode
        source.worksheet_name = effective_worksheet_name
        source.data_start_row = data_start_row
        source.updated_at = utcnow()
        if _commit:
            self._assess_mapping_from_latest_local_data(revision, persist=True)
            self._invalidate_source_reviews(source.id)
            self.db.commit()
        else:
            self.db.flush()
        shape = self._mapping_shape(revision)
        if shape is None:
            raise RuntimeError("Mapping revision persistence failed")
        return shape

    async def preview_unsaved_mapping(
        self,
        *,
        source_id: str,
        expected_source_version: int,
        worksheet_mode: str,
        worksheet_name: str | None,
        data_start_row: int,
        source_fields: list[dict[str, Any]],
        channel_mappings: list[dict[str, Any]],
        value_policy: dict[str, Any],
        worksheet_rule_mode: str = "shared",
        selected_worksheet_names: list[str] | None = None,
        duplicate_product_policy: str = "block",
        worksheet_rules: list[dict[str, Any]] | None = None,
        identity_policy_version: int = 2,
        identity_authority: dict[str, Any] | None = None,
        user: FlowHubUser,
        page: int = 1,
        page_size: int = 200,
    ) -> dict[str, Any]:
        """Preview a draft Mapping from local evidence only.

        The candidate aggregate exists only inside the savepoint. This method
        never calls a provider and never reserves Source acquisition quota.
        """

        source = self._owned_source(
            source_id, user, require_active=True, lock=True
        )
        local_data = self._latest_local_validation_data(source)

        savepoint = self.db.begin_nested()
        try:
            self.save_mapping(
                source_id=source_id,
                expected_source_version=expected_source_version,
                worksheet_mode=worksheet_mode,
                worksheet_name=worksheet_name,
                data_start_row=data_start_row,
                source_fields=source_fields,
                channel_mappings=channel_mappings,
                value_policy=value_policy,
                worksheet_rule_mode=worksheet_rule_mode,
                selected_worksheet_names=selected_worksheet_names,
                duplicate_product_policy=duplicate_product_policy,
                worksheet_rules=worksheet_rules,
                identity_policy_version=identity_policy_version,
                identity_authority=identity_authority,
                user=user,
                _commit=False,
            )
            mapping = self.sources.latest_mapping(source.id)
            if mapping is None:
                raise RuntimeError("Draft mapping preview persistence failed")
            if local_data is None:
                records = []
            elif local_data["kind"] == "source_observation":
                records = self._mapped_external_records(local_data["worksheets"], mapping)
            else:
                records = self._mapped_sheet_records(local_data["sheetRevision"], mapping)
            result = self._shape_source_preview(
                records,
                mapping,
                page=page,
                page_size=page_size,
                sheet_revision_id=(
                    str(local_data["sourceRevisionId"]) if local_data is not None else None
                ),
                mapping_revision_id=None,
                validation_source=(local_data["evidence"] if local_data is not None else None),
            )
        finally:
            savepoint.rollback()
            self.db.expire_all()
        return result

    async def source_preview(
        self, source_id: str, user: FlowHubUser, *, page: int, page_size: int
    ) -> dict[str, Any]:
        source = self._owned_source(source_id, user, require_active=True)
        mapping = self.sources.latest_mapping(source.id)
        if mapping is None:
            raise _unprocessable("SOURCE_MAPPING_REQUIRED", "Configure Source mappings first.")
        local_data = self._latest_local_validation_data(source)
        if local_data is None:
            records = []
            revision_id = None
        elif local_data["kind"] == "source_observation":
            records = self._mapped_external_records(local_data["worksheets"], mapping)
            revision_id = str(local_data["sourceRevisionId"])
        else:
            records = self._mapped_sheet_records(local_data["sheetRevision"], mapping)
            revision_id = str(local_data["sourceRevisionId"])
        return self._shape_source_preview(
            records,
            mapping,
            page=page,
            page_size=page_size,
            sheet_revision_id=revision_id,
            mapping_revision_id=mapping.id,
            validation_source=(local_data["evidence"] if local_data is not None else None),
        )

    def _shape_source_preview(
        self,
        records: list[dict[str, Any]],
        mapping: SourceMappingRevision,
        *,
        page: int,
        page_size: int,
        sheet_revision_id: str | None,
        mapping_revision_id: str | None,
        validation_source: dict[str, Any] | None,
    ) -> dict[str, Any]:
        start = (max(page, 1) - 1) * page_size
        page_records: list[dict[str, Any]] = []
        for record in records[start : start + min(max(page_size, 1), 500)]:
            shaped = dict(record)
            shaped.pop("sourceKeyRequired", None)
            shaped["hasIssues"] = bool(record.get("issues"))
            shaped["ready"] = bool(record.get("recognized")) and not shaped["hasIssues"]
            page_records.append(shaped)
        return {
            "items": page_records,
            "total": len(records),
            "recognized": sum(1 for item in records if item["recognized"]),
            "ignored": sum(1 for item in records if not item["recognized"]),
            "issues": self._preview_issue_summary(records),
            "businessSummary": self._preview_business_summary(records, mapping),
            "identityValidation": self._identity_preview_summary(
                records,
                mapping=mapping,
                validation_source=validation_source,
            ),
            "sheetRevisionId": sheet_revision_id,
            "mappingRevisionId": mapping_revision_id,
        }

    def _identity_preview_summary(
        self,
        records: list[dict[str, Any]],
        *,
        mapping: SourceMappingRevision,
        validation_source: dict[str, Any] | None,
        listing_context: dict[tuple[str, str], Listing] | None = None,
    ) -> dict[str, Any]:
        """Summarize Source-key evidence without exposing unrelated row data."""
        mapping_references = self._identity_mapping_references(mapping)
        if (
            int(mapping.identity_policy_version or 1) < 2
            or self._identity_authority_shape(mapping)["type"] == "unspecified"
        ):
            return self._pending_identity_validation(
                mapping_references,
                validation_source,
            )
        if validation_source is None:
            return self._pending_identity_validation(mapping_references)
        identity_categories = {
            "missing_source_product_key",
            "duplicate_source_product_key",
        }
        identity_records = [
            record
            for record in records
            if record.get("sourceKeyRequired")
            and (
                any(
                    value not in {None, ""}
                    for value in dict(record.get("sourceProduct") or {}).values()
                )
                or bool(record.get("channels"))
                or bool(record.get("issues"))
            )
        ]
        missing = sum(
            1
            for record in identity_records
            if any(issue.get("category") == "missing_source_product_key" for issue in record["issues"])
        )
        duplicate_records = [
            record
            for record in identity_records
            if any(
                issue.get("category") == "duplicate_source_product_key"
                for issue in record["issues"]
            )
        ]
        (
            binding_conflicts,
            binding_conflict_keys,
            binding_context_fingerprint,
        ) = self._identity_binding_conflicts(
            identity_records,
            source_id=mapping.source_id,
            listing_context=listing_context,
        )
        valid = sum(
            1
            for record in identity_records
            if _normalize_source_product_key(
                record.get("sourceProduct", {}).get("source_key")
            )
            and _normalize_source_product_key(
                record.get("sourceProduct", {}).get("source_key")
            )
            not in binding_conflict_keys
            and not any(
                issue.get("category") in identity_categories
                for issue in record["issues"]
            )
        )
        missing_rows = [
            {
                "worksheetName": str(record.get("worksheetName") or ""),
                "rowNumber": int(record.get("rowNumber") or 0),
            }
            for record in identity_records
            if any(
                issue.get("category") == "missing_source_product_key"
                for issue in record["issues"]
            )
        ][:IDENTITY_CONFLICT_GROUP_LIMIT]
        duplicate_groups_by_key: dict[str, dict[str, Any]] = {}
        for record in duplicate_records:
            display_value = str(
                record.get("sourceProduct", {}).get("source_key") or ""
            ).strip()
            normalized_key = _normalize_source_product_key(display_value)
            group = duplicate_groups_by_key.setdefault(
                normalized_key,
                {"keyValue": display_value[:240], "rows": []},
            )
            group["rows"].append(
                {
                    "worksheetName": str(record.get("worksheetName") or ""),
                    "rowNumber": int(record.get("rowNumber") or 0),
                }
            )
        duplicate_groups = [
            {
                "keyValue": group["keyValue"],
                "rows": group["rows"][:IDENTITY_CONFLICT_ROWS_PER_GROUP_LIMIT],
            }
            for _, group in sorted(duplicate_groups_by_key.items())
        ][:IDENTITY_CONFLICT_GROUP_LIMIT]
        duplicate_rows = len(duplicate_records)
        return {
            "status": (
                "blocked"
                if missing or duplicate_rows or binding_conflicts
                else "pass"
            ),
            "participatingRowCount": len(identity_records),
            "validKeyCount": valid,
            "missingKeyCount": missing,
            "duplicateKeyCount": len(duplicate_groups_by_key),
            "duplicateRowCount": duplicate_rows,
            "bindingConflictCount": len(binding_conflict_keys),
            "bindingContextFingerprint": binding_context_fingerprint,
            "missingRows": missing_rows,
            "duplicateGroups": duplicate_groups,
            "bindingConflicts": binding_conflicts,
            "mappingReferences": mapping_references,
            "evidence": validation_source,
        }

    def _identity_binding_conflicts(
        self,
        records: list[dict[str, Any]],
        *,
        source_id: str,
        listing_context: dict[tuple[str, str], Listing] | None = None,
    ) -> tuple[list[dict[str, Any]], set[str], str]:
        """Return conflicts plus the exact local Listing/binding evidence context."""

        listing_identities = {
            (
                str(channel.get("channelId") or ""),
                str(channel.get("fields", {}).get("external_id") or "").strip(),
            )
            for record in records
            for channel in record.get("channels") or []
            if str(channel.get("fields", {}).get("external_id") or "").strip()
        }
        listings = listing_context
        if listings is None:
            listings = {
                (item.channel_id, item.external_primary_id): item
                for item in (
                    self.db.query(Listing)
                    .filter(
                        tuple_(Listing.channel_id, Listing.external_primary_id).in_(
                            sorted(listing_identities)
                        )
                    )
                    .all()
                    if listing_identities
                    else []
                )
            }
        records_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            normalized_key = _normalize_source_product_key(
                record.get("sourceProduct", {}).get("source_key")
            )
            if normalized_key:
                records_by_key[normalized_key].append(record)
        key_hashes = {
            _source_product_key_hash(normalized_key)
            for normalized_key in records_by_key
        }
        bindings = {
            item.source_key_hash: item
            for item in (
                self.db.query(SourceProductIdentity)
                .filter(
                    SourceProductIdentity.source_id == source_id,
                    SourceProductIdentity.normalization_version
                    == SOURCE_KEY_NORMALIZATION_VERSION,
                    SourceProductIdentity.source_key_hash.in_(key_hashes),
                )
                .all()
                if key_hashes
                else []
            )
        }
        conflicts: list[dict[str, Any]] = []
        conflict_keys: set[str] = set()
        binding_context: list[dict[str, Any]] = []
        for normalized_key, key_records in sorted(records_by_key.items()):
            canonical_ids: set[str] = set()
            channel_identities = {
                (
                    str(channel.get("channelId") or ""),
                    str(
                        channel.get("fields", {}).get("external_id") or ""
                    ).strip(),
                )
                for record in key_records
                for channel in record.get("channels") or []
                if str(
                    channel.get("fields", {}).get("external_id") or ""
                ).strip()
            }
            for record in key_records:
                for channel in record.get("channels") or []:
                    listing = listings.get(
                        (
                            str(channel.get("channelId") or ""),
                            str(
                                channel.get("fields", {}).get("external_id") or ""
                            ).strip(),
                        )
                    )
                    if (
                        listing is not None
                        and listing.enabled
                        and listing.mapping_state == "resolved"
                    ):
                        canonical_ids.add(listing.canonical_product_id)
            binding = bindings.get(_source_product_key_hash(normalized_key))
            effective_canonical_product_id = (
                binding.canonical_product_id
                if binding is not None
                else next(iter(canonical_ids))
                if len(canonical_ids) == 1
                else None
            )
            binding_context.append(
                {
                    "sourceKeyHash": _source_product_key_hash(normalized_key),
                    "effectiveBinding": (
                        {
                            "canonicalProductId": effective_canonical_product_id,
                            "normalizationVersion": SOURCE_KEY_NORMALIZATION_VERSION,
                        }
                        if effective_canonical_product_id is not None
                        else None
                    ),
                    "listings": [
                        {
                            "channelId": channel_id,
                            "externalIdentifierHash": checksum(
                                {
                                    "channelId": channel_id,
                                    "externalPrimaryId": external_id,
                                }
                            ),
                            "listingId": (
                                listings[(channel_id, external_id)].id
                                if (channel_id, external_id) in listings
                                else None
                            ),
                            "mappingVersion": (
                                listings[(channel_id, external_id)].mapping_version
                                if (channel_id, external_id) in listings
                                else None
                            ),
                            "canonicalProductId": (
                                listings[(channel_id, external_id)].canonical_product_id
                                if (channel_id, external_id) in listings
                                else None
                            ),
                            "mappingState": (
                                listings[(channel_id, external_id)].mapping_state
                                if (channel_id, external_id) in listings
                                else None
                            ),
                            "enabled": (
                                listings[(channel_id, external_id)].enabled
                                if (channel_id, external_id) in listings
                                else None
                            ),
                        }
                        for channel_id, external_id in sorted(channel_identities)
                    ],
                }
            )
            if binding is None:
                conflicting_ids = set(canonical_ids) if len(canonical_ids) > 1 else set()
            else:
                conflicting_ids = canonical_ids - {binding.canonical_product_id}
            if not conflicting_ids:
                continue
            conflict_keys.add(normalized_key)
            if len(conflicts) >= IDENTITY_CONFLICT_GROUP_LIMIT:
                continue
            conflicts.append(
                {
                    "keyValue": normalized_key[:240],
                    "rows": [
                        {
                            "worksheetName": str(record.get("worksheetName") or ""),
                            "rowNumber": int(record.get("rowNumber") or 0),
                        }
                        for record in key_records
                    ][:IDENTITY_CONFLICT_ROWS_PER_GROUP_LIMIT],
                    "boundCanonicalProductId": (
                        binding.canonical_product_id if binding is not None else None
                    ),
                    "conflictingCanonicalProductIds": sorted(conflicting_ids),
                }
            )
        return (
            conflicts,
            conflict_keys,
            checksum(
                {
                    "contract": "source-product-binding-context-v1",
                    "normalizationVersion": SOURCE_KEY_NORMALIZATION_VERSION,
                    "keys": binding_context,
                }
            ),
        )

    @staticmethod
    def _pending_identity_validation(
        mapping_references: list[dict[str, Any]],
        evidence: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        return {
            "status": "pending",
            "participatingRowCount": None,
            "validKeyCount": None,
            "missingKeyCount": None,
            "duplicateKeyCount": None,
            "duplicateRowCount": None,
            "bindingConflictCount": None,
            "bindingContextFingerprint": None,
            "missingRows": [],
            "duplicateGroups": [],
            "bindingConflicts": [],
            "mappingReferences": mapping_references,
            "evidence": evidence
            or {
                "kind": "none",
                "sourceRevisionId": None,
                "datasetId": None,
                "snapshotId": None,
                "snapshotVersion": None,
                "label": None,
                "validatedAt": None,
            },
        }

    def _latest_local_validation_data(
        self, source: SourceProfile
    ) -> dict[str, Any] | None:
        """Return replayable local rows and immutable evidence identity only."""
        sheet = self.sheets.for_source(source.id)
        if sheet is not None:
            revision = self.sheets.latest_revision(sheet.id)
            if revision is None:
                return None
            return {
                "kind": "flowhub_sheet_revision",
                "sourceRevisionId": revision.id,
                "sheetRevision": revision,
                "dataset": None,
                "evidence": {
                    "kind": "flowhub_sheet_revision",
                    "sourceRevisionId": revision.id,
                    "datasetId": None,
                    "snapshotId": None,
                    "snapshotVersion": revision.version,
                    "label": f"FlowHub Sheet revision {revision.version}",
                    "validatedAt": None,
                },
            }

        expected_resource_scope: str | None = None
        expected_binding_fingerprint: str | None = None
        if self._is_nextcloud_connector(source.external_source_id):
            reader = SpreadsheetSourceReadService(
                self.db,
                connector_id=source.external_source_id or LEGACY_EXTERNAL_SOURCE_ID,
            )
            expected_resource_scope = reader.configured_resource_scope()
            expected_binding_fingerprint = reader.configured_binding_fingerprint()
            if (
                expected_resource_scope is None
                or expected_binding_fingerprint is None
            ):
                # A dataset captured for a previous workbook binding must never
                # validate the mapping for an unbound or reconfigured Source.
                return None

        dataset_query = (
            self.db.query(SourceObservationDataset)
            .join(
                SourceObservation,
                SourceObservation.id == SourceObservationDataset.observation_id,
            )
            .filter(SourceObservationDataset.source_id == source.id)
        )
        if expected_resource_scope is not None:
            dataset_query = dataset_query.filter(
                SourceObservationDataset.resource_scope == expected_resource_scope,
                SourceObservationDataset.binding_fingerprint
                == expected_binding_fingerprint,
            )
        dataset = (
            dataset_query
            .order_by(
                SourceObservation.observation_version.desc(),
                SourceObservationDataset.created_at.desc(),
            )
            .first()
        )
        if dataset is None:
            return None
        worksheet_rows = (
            self.db.query(SourceObservationWorksheetDataset)
            .filter(SourceObservationWorksheetDataset.dataset_id == dataset.id)
            .order_by(SourceObservationWorksheetDataset.worksheet_order)
            .all()
        )
        worksheets = {
            item.worksheet_name: list(item.rows_json or []) for item in worksheet_rows
        }
        return {
            "kind": "source_observation",
            "sourceRevisionId": dataset.observation_id,
            "sheetRevision": None,
            "dataset": dataset,
            "worksheets": worksheets,
            "evidence": {
                "kind": "source_observation",
                "sourceRevisionId": dataset.observation_id,
                "datasetId": dataset.id,
                "snapshotId": dataset.source_snapshot_id,
                "snapshotVersion": dataset.source_snapshot_version,
                "label": (
                    f"Snapshot #{dataset.source_snapshot_id}"
                    if dataset.source_snapshot_id is not None
                    else "Local Source observation"
                ),
                "validatedAt": None,
            },
        }

    def _identity_mapping_references(
        self, mapping: SourceMappingRevision
    ) -> list[dict[str, Any]]:
        _, _, rules = self._worksheet_rule_configs(mapping)
        references: list[dict[str, Any]] = []
        for rule in rules:
            if not bool(rule.get("enabled", True)):
                continue
            for raw_field in rule.get("sourceFields") or []:
                field = (
                    raw_field.get("field")
                    if isinstance(raw_field, dict)
                    else raw_field.field
                )
                if field != "source_key":
                    continue
                reference_type = (
                    raw_field.get("referenceType")
                    or raw_field.get("reference_type")
                    if isinstance(raw_field, dict)
                    else raw_field.reference_type
                )
                reference_value = (
                    raw_field.get("referenceValue")
                    or raw_field.get("reference_value")
                    if isinstance(raw_field, dict)
                    else raw_field.reference_value
                )
                if reference_type == "disabled":
                    continue
                references.append(
                    {
                        "field": "source_key",
                        "worksheetName": str(rule.get("worksheetName") or "*"),
                        "referenceType": str(reference_type),
                        "referenceValue": reference_value,
                    }
                )
        return references

    def _identity_fingerprint(self, mapping: SourceMappingRevision) -> str:
        rule_mode, duplicate_policy, rules = self._worksheet_rule_configs(mapping)
        authority = self._identity_authority_shape(mapping)
        return checksum(
            {
                "algorithm": IDENTITY_ASSESSMENT_ALGORITHM_VERSION,
                "identityPolicyVersion": mapping.identity_policy_version,
                "identityAuthority": {
                    "type": authority["type"],
                    "systemIdentifier": authority["systemIdentifier"],
                },
                "worksheetMode": mapping.worksheet_mode,
                "worksheetRuleMode": rule_mode,
                "duplicateProductPolicy": duplicate_policy,
                "rules": [
                    {
                        "worksheetName": rule.get("worksheetName"),
                        "enabled": bool(rule.get("enabled", True)),
                        "dataStartRow": int(rule.get("dataStartRow") or 1),
                        "sourceKey": [
                            reference
                            for reference in self._identity_mapping_references(mapping)
                            if reference["worksheetName"]
                            == str(rule.get("worksheetName") or "*")
                        ],
                    }
                    for rule in rules
                ],
            }
        )

    def _assess_mapping_from_latest_local_data(
        self,
        mapping: SourceMappingRevision,
        *,
        persist: bool,
    ) -> dict[str, Any]:
        source = self.db.get(SourceProfile, mapping.source_id)
        if source is None:
            raise RuntimeError("Mapping Source is missing.")
        mapping_references = self._identity_mapping_references(mapping)
        if (
            int(mapping.identity_policy_version or 1) < 2
            or self._identity_authority_shape(mapping)["type"] == "unspecified"
        ):
            return self._pending_identity_validation(mapping_references)
        local_data = self._latest_local_validation_data(source)
        if local_data is None:
            return self._pending_identity_validation(mapping_references)
        try:
            if local_data["kind"] == "source_observation":
                records = self._mapped_external_records(
                    local_data["worksheets"], mapping
                )
            else:
                records = self._mapped_sheet_records(
                    local_data["sheetRevision"], mapping
                )
        except HTTPException as exc:
            detail = exc.detail if isinstance(exc.detail, dict) else {}
            if detail.get("code") != "WORKSHEET_NOT_FOUND":
                raise
            return self._pending_identity_validation(
                mapping_references,
                dict(local_data["evidence"]),
            )
        evidence = dict(local_data["evidence"])
        evidence["validatedAt"] = _iso_utc_timestamp(utcnow())
        summary = self._identity_preview_summary(
            records,
            mapping=mapping,
            validation_source=evidence,
        )
        if persist and summary["status"] != "pending":
            assessment = self._persist_identity_assessment(
                mapping=mapping,
                local_data=local_data,
                summary=summary,
            )
            summary["evidence"] = {
                **evidence,
                "validatedAt": _iso_utc_timestamp(assessment.validated_at),
            }
        return summary

    def _persist_identity_assessment(
        self,
        *,
        mapping: SourceMappingRevision,
        local_data: dict[str, Any],
        summary: dict[str, Any],
    ) -> SourceMappingIdentityAssessment:
        source_revision_id = str(local_data["sourceRevisionId"])
        identity_fingerprint = self._identity_fingerprint(mapping)
        binding_context_fingerprint = str(
            summary.get("bindingContextFingerprint") or ""
        )
        if len(binding_context_fingerprint) != 64:
            raise RuntimeError("Identity assessment is missing its binding context.")
        existing = (
            self.db.query(SourceMappingIdentityAssessment)
            .filter(
                SourceMappingIdentityAssessment.mapping_revision_id == mapping.id,
                SourceMappingIdentityAssessment.source_revision_kind
                == local_data["kind"],
                SourceMappingIdentityAssessment.source_revision_id
                == source_revision_id,
                SourceMappingIdentityAssessment.identity_fingerprint
                == identity_fingerprint,
                SourceMappingIdentityAssessment.binding_context_fingerprint
                == binding_context_fingerprint,
                SourceMappingIdentityAssessment.algorithm_version
                == IDENTITY_ASSESSMENT_ALGORITHM_VERSION,
            )
            .one_or_none()
        )
        if existing is not None:
            return existing
        validated_at = utcnow()
        document = {
            "mappingRevisionId": mapping.id,
            "sourceRevisionKind": local_data["kind"],
            "sourceRevisionId": source_revision_id,
            "identityFingerprint": identity_fingerprint,
            "bindingContextFingerprint": binding_context_fingerprint,
            "status": summary["status"],
            "participatingRowCount": summary["participatingRowCount"],
            "validKeyCount": summary["validKeyCount"],
            "missingKeyCount": summary["missingKeyCount"],
            "duplicateKeyCount": summary["duplicateKeyCount"],
            "duplicateRowCount": summary["duplicateRowCount"],
            "bindingConflictCount": summary["bindingConflictCount"],
            "missingRows": summary["missingRows"],
            "duplicateGroups": summary["duplicateGroups"],
            "bindingConflicts": summary["bindingConflicts"],
            "mappingReferences": summary["mappingReferences"],
            "algorithmVersion": IDENTITY_ASSESSMENT_ALGORITHM_VERSION,
        }
        assessment = SourceMappingIdentityAssessment(
            id=_id(),
            source_id=mapping.source_id,
            mapping_revision_id=mapping.id,
            dataset_id=(
                local_data["dataset"].id
                if local_data["kind"] == "source_observation"
                else None
            ),
            sheet_revision_id=(
                local_data["sheetRevision"].id
                if local_data["kind"] == "flowhub_sheet_revision"
                else None
            ),
            source_revision_kind=local_data["kind"],
            source_revision_id=source_revision_id,
            identity_fingerprint=identity_fingerprint,
            binding_context_fingerprint=binding_context_fingerprint,
            status=str(summary["status"]),
            participating_row_count=int(summary["participatingRowCount"] or 0),
            valid_key_count=int(summary["validKeyCount"] or 0),
            missing_key_count=int(summary["missingKeyCount"] or 0),
            duplicate_key_count=int(summary["duplicateKeyCount"] or 0),
            duplicate_row_count=int(summary["duplicateRowCount"] or 0),
            binding_conflict_count=int(summary["bindingConflictCount"] or 0),
            missing_rows_json=list(summary["missingRows"]),
            duplicate_groups_json=list(summary["duplicateGroups"]),
            binding_conflicts_json=list(summary["bindingConflicts"]),
            mapping_references_json=list(summary["mappingReferences"]),
            algorithm_version=IDENTITY_ASSESSMENT_ALGORITHM_VERSION,
            checksum=checksum(document),
            validated_at=validated_at,
        )
        savepoint = self.db.begin_nested()
        try:
            self.db.add(assessment)
            self.db.flush()
            savepoint.commit()
            return assessment
        except IntegrityError:
            savepoint.rollback()
            concurrent = (
                self.db.query(SourceMappingIdentityAssessment)
                .filter(
                    SourceMappingIdentityAssessment.mapping_revision_id
                    == mapping.id,
                    SourceMappingIdentityAssessment.source_revision_kind
                    == local_data["kind"],
                    SourceMappingIdentityAssessment.source_revision_id
                    == source_revision_id,
                    SourceMappingIdentityAssessment.identity_fingerprint
                    == identity_fingerprint,
                    SourceMappingIdentityAssessment.binding_context_fingerprint
                    == binding_context_fingerprint,
                    SourceMappingIdentityAssessment.algorithm_version
                    == IDENTITY_ASSESSMENT_ALGORITHM_VERSION,
                )
                .one_or_none()
            )
            if concurrent is None:
                raise
            return concurrent

    def validate_saved_mapping_identity(
        self, source_id: str, user: FlowHubUser
    ) -> dict[str, Any]:
        """Persist local evidence for the latest Mapping without provider I/O."""
        source = self._owned_source(
            source_id, user, require_active=True, lock=True
        )
        mapping = self.sources.latest_mapping(source.id)
        if mapping is None:
            raise _unprocessable(
                "SOURCE_MAPPING_REQUIRED", "Configure Source mappings first."
            )
        summary = self._assess_mapping_from_latest_local_data(mapping, persist=True)
        if summary["status"] != "pending":
            self.db.commit()
        return summary

    async def snapshot_candidates(self, source_id: str, user: FlowHubUser) -> dict[str, Any]:
        """Resolve candidates from one pinned local evidence cohort only."""
        source = self._owned_source(source_id, user, require_active=True, lock=True)
        mapping = self.sources.latest_mapping(source.id)
        sheet = self.sheets.for_source(source.id)
        if mapping is None:
            raise _unprocessable("SOURCE_MAPPING_REQUIRED", "Configure Source mappings first.")
        if int(mapping.identity_policy_version or 1) < 2:
            raise _unprocessable(
                "SOURCE_IDENTITY_POLICY_UPGRADE_REQUIRED",
                "Save this Mapping with Identity Authority and Source Product Key before opening a Workspace.",
            )
        if self._identity_authority_shape(mapping)["type"] == "unspecified":
            raise _unprocessable(
                "SOURCE_IDENTITY_AUTHORITY_REQUIRED",
                "Choose the system that owns the Source Product Key before opening a Workspace.",
            )
        local_data = self._latest_local_validation_data(source)
        if local_data is None:
            raise _unprocessable(
                "SOURCE_IDENTITY_VALIDATION_PENDING",
                "Source data must be read explicitly before identity validation can be completed.",
            )
        if sheet is None:
            if local_data["kind"] != "source_observation":
                raise RuntimeError("External Source local evidence kind is invalid.")
            dataset = local_data["dataset"]
            records = self._mapped_external_records(local_data["worksheets"], mapping)
            revision_shape = {
                "id": f"external:{dataset.source_snapshot_id}:{dataset.source_snapshot_version}",
                "version": int(dataset.source_snapshot_version),
                "checksum": str(dataset.workbook_checksum),
                "formulaEngineVersion": dataset.formula_evaluation_version,
                "observationId": dataset.observation_id,
                "datasetId": dataset.id,
            }
        else:
            if local_data["kind"] != "flowhub_sheet_revision":
                raise RuntimeError("Managed Sheet local evidence kind is invalid.")
            revision = local_data["sheetRevision"]
            if revision is None:
                raise _unprocessable("SHEET_REVISION_REQUIRED", "Save the FlowHub Sheet first.")
            records = self._mapped_sheet_records(revision, mapping)
            revision_shape = {
                "id": revision.id,
                "version": revision.version,
                "checksum": revision.checksum,
                "formulaEngineVersion": revision.formula_engine_version,
            }
        identities = {
            (channel["channelId"], str(channel["fields"]["external_id"]).strip())
            for record in records
            for channel in record["channels"]
            if str(channel["fields"].get("external_id") or "").strip()
        }
        # One pinned in-memory Listing cohort feeds identity assessment,
        # candidate resolution, binding proposals, and final lock verification.
        listings = {
            (item.channel_id, item.external_primary_id): item
            for item in self.db.query(Listing)
            .filter(
                tuple_(Listing.channel_id, Listing.external_primary_id).in_(
                    sorted(identities)
                )
            )
            .all()
        } if identities else {}
        identity_evidence = dict(local_data["evidence"])
        identity_evidence["validatedAt"] = _iso_utc_timestamp(utcnow())
        identity_validation = self._identity_preview_summary(
            records,
            mapping=mapping,
            validation_source=identity_evidence,
            listing_context=listings,
        )
        assessment = self._persist_identity_assessment(
            mapping=mapping,
            local_data=local_data,
            summary=identity_validation,
        )
        identity_validation["evidence"] = {
            **identity_evidence,
            "validatedAt": _iso_utc_timestamp(assessment.validated_at),
        }
        if identity_validation["status"] != "pass":
            # BLOCKED evidence is useful after the failed activation and does
            # not include any provider side effect.
            self.db.commit()
            raise _unprocessable(
                "SOURCE_IDENTITY_VALIDATION_BLOCKED",
                "Source Product identity must pass before this Mapping can be activated.",
                {"identityValidation": identity_validation},
            )
        listing_ids = [item.id for item in listings.values()]
        caches = {
            item.listing_id: item
            for item in self.db.query(ChannelCache)
            .filter(ChannelCache.listing_id.in_(listing_ids))
            .all()
        } if listing_ids else {}
        product_ids = {item.canonical_product_id for item in listings.values()}
        products = {
            item.id: item
            for item in self.db.query(CanonicalProduct)
            .filter(CanonicalProduct.id.in_(product_ids))
            .all()
        } if product_ids else {}
        source_key_hashes = {
            _source_product_key_hash(
                record.get("sourceProduct", {}).get("source_key")
            )
            for record in records
            if str(record.get("sourceProduct", {}).get("source_key") or "").strip()
        }
        source_identity_bindings = {
            item.source_key_hash: item
            for item in self.db.query(SourceProductIdentity)
            .filter(
                SourceProductIdentity.source_id == source.id,
                SourceProductIdentity.normalization_version
                == SOURCE_KEY_NORMALIZATION_VERSION,
                SourceProductIdentity.source_key_hash.in_(source_key_hashes),
            )
            .all()
        } if source_key_hashes else {}
        # A durable Source-key binding may only be proposed from the complete
        # resolved Listing cohort for that key. Candidate-specific failures
        # (for example a missing cache or invalid target value) must not omit
        # a Listing from the evidence that is locked and rechecked at commit.
        binding_evidence_candidates: dict[str, dict[str, Any]] = {}
        for record in records:
            normalized_key = _normalize_source_product_key(
                record.get("sourceProduct", {}).get("source_key")
            )
            if not normalized_key:
                continue
            source_key_hash = _source_product_key_hash(normalized_key)
            candidate_evidence = binding_evidence_candidates.setdefault(
                source_key_hash,
                {
                    "normalizedSourceKey": normalized_key,
                    "complete": True,
                    "canonicalProductIds": set(),
                    "listingEvidence": {},
                },
            )
            record_channels = list(record.get("channels") or [])
            if not record_channels:
                candidate_evidence["complete"] = False
            for channel in record_channels:
                channel_id = str(channel.get("channelId") or "")
                external_id = str(
                    channel.get("fields", {}).get("external_id") or ""
                ).strip()
                listing = listings.get((channel_id, external_id))
                if (
                    not external_id
                    or listing is None
                    or not listing.enabled
                    or listing.mapping_state != "resolved"
                ):
                    candidate_evidence["complete"] = False
                    continue
                candidate_evidence["canonicalProductIds"].add(
                    listing.canonical_product_id
                )
                candidate_evidence["listingEvidence"][listing.id] = {
                    "listingId": listing.id,
                    "canonicalProductId": listing.canonical_product_id,
                    "mappingVersion": listing.mapping_version,
                }
        complete_binding_evidence = {
            source_key_hash: {
                "normalizedSourceKey": evidence["normalizedSourceKey"],
                "canonicalProductId": next(iter(evidence["canonicalProductIds"])),
                "listingEvidence": sorted(
                    evidence["listingEvidence"].values(),
                    key=lambda item: item["listingId"],
                ),
            }
            for source_key_hash, evidence in binding_evidence_candidates.items()
            if evidence["complete"]
            and len(evidence["canonicalProductIds"]) == 1
            and evidence["listingEvidence"]
        }
        channel_contracts = {
            item.id: item
            for item in self.db.query(WorkspaceChannel)
            .filter(WorkspaceChannel.id.in_({channel_id for channel_id, _ in identities}))
            .all()
        } if identities else {}
        candidates: list[dict[str, Any]] = []
        issues: list[dict[str, Any]] = []
        identity_binding_proposals: dict[str, dict[str, Any]] = {}
        product_identity: dict[str, str] = {}
        source_channel_listings: dict[tuple[str, str], set[str]] = defaultdict(set)
        seen_listing_ids: set[str] = set()
        for record in records:
            source_product = record["sourceProduct"]
            # The accepted Source Product Key groups Source rows.  Product
            # Name is display evidence only and is never an identity fallback.
            group_key = _normalize_source_product_key(source_product.get("source_key"))
            blocked_channels: set[str] = set()
            global_block = False
            for row_issue in record.get("issues", []):
                issue_channel = row_issue.get("channelId")
                if issue_channel:
                    blocked_channels.add(str(issue_channel))
                else:
                    global_block = True
                issues.append(
                    self._candidate_issue(
                        record,
                        str(issue_channel) if issue_channel else None,
                        str(row_issue["category"]),
                        str(row_issue["category"]).upper(),
                        str(row_issue["message"]),
                        "Correct the Source row or its explicit Mapping policy.",
                        {},
                    )
                )
            if global_block:
                continue
            if not record["recognized"]:
                continue
            policy = dict(DEFAULT_VALUE_POLICY) | dict(
                record.get("valuePolicy") or mapping.value_policy_json
            )
            for channel in record["channels"]:
                channel_id = channel["channelId"]
                if channel_id in blocked_channels:
                    continue
                fields = channel["fields"]
                external_id = str(fields.get("external_id") or "").strip()
                listing = listings.get((channel_id, external_id))
                if listing is None:
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "missing_mapping",
                            "LISTING_NOT_MAPPED",
                            "No Channel Listing matches this External Listing ID.",
                            "Map the listing explicitly before Review.",
                            {"external_id": external_id},
                        )
                    )
                    continue
                if not listing.enabled or listing.mapping_state != "resolved":
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "mapping_not_ready",
                            "LISTING_IDENTITY_NOT_RESOLVED",
                            "The Channel Listing identity is disabled or unresolved.",
                            "Resolve and enable the Listing mapping before opening a Workspace.",
                            {
                                "listing_id": listing.id,
                                "mapping_state": listing.mapping_state,
                                "enabled": listing.enabled,
                            },
                        )
                    )
                    continue
                if listing.id in seen_listing_ids:
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "duplicate_rows",
                            "DUPLICATE_LISTING_ROW",
                            "The same Listing appears more than once in this Source revision.",
                            "Keep one authoritative row for each Listing.",
                            {"listing_id": listing.id},
                        )
                    )
                    continue
                previous_product = product_identity.get(group_key)
                source_key_value = str(source_product.get("source_key") or "").strip()
                source_key_hash = (
                    _source_product_key_hash(source_key_value)
                    if source_key_value
                    else None
                )
                existing_source_identity = (
                    source_identity_bindings.get(source_key_hash)
                    if source_key_hash is not None
                    else None
                )
                if (
                    existing_source_identity is not None
                    and existing_source_identity.canonical_product_id
                    != listing.canonical_product_id
                ):
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "mapping_conflict",
                            "SOURCE_PRODUCT_IDENTITY_CONFLICT",
                            "This Source Product Key is already bound to a different Canonical Product.",
                            "Correct the Source key or Listing mapping before continuing.",
                            {
                                "bound_product_id": existing_source_identity.canonical_product_id,
                                "conflicting_product_id": listing.canonical_product_id,
                            },
                        )
                    )
                    continue
                if previous_product and previous_product != listing.canonical_product_id:
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "mapping_conflict",
                            "SOURCE_PRODUCT_MAPPING_CONFLICT",
                            "Channel Listings under this Source Product resolve to different products.",
                            "Correct the Source key or Listing mappings before continuing.",
                            {
                                "previous_product_id": previous_product,
                                "conflicting_product_id": listing.canonical_product_id,
                            },
                        )
                    )
                    continue
                channel_group = (group_key, channel_id)
                channel_contract = channel_contracts.get(channel_id)
                if channel_contract is None:
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "unavailable_capability",
                            "CHANNEL_CONTRACT_UNAVAILABLE",
                            "The Channel capability contract is unavailable.",
                            "Refresh Channel configuration before creating a Workspace.",
                            {},
                        )
                    )
                    continue
                if (
                    channel_contract is not None
                    and not bool(
                        (channel_contract.capabilities_json or {}).get(
                            "supportsMultipleListings", False
                        )
                    )
                    and source_channel_listings[channel_group]
                    and listing.id not in source_channel_listings[channel_group]
                ):
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "mapping_conflict",
                            "CHANNEL_LISTING_CARDINALITY",
                            "This Channel supports at most one Listing for a Source Product.",
                            "Keep one Listing for this Channel or separate the Source Products.",
                            {"listing_id": listing.id},
                        )
                    )
                    continue
                product_identity[group_key] = listing.canonical_product_id
                source_channel_listings[channel_group].add(listing.id)
                product = products.get(listing.canonical_product_id)
                cache = caches.get(listing.id)
                if product is None or cache is None:
                    issues.append(
                        self._candidate_issue(
                            record,
                            channel_id,
                            "unavailable_cache",
                            "LISTING_CACHE_UNAVAILABLE",
                            "The mapped Listing or its Channel Cache is unavailable.",
                            "Refresh the Channel Cache before creating a Workspace.",
                            {"listing_id": listing.id},
                        )
                    )
                    continue
                targets, classification, classification_warnings = self._classify_channel_targets(
                    fields, policy, channel_contract, cache
                )
                if classification["blockers"]:
                    for code in classification["blockers"]:
                        issues.append(
                            self._candidate_issue(
                                record,
                                channel_id,
                                "invalid_value",
                                str(code),
                                "The Source value or its monetary contract cannot be safely classified.",
                                "Correct the Source value or Mapping currency/precision evidence.",
                                {"classification": classification},
                            )
                        )
                    continue
                for warning in classification_warnings:
                    issues.append(
                        {
                            **self._candidate_issue(
                                record,
                                channel_id,
                                "warning",
                                str(warning["code"]),
                                "The mapped Price is unusable and requests out of stock.",
                                "Correct the Source Price if out of stock was not intended.",
                                {"field": warning["field"], "classification": classification},
                            ),
                            "severity": "warning",
                        }
                    )
                seen_listing_ids.add(listing.id)
                candidates.append(
                    {
                        "sourceRowKey": record["rowKey"],
                        "sourceRowNumber": record["rowNumber"],
                        "sourceProduct": source_product,
                        "canonicalProductId": product.id,
                        "listingId": listing.id,
                        "channelId": listing.channel_id,
                        "mappingVersion": listing.mapping_version,
                        "cacheVersion": cache.cache_version,
                        "targets": targets,
                        "classification": classification,
                    }
                )
                complete_evidence = (
                    complete_binding_evidence.get(source_key_hash)
                    if source_key_hash is not None
                    else None
                )
                if complete_evidence is not None:
                    if (
                        complete_evidence["canonicalProductId"]
                        != listing.canonical_product_id
                    ):
                        raise RuntimeError(
                            "Conflicting Canonical Products escaped candidate validation."
                        )
                    identity_binding_proposals.setdefault(
                        source_key_hash,
                        {
                            "sourceKeyHash": source_key_hash,
                            "normalizedSourceKey": complete_evidence[
                                "normalizedSourceKey"
                            ],
                            "normalizationVersion": SOURCE_KEY_NORMALIZATION_VERSION,
                            "canonicalProductId": complete_evidence[
                                "canonicalProductId"
                            ],
                            "listingEvidence": complete_evidence[
                                "listingEvidence"
                            ],
                        },
                    )
        source_shape = self._source_shape(source)
        mapping_shape = self._mapping_shape(mapping)
        if mapping_shape is None:
            raise RuntimeError("Source Mapping disappeared during candidate analysis.")
        exact_readiness = self._mapping_readiness(identity_validation)
        source_shape["mappingReadiness"] = exact_readiness
        mapping_shape["identityValidation"] = identity_validation
        mapping_shape["mappingReadiness"] = exact_readiness
        return {
            "source": source_shape,
            "mapping": mapping_shape,
            "sheetRevision": revision_shape,
            "identityBindingProposals": [
                {
                    **proposal,
                    "listingEvidence": sorted(
                        proposal["listingEvidence"],
                        key=lambda item: item["listingId"],
                    ),
                }
                for _, proposal in sorted(identity_binding_proposals.items())
            ],
            "candidates": candidates,
            "issues": issues,
            "summary": {
                "sourceProducts": len({str(item["sourceProduct"].get("source_key") or item["sourceProduct"].get("name")) for item in candidates}),
                "listings": len(candidates),
                "blocked": len(issues),
            },
        }

    # -- FlowHub Sheet ------------------------------------------------------

    def create_sheet(
        self, *, name: str, columns: list[dict[str, Any]], user: FlowHubUser
    ) -> dict[str, Any]:
        source_shape = self.create_source(
            name=name,
            source_kind="flowhub_sheet",
            external_source_id=None,
            worksheet_mode="selected",
            worksheet_name="Sheet1",
            data_start_row=1,
            user=user,
        )
        sheet = self.sheets.for_source(source_shape["id"])
        if sheet is None:
            raise RuntimeError("Sheet persistence failed")
        if columns:
            self.save_sheet_revision(
                sheet_id=sheet.id,
                expected_version=0,
                columns=columns,
                rows=[],
                user=user,
            )
        return self.get_sheet(sheet.id, user, page=1, page_size=200)

    def get_sheet(
        self,
        sheet_id: str,
        user: FlowHubUser,
        *,
        page: int,
        page_size: int,
        search: str | None = None,
        sort_column: str | None = None,
        sort_direction: str = "asc",
    ) -> dict[str, Any]:
        sheet = self._owned_sheet(sheet_id, user)
        revision = self.sheets.latest_revision(sheet.id)
        if revision is None:
            return {
                "id": sheet.id,
                "sourceId": sheet.source_id,
                "name": sheet.name,
                "version": 0,
                "revisionId": None,
                "columns": [],
                "rows": [],
                "total": 0,
                "page": page,
                "pageSize": page_size,
            }
        page_size = min(max(page_size, 1), 500)
        columns = self.sheets.columns(revision.id)
        column_keys = {item.column_key for item in columns}
        if sort_column and sort_column not in column_keys:
            raise _unprocessable("SHEET_SORT_COLUMN_INVALID", "Sort requires a persisted Column identity.")
        if sort_direction not in {"asc", "desc"}:
            raise _unprocessable("SHEET_SORT_DIRECTION_INVALID", "Sort direction must be asc or desc.")
        rows, total = self.sheets.rows(
            revision.id,
            offset=(max(page, 1) - 1) * page_size,
            limit=page_size,
            search=(search or "").strip()[:240] or None,
            sort_column=sort_column,
            sort_direction=sort_direction,
        )
        cells = self.sheets.cells(revision.id, [row.id for row in rows])
        by_row: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
        for cell in cells:
            by_row[cell.row_id][cell.column_key] = {
                "raw": cell.raw_value,
                "value": cell.calculated_value,
                "formula": cell.formula_expression,
                "error": cell.calculation_error,
            }
        return {
            "id": sheet.id,
            "sourceId": sheet.source_id,
            "name": sheet.name,
            "version": sheet.current_version,
            "revisionId": revision.id,
            "revisionChecksum": revision.checksum,
            "formulaEngineVersion": revision.formula_engine_version,
            "columns": [self._column_shape(item) for item in columns],
            "rows": [
                {
                    "rowKey": row.row_key,
                    "position": row.position,
                    "cells": by_row.get(row.id, {}),
                }
                for row in rows
            ],
            "total": total,
            "page": page,
            "pageSize": page_size,
        }

    def save_sheet_revision(
        self,
        *,
        sheet_id: str,
        expected_version: int,
        columns: list[dict[str, Any]],
        rows: list[dict[str, Any]],
        user: FlowHubUser,
    ) -> dict[str, Any]:
        sheet = self._owned_sheet(sheet_id, user)
        self._owned_source(sheet.source_id, user, require_active=True, lock=True)
        if sheet.current_version != expected_version:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "SHEET_VERSION_CONFLICT", "message": "Sheet was changed elsewhere."},
            )
        normalized_columns = self._normalize_columns(columns)
        normalized_rows = self._normalize_rows(rows, normalized_columns)
        formula_results = self._calculate(normalized_columns, normalized_rows)
        version = sheet.current_version + 1
        document = {
            "sheetId": sheet.id,
            "version": version,
            "columns": normalized_columns,
            "rows": normalized_rows,
            "formulaEngine": FORMULA_ENGINE_VERSION,
        }
        revision = SheetRevision(
            id=_id(),
            sheet_id=sheet.id,
            version=version,
            checksum=checksum(document),
            formula_engine_version=FORMULA_ENGINE_VERSION,
            row_count=len(normalized_rows),
            column_count=len(normalized_columns),
            created_by_user_id=user.id,
        )
        self.db.add(revision)
        self.db.flush()
        column_models = [
            SheetColumn(
                id=_id(),
                revision_id=revision.id,
                column_key=item["columnKey"],
                name=item["name"],
                position=item["position"],
                data_type=item["dataType"],
            )
            for item in normalized_columns
        ]
        self.db.add_all(column_models)
        self.db.flush()
        row_models: list[SheetRow] = []
        cell_models: list[SheetCell] = []
        column_position = {item["columnKey"]: item["position"] for item in normalized_columns}
        for item in normalized_rows:
            row_model = SheetRow(
                id=_id(),
                revision_id=revision.id,
                row_key=item["rowKey"],
                position=item["position"],
            )
            row_models.append(row_model)
            for column_key, raw in item["values"].items():
                reference = f"{column_name(column_position[column_key])}{item['position']}"
                result = formula_results.get(reference)
                formula = str(raw) if raw is not None and str(raw).lstrip().startswith("=") else None
                cell_models.append(
                    SheetCell(
                        id=_id(),
                        revision_id=revision.id,
                        row_id=row_model.id,
                        column_key=column_key,
                        raw_value=None if raw is None else str(raw),
                        calculated_value=result.value if result else None if raw is None else str(raw),
                        formula_expression=formula,
                        formula_dependencies_json=list(result.dependencies) if result else [],
                        calculation_error=result.error if result else None,
                    )
                )
        # One flush/commit for the full bulk revision, never one transaction per cell.
        self.db.add_all(row_models)
        # SheetCell stores scalar foreign keys rather than ORM relationships, so
        # SQLAlchemy cannot infer that pending rows must be inserted first.
        self.db.flush()
        self.db.add_all(cell_models)
        sheet.current_version = version
        sheet.updated_at = utcnow()
        self.db.commit()
        return self.get_sheet(sheet.id, user, page=1, page_size=200)

    def patch_sheet_revision(
        self,
        *,
        sheet_id: str,
        expected_version: int,
        changes: list[dict[str, Any]],
        column_names: dict[str, str] | None = None,
        user: FlowHubUser,
    ) -> dict[str, Any]:
        """Create a full immutable revision from a bounded, identity-based cell patch."""
        sheet = self._owned_sheet(sheet_id, user)
        if sheet.current_version != expected_version:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "SHEET_VERSION_CONFLICT", "message": "Sheet was changed elsewhere."},
            )
        revision = self.sheets.latest_revision(sheet.id)
        if revision is None:
            raise _unprocessable("SHEET_REVISION_REQUIRED", "Create the first Sheet revision first.")
        columns = [self._column_shape(item) for item in self.sheets.columns(revision.id)]
        column_names = column_names or {}
        column_keys = {item["columnKey"] for item in columns}
        if set(column_names) - column_keys:
            raise _unprocessable(
                "SHEET_COLUMN_IDENTITY_INVALID",
                "Column name changes require persisted Column identities.",
            )
        for column in columns:
            replacement = column_names.get(column["columnKey"])
            if replacement is not None:
                column["name"] = _clean_name(replacement, column["name"])
        row_models = self.sheets.all_rows(revision.id)
        cells = self.sheets.cells(revision.id, [item.id for item in row_models])
        values_by_row: dict[str, dict[str, str | None]] = defaultdict(dict)
        row_key_by_id = {item.id: item.row_key for item in row_models}
        for cell in cells:
            values_by_row[row_key_by_id[cell.row_id]][cell.column_key] = cell.raw_value
        row_keys = {item.row_key for item in row_models}
        seen: set[tuple[str, str]] = set()
        for change in changes:
            row_key = str(change.get("row_key") or change.get("rowKey") or "")
            column_key = str(change.get("column_key") or change.get("columnKey") or "")
            identity = (row_key, column_key)
            if row_key not in row_keys or column_key not in column_keys or identity in seen:
                raise _unprocessable(
                    "SHEET_PATCH_IDENTITY_INVALID",
                    "Cell patches require unique persisted Row and Column identities.",
                )
            seen.add(identity)
            value = change.get("value")
            values_by_row[row_key][column_key] = None if value is None else str(value)
        rows = [
            {
                "rowKey": item.row_key,
                "position": item.position,
                "values": values_by_row[item.row_key],
            }
            for item in row_models
        ]
        return self.save_sheet_revision(
            sheet_id=sheet.id,
            expected_version=expected_version,
            columns=columns,
            rows=rows,
            user=user,
        )

    def append_sheet_rows(
        self,
        *,
        sheet_id: str,
        expected_version: int,
        count: int,
        user: FlowHubUser,
    ) -> dict[str, Any]:
        sheet = self._owned_sheet(sheet_id, user)
        if sheet.current_version != expected_version:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "SHEET_VERSION_CONFLICT", "message": "Sheet was changed elsewhere."},
            )
        revision = self.sheets.latest_revision(sheet.id)
        if revision is None:
            raise _unprocessable("SHEET_REVISION_REQUIRED", "Create the first Sheet revision first.")
        if count < 1 or revision.row_count + count > MAX_SHEET_ROWS:
            raise _unprocessable("SHEET_ROW_LIMIT", f"A Sheet supports at most {MAX_SHEET_ROWS} rows.")
        columns = [self._column_shape(item) for item in self.sheets.columns(revision.id)]
        row_models = self.sheets.all_rows(revision.id)
        cells = self.sheets.cells(revision.id, [item.id for item in row_models])
        row_key_by_id = {item.id: item.row_key for item in row_models}
        values_by_row: dict[str, dict[str, str | None]] = defaultdict(dict)
        for cell in cells:
            values_by_row[row_key_by_id[cell.row_id]][cell.column_key] = cell.raw_value
        rows = [
            {"rowKey": item.row_key, "position": item.position, "values": values_by_row[item.row_key]}
            for item in row_models
        ]
        last_position = max((item.position for item in row_models), default=0)
        rows.extend(
            {"rowKey": _id(), "position": last_position + offset, "values": {}}
            for offset in range(1, count + 1)
        )
        return self.save_sheet_revision(
            sheet_id=sheet.id,
            expected_version=expected_version,
            columns=columns,
            rows=rows,
            user=user,
        )

    def calculate(
        self, *, columns: list[dict[str, Any]], rows: list[dict[str, Any]]
    ) -> dict[str, Any]:
        normalized_columns = self._normalize_columns(columns)
        normalized_rows = self._normalize_rows(rows, normalized_columns)
        results = self._calculate(normalized_columns, normalized_rows)
        return {
            "engineVersion": FORMULA_ENGINE_VERSION,
            "cells": {
                reference: {
                    "value": result.value,
                    "dependencies": list(result.dependencies),
                    "error": result.error,
                }
                for reference, result in results.items()
            },
        }

    # -- Import -------------------------------------------------------------

    def preview_import(
        self, *, filename: str, content_base64: str, worksheet_name: str | None
    ) -> dict[str, Any]:
        content = self._decode_import(content_base64)
        sheets = self._read_import(filename, content)
        selected = worksheet_name or next(iter(sheets), None)
        if selected is None or selected not in sheets:
            raise _unprocessable("WORKSHEET_NOT_FOUND", "Select an available worksheet.")
        rows = sheets[selected]
        width = max((len(row) for row in rows), default=0)
        headers = [str(value or f"Column {column_name(index)}") for index, value in enumerate(rows[0] if rows else [], start=1)]
        return {
            "filename": filename,
            "sourceChecksum": checksum(content.hex()),
            "worksheets": list(sheets),
            "selectedWorksheet": selected,
            "rowCount": len(rows),
            "columnCount": width,
            "headers": headers,
            "previewRows": rows[:50],
            "truncated": len(rows) > 50,
        }

    def import_sheet(
        self,
        *,
        name: str,
        filename: str,
        content_base64: str,
        worksheet_name: str,
        expected_checksum: str,
        data_start_row: int,
        user: FlowHubUser,
    ) -> dict[str, Any]:
        content = self._decode_import(content_base64)
        source_checksum = checksum(content.hex())
        if source_checksum != expected_checksum:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {"code": "IMPORT_CONTENT_CHANGED", "message": "Import file changed after preview."},
            )
        sheets = self._read_import(filename, content)
        if worksheet_name not in sheets:
            raise _unprocessable("WORKSHEET_NOT_FOUND", "Selected worksheet does not exist.")
        source = self.create_source(
            name=name,
            source_kind="imported_sheet",
            external_source_id=None,
            worksheet_mode="selected",
            worksheet_name=worksheet_name,
            data_start_row=data_start_row,
            user=user,
        )
        sheet = self.sheets.for_source(source["id"])
        if sheet is None:
            raise RuntimeError("Imported sheet persistence failed")
        imported_rows = sheets[worksheet_name]
        width = max((len(row) for row in imported_rows), default=0)
        columns = [
            {
                "columnKey": f"col-{column_name(index).lower()}",
                "name": _clean_name(
                    imported_rows[0][index - 1] if imported_rows and len(imported_rows[0]) >= index else None,
                    f"Column {column_name(index)}",
                ),
                "position": index,
                "dataType": "text",
            }
            for index in range(1, width + 1)
        ]
        rows = [
            {
                "rowKey": _id(),
                "position": index,
                "values": {
                    columns[column_index]["columnKey"]: value
                    for column_index, value in enumerate(row)
                    if column_index < len(columns) and value is not None
                },
            }
            for index, row in enumerate(imported_rows, start=1)
        ]
        result = self.save_sheet_revision(
            sheet_id=sheet.id,
            expected_version=0,
            columns=columns,
            rows=rows,
            user=user,
        )
        self.db.add(
            SheetImportJob(
                id=_id(),
                sheet_id=sheet.id,
                source_type="xlsx" if filename.lower().endswith(".xlsx") else "csv",
                source_filename=filename[:500],
                worksheet_name=worksheet_name,
                imported_row_count=len(rows),
                mapping_version=0,
                source_checksum=source_checksum,
                status="completed",
                metadata_json={"original_unchanged": True, "available_worksheets": list(sheets)},
                created_by_user_id=user.id,
            )
        )
        self.db.commit()
        return result

    # -- Data Quality -------------------------------------------------------

    async def scan_data_quality(
        self,
        *,
        user: FlowHubUser,
        source_id: str | None,
    ) -> dict[str, Any]:
        """Evaluate each selected Source once and persist an explicit scan result."""
        if source_id:
            sources = [self._owned_source(source_id, user, require_active=True)]
        else:
            sources = [
                source
                for source in self.sources.list_for_user(user.id)
                if source.status == "active"
            ]
        source_ids = [source.id for source in sources]
        scan = SourceDataQualityScan(
            id=_id(),
            owner_user_id=user.id,
            source_id=source_id,
            source_ids_json=source_ids,
            source_results_json={},
            status="checking",
            sources_checked=0,
            products_checked=0,
            issue_count=0,
            blocking_issue_count=0,
            warning_count=0,
            affected_product_count=0,
            affected_channel_count=0,
            affected_source_count=0,
            previous_issue_count=None,
            resolved_since_previous=0,
            error_code=None,
            created_at=utcnow(),
            checked_at=None,
        )
        self.db.add(scan)
        self.db.add_all(
            SourceDataQualityScanSource(scan_id=scan.id, source_id=item)
            for item in source_ids
        )
        self.db.commit()

        try:
            pending_issues: list[SourceDataQualityIssue] = []
            all_product_rows: set[tuple[str, str]] = set()
            affected_products: set[tuple[str, str]] = set()
            affected_channels: set[str] = set()
            affected_sources: set[str] = set()
            source_results: dict[str, dict[str, int]] = {}
            blocking_count = 0
            warning_count = 0

            for source in sources:
                # Candidate analysis replays one pinned local evidence cohort.
                # Data Quality never performs a scan-specific Source read.
                try:
                    analysis = await self.snapshot_candidates(source.id, user)
                except HTTPException as exc:
                    detail: dict[str, Any] = (
                        exc.detail if isinstance(exc.detail, dict) else {}
                    )
                    code = str(detail.get("code") or "")
                    configuration_issues = {
                        "SOURCE_MAPPING_REQUIRED": (
                            "mapping_not_configured",
                            "Source columns have not been configured.",
                            "Choose the Source Product and Channel columns before running the check.",
                        ),
                        "SHEET_REVISION_REQUIRED": (
                            "source_not_saved",
                            "The FlowHub Sheet has not been saved yet.",
                            "Save the Sheet, then run the Data Quality check again.",
                        ),
                        "SOURCE_IDENTITY_POLICY_UPGRADE_REQUIRED": (
                            "identity_mapping_not_ready",
                            "The saved Mapping uses a historical identity policy.",
                            "Save Identity Authority and Source Product Key before running the check.",
                        ),
                        "SOURCE_IDENTITY_AUTHORITY_REQUIRED": (
                            "identity_authority_required",
                            "The Source Product Key has no explicit Identity Authority.",
                            "Choose the system that owns the Source Product Key, then save the Mapping.",
                        ),
                        "SOURCE_IDENTITY_VALIDATION_PENDING": (
                            "identity_validation_pending",
                            "No compatible local Source data is available for identity validation.",
                            "Use Read Source explicitly, then run the Data Quality check again.",
                        ),
                        "SOURCE_IDENTITY_VALIDATION_BLOCKED": (
                            "identity_validation_blocked",
                            "Source Product identity validation is blocked.",
                            "Correct the missing, duplicate, or conflicting Source Product Keys.",
                        ),
                    }
                    if code not in configuration_issues:
                        raise
                    category, summary, action = configuration_issues[code]
                    analysis = {
                        "candidates": [],
                        "issues": [
                            {
                                "sourceRowKey": None,
                                "sourceRowNumber": None,
                                "worksheetName": source.worksheet_name,
                                "channelId": None,
                                "sourceProductName": None,
                                "mappingState": "unmapped",
                                "category": category,
                                "severity": "blocked",
                                "code": code,
                                "summary": summary,
                                "recommendedAction": action,
                                "technicalDetails": (
                                    dict(detail.get("details") or {})
                                    if code == "SOURCE_IDENTITY_VALIDATION_BLOCKED"
                                    else {}
                                ),
                            }
                        ],
                    }
                analysis_issues = [dict(item) for item in analysis.get("issues", [])]
                source_product_rows: set[str] = {
                    str(item.get("sourceRowKey") or "")
                    for item in analysis.get("candidates", [])
                    if item.get("sourceRowKey")
                }
                source_product_rows.update(
                    str(item.get("sourceRowKey") or "")
                    for item in analysis_issues
                    if item.get("sourceRowKey")
                )
                all_product_rows.update((source.id, row_key) for row_key in source_product_rows)

                source_blocking = 0
                source_warnings = 0
                source_affected_products: set[str] = set()
                source_affected_channels: set[str] = set()
                for issue in analysis_issues:
                    severity = str(issue.get("severity") or "blocked")
                    if severity not in {"warning", "error", "blocked"}:
                        severity = "blocked"
                    if severity == "warning":
                        source_warnings += 1
                        warning_count += 1
                    else:
                        source_blocking += 1
                        blocking_count += 1
                    row_key = str(issue.get("sourceRowKey") or "")
                    if row_key:
                        source_affected_products.add(row_key)
                        affected_products.add((source.id, row_key))
                    channel_id = str(issue.get("channelId") or "").strip() or None
                    if channel_id:
                        source_affected_channels.add(channel_id)
                        affected_channels.add(channel_id)
                    affected_sources.add(source.id)
                    technical_details = dict(issue.get("technicalDetails") or {})
                    listing_id = str(technical_details.get("listing_id") or "").strip()
                    listing = self.db.get(Listing, listing_id) if listing_id else None
                    pending_issues.append(
                        SourceDataQualityIssue(
                            id=_id(),
                            scan_id=scan.id,
                            source_id=source.id,
                            snapshot_id=None,
                            worksheet_name=(
                                str(issue.get("worksheetName"))[:240]
                                if issue.get("worksheetName") is not None
                                else None
                            ),
                            source_row_key=self._data_quality_source_row_key(
                                issue.get("sourceRowKey")
                            ),
                            source_product_name=(
                                str(issue.get("sourceProductName"))[:240]
                                if issue.get("sourceProductName") is not None
                                else None
                            ),
                            mapping_state=(
                                str(issue.get("mappingState"))[:40]
                                if issue.get("mappingState") is not None
                                else None
                            ),
                            channel_id=channel_id,
                            canonical_product_id=(
                                listing.canonical_product_id if listing is not None else None
                            ),
                            category=str(issue.get("category") or "validation")[:80],
                            severity=severity,
                            code=str(issue.get("code") or "SOURCE_VALIDATION_FAILED")[:120],
                            summary=str(issue.get("summary") or "Source validation failed.")[:500],
                            recommended_action=str(
                                issue.get("recommendedAction")
                                or "Review the Source row and its configured columns."
                            )[:1000],
                            technical_details_json=technical_details,
                            created_at=utcnow(),
                        )
                    )
                source_results[source.id] = {
                    "productsChecked": len(source_product_rows),
                    "issueCount": len(analysis_issues),
                    "blockingIssues": source_blocking,
                    "warnings": source_warnings,
                    "affectedProducts": len(source_affected_products),
                    "affectedChannels": len(source_affected_channels),
                    "affectedSources": 1 if analysis_issues else 0,
                }

            previous = self.issues.previous_completed_scan(
                user_id=user.id,
                source_id=source_id,
                exclude_scan_id=scan.id,
            )
            previous_issue_count: int | None = None
            previous_issue_identities: set[
                tuple[str, str, str, str, str, str, str, str]
            ] = set()
            if previous is not None:
                if source_id is not None and previous.source_id is None:
                    previous_source_result = dict(
                        previous.source_results_json.get(source_id) or {}
                    )
                    previous_issue_count = int(
                        previous_source_result.get("issueCount", 0)
                    )
                else:
                    previous_issue_count = previous.issue_count
                previous_issue_identities = self.issues.issue_identity_keys(
                    previous.id,
                    source_id=source_id,
                )
            current_issue_identities = {
                self._data_quality_issue_identity(item) for item in pending_issues
            }
            persisted_scan = self.db.get(SourceDataQualityScan, scan.id)
            if persisted_scan is None:
                raise RuntimeError("Data Quality scan persistence was lost")
            scan = persisted_scan
            scan.sources_checked = len(sources)
            scan.products_checked = len(all_product_rows)
            scan.issue_count = len(pending_issues)
            scan.blocking_issue_count = blocking_count
            scan.warning_count = warning_count
            scan.affected_product_count = len(affected_products)
            scan.affected_channel_count = len(affected_channels)
            scan.affected_source_count = len(affected_sources)
            scan.previous_issue_count = previous_issue_count
            scan.resolved_since_previous = len(
                previous_issue_identities - current_issue_identities
            )
            scan.source_results_json = source_results
            self.db.add_all(pending_issues)
            # Issue rows must be persisted while the durable scan is still in
            # its only mutable state.  The database seals terminal scans and
            # rejects any issue appended after this flush.
            self.db.flush()
            scan.status = "completed"
            scan.checked_at = utcnow()
            self.db.commit()
            return {"summary": self._data_quality_summary(scan, source_id=source_id)}
        except Exception as exc:
            self.db.rollback()
            persisted = self.db.get(SourceDataQualityScan, scan.id)
            if persisted is not None:
                persisted.status = "failed"
                persisted.sources_checked = len(sources)
                persisted.error_code = self._data_quality_error_code(exc)
                persisted.checked_at = utcnow()
                self.db.commit()
            raise

    def data_quality(
        self,
        *,
        user: FlowHubUser,
        source_id: str | None,
        channel_id: str | None,
        worksheet: str | None,
        category: str | None,
        severity: str | None,
        product: str | None,
        mapping_state: str | None,
        page: int,
        page_size: int,
    ) -> dict[str, Any]:
        normalized_source = self._data_quality_filter(source_id)
        normalized_channel = self._data_quality_filter(channel_id)
        normalized_worksheet = self._data_quality_filter(worksheet)
        normalized_category = self._data_quality_filter(category)
        normalized_severity = self._data_quality_filter(severity)
        normalized_product = self._data_quality_filter(product)
        normalized_mapping_state = self._data_quality_filter(mapping_state)
        scan = self.issues.latest_scan(user_id=user.id, source_id=normalized_source)
        if scan is None:
            return {
                "items": [],
                "counts": {},
                "total": 0,
                "page": page,
                "pageSize": page_size,
                "summary": self._data_quality_summary(None, source_id=normalized_source),
            }
        items, total, counts = self.issues.list(
            user_id=user.id,
            scan_id=scan.id,
            source_id=normalized_source,
            channel_id=normalized_channel,
            worksheet=normalized_worksheet,
            category=normalized_category,
            severity=normalized_severity,
            product=normalized_product,
            mapping_state=normalized_mapping_state,
            page=max(page, 1),
            page_size=min(max(page_size, 1), 200),
        )
        return {
            "items": [
                {
                    "id": item.id,
                    "scanId": item.scan_id,
                    "sourceId": item.source_id,
                    "sourceRowKey": item.source_row_key,
                    "worksheet": item.worksheet_name,
                    "sourceProductName": item.source_product_name,
                    "mappingState": item.mapping_state,
                    "channelId": item.channel_id,
                    "category": item.category,
                    "severity": item.severity,
                    "code": item.code,
                    "summary": item.summary,
                    "recommendedAction": item.recommended_action,
                    "technicalDetails": item.technical_details_json,
                }
                for item in items
            ],
            "counts": dict(counts),
            "total": total,
            "page": page,
            "pageSize": page_size,
            "summary": self._data_quality_summary(scan, source_id=normalized_source),
        }

    def _data_quality_summary(
        self,
        scan: SourceDataQualityScan | None,
        *,
        source_id: str | None,
    ) -> dict[str, Any]:
        if scan is None:
            return {
                "state": "never_checked",
                "totalIssues": 0,
                "blockingIssues": 0,
                "warnings": 0,
                "affectedProducts": 0,
                "affectedChannels": 0,
                "affectedSources": 0,
                "resolvedSinceLastRead": 0,
                "trendSinceLastRead": None,
                "productsChecked": 0,
                "sourcesChecked": 0,
                "checkedAt": None,
                "scanId": None,
                "errorCode": None,
                "categories": [],
            }
        scoped = (
            dict(scan.source_results_json.get(source_id) or {})
            if source_id is not None and scan.source_id is None
            else {}
        )
        total_issues = int(scoped.get("issueCount", scan.issue_count))
        if scan.status == "checking":
            state = "checking"
        elif scan.status == "failed":
            state = "failed"
        else:
            state = "issues_found" if total_issues else "healthy"
        scoped_to_global_scan = bool(scoped)
        categories = self.issues.categories(scan.id, source_id=source_id)
        return {
            "state": state,
            "totalIssues": total_issues,
            "blockingIssues": int(scoped.get("blockingIssues", scan.blocking_issue_count)),
            "warnings": int(scoped.get("warnings", scan.warning_count)),
            "affectedProducts": int(
                scoped.get("affectedProducts", scan.affected_product_count)
            ),
            "affectedChannels": int(
                scoped.get("affectedChannels", scan.affected_channel_count)
            ),
            "affectedSources": int(
                scoped.get("affectedSources", scan.affected_source_count)
            ),
            "resolvedSinceLastRead": (
                0 if scoped_to_global_scan else scan.resolved_since_previous
            ),
            "trendSinceLastRead": (
                None
                if scoped_to_global_scan or scan.previous_issue_count is None
                else scan.issue_count - scan.previous_issue_count
            ),
            "productsChecked": int(scoped.get("productsChecked", scan.products_checked)),
            "sourcesChecked": 1 if scoped_to_global_scan else scan.sources_checked,
            "checkedAt": _utc_timestamp(scan.checked_at),
            "scanId": scan.id,
            "errorCode": scan.error_code,
            "categories": [
                {"category": category, "count": count}
                for category, count in sorted(
                    categories.items(), key=lambda item: (-item[1], item[0])
                )
            ],
        }

    @staticmethod
    def _data_quality_filter(value: str | None) -> str | None:
        normalized = str(value or "").strip()
        return None if not normalized or normalized.casefold() == "all" else normalized

    @staticmethod
    def _data_quality_source_row_key(value: object) -> str | None:
        """Return a complete bounded Source-row identity without truncation."""
        normalized = str(value or "").strip()
        if not normalized:
            return None
        if len(normalized) > 512:
            raise _unprocessable(
                "SOURCE_ROW_IDENTITY_TOO_LONG",
                "The Source row identity exceeds the supported length.",
            )
        return normalized

    @staticmethod
    def _data_quality_issue_identity(
        issue: SourceDataQualityIssue,
    ) -> tuple[str, str, str, str, str, str, str, str]:
        return (
            str(issue.source_id or ""),
            str(issue.worksheet_name or ""),
            str(issue.source_row_key or ""),
            str(issue.source_product_name or ""),
            str(issue.channel_id or ""),
            str(issue.mapping_state or ""),
            str(issue.category or ""),
            str(issue.code or ""),
        )

    @staticmethod
    def _data_quality_error_code(exc: Exception) -> str:
        if isinstance(exc, HTTPException) and isinstance(exc.detail, dict):
            return str(exc.detail.get("code") or "DATA_QUALITY_SCAN_FAILED")[:120]
        return type(exc).__name__.upper()[:120] or "DATA_QUALITY_SCAN_FAILED"

    # -- Helpers ------------------------------------------------------------

    def _owned_source(
        self,
        source_id: str,
        user: FlowHubUser,
        *,
        require_active: bool = False,
        lock: bool = False,
    ) -> SourceProfile:
        query = self.db.query(SourceProfile).filter(SourceProfile.id == source_id)
        if lock:
            query = query.with_for_update()
        source = query.populate_existing().one_or_none()
        if (
            source is None
            or source.status == "deleted"
            or (source.owner_user_id != user.id and user.role != "admin")
        ):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Source not found.")
        if require_active and source.status == "archived":
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_ARCHIVED",
                    "message": "Archived Sources are read-only and cannot start new processing.",
                },
            )
        if require_active and source.status != "active":
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                {
                    "code": "SOURCE_DISABLED",
                    "message": "Disabled Sources cannot start new processing.",
                },
            )
        return source

    def _source_lifecycle_impact(self, source: SourceProfile) -> dict[str, Any]:
        sheet = self.sheets.for_source(source.id)
        sheet_revision_count = (
            self.db.query(SheetRevision).filter(SheetRevision.sheet_id == sheet.id).count()
            if sheet is not None
            else 0
        )
        import_count = (
            self.db.query(SheetImportJob).filter(SheetImportJob.sheet_id == sheet.id).count()
            if sheet is not None
            else 0
        )
        workspace_rows = (
            self.db.query(WorkspaceSnapshot, UnifiedWorkspace)
            .join(UnifiedWorkspace, UnifiedWorkspace.id == WorkspaceSnapshot.workspace_id)
            .filter(WorkspaceSnapshot.entry_point == "source")
            .all()
        )
        matching_workspaces = [
            (snapshot, workspace)
            for snapshot, workspace in workspace_rows
            if str((snapshot.source_metadata_json or {}).get("source_id") or "") == source.id
        ]
        workspace_binding_count = self.db.query(WorkspaceSourceBinding).filter(
            WorkspaceSourceBinding.source_id == source.id
        ).count()
        bound_workspace_rows = (
            self.db.query(WorkspaceSourceBinding.workspace_id, UnifiedWorkspace.status)
            .join(UnifiedWorkspace, UnifiedWorkspace.id == WorkspaceSourceBinding.workspace_id)
            .filter(WorkspaceSourceBinding.source_id == source.id)
            .all()
        )
        active_workspace_ids = {
            workspace_id
            for workspace_id, workspace_status in bound_workspace_rows
            if workspace_status == "active"
        }
        acquisition_runs = self.db.query(AcquisitionRun).filter(
            AcquisitionRun.source_id == source.id
        )
        active_acquisition_count = acquisition_runs.filter(
            AcquisitionRun.status.in_(ACTIVE_RUN_STATUSES)
        ).count()
        scan_ids = {
            scan.id
            for scan in self.db.query(SourceDataQualityScan)
            .filter(SourceDataQualityScan.owner_user_id == source.owner_user_id)
            .all()
            if scan.source_id == source.id or source.id in scan.source_ids_json
        }
        scan_ids.update(
            scan_id
            for (scan_id,) in self.db.query(SourceDataQualityScanSource.scan_id)
            .filter(SourceDataQualityScanSource.source_id == source.id)
            .all()
        )
        audit_history_count = sum(
            1
            for audit in self.db.query(UnifiedAuditEntry).all()
            if str(
                (audit.metadata_json or {}).get("sourceId")
                or (audit.metadata_json or {}).get("source_id")
                or ""
            )
            == source.id
        )
        protected_counts = {
            "mappingRevisions": self.db.query(SourceMappingRevision)
            .filter(SourceMappingRevision.source_id == source.id)
            .count(),
            "sheetRevisions": sheet_revision_count,
            "importJobs": import_count,
            "dataQualityIssues": self.db.query(SourceDataQualityIssue)
            .filter(SourceDataQualityIssue.source_id == source.id)
            .count(),
            "dataQualityScans": len(scan_ids),
            "workspaceSnapshots": len(matching_workspaces),
            "workspaceBindings": workspace_binding_count,
            "acquisitionRuns": acquisition_runs.count(),
            "sourceObservationVersionHeads": self.db.query(SourceObservationVersionHead)
            .filter(SourceObservationVersionHead.source_id == source.id)
            .count(),
            "sourceObservations": self.db.query(SourceObservation)
            .filter(SourceObservation.source_id == source.id)
            .count(),
            "sourceObservationDatasets": self.db.query(SourceObservationDataset)
            .filter(SourceObservationDataset.source_id == source.id)
            .count(),
            "identityAssessments": self.db.query(SourceMappingIdentityAssessment)
            .filter(SourceMappingIdentityAssessment.source_id == source.id)
            .count(),
            "sourceProductIdentities": self.db.query(SourceProductIdentity)
            .filter(SourceProductIdentity.source_id == source.id)
            .count(),
            "mappingSchemaExpectations": self.db.query(SourceMappingSchemaExpectation)
            .filter(SourceMappingSchemaExpectation.source_id == source.id)
            .count(),
            "schemaAssessments": self.db.query(SourceSchemaAssessment)
            .filter(SourceSchemaAssessment.source_id == source.id)
            .count(),
            "businessEvents": self.db.query(BusinessEvent)
            .filter(
                BusinessEvent.primary_scope_type == "source",
                BusinessEvent.primary_scope_id == source.id,
            )
            .count(),
            "auditHistory": audit_history_count,
            "currencyProfiles": self.db.query(CurrencyProfile)
            .filter(
                CurrencyProfile.scope == "source",
                CurrencyProfile.scope_reference == source.id,
            )
            .count(),
            "sourceSnapshots": (
                self.db.query(DlSourceSnapshot)
                .filter(DlSourceSnapshot.connector_id == source.external_source_id)
                .count()
                if source.external_source_id
                else 0
            ),
        }
        protected_history = {
            key: count for key, count in protected_counts.items() if count > 0
        }
        blockers = {
            key: count
            for key, count in {
                "activeWorkspaces": len(
                    {
                        workspace.id
                        for _, workspace in matching_workspaces
                        if workspace.status == "active"
                    }
                    | active_workspace_ids
                ),
                "activeAcquisitionRuns": active_acquisition_count,
            }.items()
            if count > 0
        }
        action = (
            "none"
            if source.status != "active"
            else "blocked"
            if blockers
            else "archive"
            if protected_history
            else "delete"
        )
        return {
            "sourceId": source.id,
            "sourceName": source.name,
            "sourceVersion": source.version,
            "sourceStatus": source.status,
            "action": action,
            "blockers": blockers,
            "protectedHistory": protected_history,
            "archiveAllowed": source.status == "active" and not blockers,
            "permanentDeleteAllowed": not blockers,
            "permanentDeletePolicy": "immutable_history_tombstone",
        }

    def _disable_external_connector(self, source: SourceProfile) -> bool:
        """Retire only the connector bound to this Source in the same transaction."""
        if not source.external_source_id:
            return False
        connector = (
            self.db.query(IntegrationConnectorInstance)
            .filter(IntegrationConnectorInstance.id == source.external_source_id)
            .with_for_update()
            .one_or_none()
        )
        if connector is None:
            return False
        changed = connector.enabled or connector.status != "disabled"
        connector.enabled = False
        connector.status = "disabled"
        connector.updated_at = utcnow()
        return changed

    def _append_source_lifecycle_audit(
        self,
        *,
        event_type: str,
        user: FlowHubUser,
        reason: str,
        metadata: dict[str, Any],
    ) -> None:
        bus = DomainEventBus()
        bus.subscribe(PersistenceAuditSubscriber(self.db, _id))
        bus.publish(
            DomainEvent(
                event_type=event_type,
                correlation_id=f"source-lifecycle:{_id()}",
                user_id=user.id,
                attributes={"reason": reason, "metadata": metadata},
            )
        )

    def _ensure_channels(self) -> None:
        # Reuse the v1.2 connector capability registry.  The local import keeps
        # the Source module independent from the Workspace service at import time.
        from app.flowhub.unified_workspace.services import UnifiedWorkspaceService

        UnifiedWorkspaceService(self.db)._seed_channels()

    def _owned_sheet(self, sheet_id: str, user: FlowHubUser) -> FlowHubSheet:
        sheet = self.sheets.get(sheet_id)
        if sheet is None or (sheet.owner_user_id != user.id and user.role != "admin"):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Sheet not found.")
        return sheet

    def _source_shape(self, source: SourceProfile) -> dict[str, Any]:
        mapping = self.sources.latest_mapping(source.id)
        sheet = self.sheets.for_source(source.id)
        currency_profile = PricingMatrixService(self.db).unit_declaration(
            "source", source.id
        )
        return {
            "id": source.id,
            "name": source.name,
            "sourceKind": source.source_kind,
            "externalSourceId": source.external_source_id,
            "worksheetMode": source.worksheet_mode,
            "worksheetName": source.worksheet_name,
            "dataStartRow": source.data_start_row,
            "status": source.status,
            "archivedAt": _utc_timestamp(source.archived_at),
            "version": source.version,
            "mappingVersion": mapping.version if mapping else 0,
            "mappingReadiness": (
                self._mapping_readiness(
                    self._mapping_identity_validation_shape(mapping)
                )
                if mapping is not None
                else "not_configured"
            ),
            "sheetId": sheet.id if sheet else None,
            "createdAt": _utc_timestamp(source.created_at),
            "updatedAt": _utc_timestamp(source.updated_at),
            "currencyProfile": currency_profile,
        }

    @staticmethod
    def _mapping_readiness(identity_validation: dict[str, Any]) -> str:
        return {
            "pass": "ready",
            "blocked": "identity_validation_blocked",
            "pending": "identity_validation_pending",
        }[str(identity_validation["status"])]

    def _mapping_identity_validation_shape(
        self, mapping: SourceMappingRevision
    ) -> dict[str, Any]:
        references = self._identity_mapping_references(mapping)
        source = self.db.get(SourceProfile, mapping.source_id)
        if source is None:
            return self._pending_identity_validation(references)
        local_data = self._latest_local_validation_data(source)
        if local_data is None:
            return self._pending_identity_validation(references)
        evidence = dict(local_data["evidence"])
        if (
            int(mapping.identity_policy_version or 1) < 2
            or self._identity_authority_shape(mapping)["type"] == "unspecified"
        ):
            return self._pending_identity_validation(references, evidence)
        try:
            records = (
                self._mapped_external_records(local_data["worksheets"], mapping)
                if local_data["kind"] == "source_observation"
                else self._mapped_sheet_records(local_data["sheetRevision"], mapping)
            )
        except HTTPException as exc:
            detail = exc.detail if isinstance(exc.detail, dict) else {}
            if detail.get("code") != "WORKSHEET_NOT_FOUND":
                raise
            return self._pending_identity_validation(references, evidence)
        current_summary = self._identity_preview_summary(
            records,
            mapping=mapping,
            validation_source=evidence,
        )
        binding_context_fingerprint = str(
            current_summary.get("bindingContextFingerprint") or ""
        )
        assessment = (
            self.db.query(SourceMappingIdentityAssessment)
            .filter(
                SourceMappingIdentityAssessment.mapping_revision_id == mapping.id,
                SourceMappingIdentityAssessment.source_revision_kind
                == local_data["kind"],
                SourceMappingIdentityAssessment.source_revision_id
                == str(local_data["sourceRevisionId"]),
                SourceMappingIdentityAssessment.identity_fingerprint
                == self._identity_fingerprint(mapping),
                SourceMappingIdentityAssessment.binding_context_fingerprint
                == binding_context_fingerprint,
                SourceMappingIdentityAssessment.algorithm_version
                == IDENTITY_ASSESSMENT_ALGORITHM_VERSION,
            )
            .one_or_none()
        )
        if assessment is None:
            return self._pending_identity_validation(references, evidence)
        evidence["validatedAt"] = _iso_utc_timestamp(assessment.validated_at)
        return {
            "status": assessment.status,
            "participatingRowCount": assessment.participating_row_count,
            "validKeyCount": assessment.valid_key_count,
            "missingKeyCount": assessment.missing_key_count,
            "duplicateKeyCount": assessment.duplicate_key_count,
            "duplicateRowCount": assessment.duplicate_row_count,
            "bindingConflictCount": assessment.binding_conflict_count,
            "bindingContextFingerprint": assessment.binding_context_fingerprint,
            "missingRows": list(assessment.missing_rows_json or []),
            "duplicateGroups": list(assessment.duplicate_groups_json or []),
            "bindingConflicts": list(assessment.binding_conflicts_json or []),
            "mappingReferences": list(assessment.mapping_references_json or []),
            "evidence": evidence,
        }

    def _mapping_shape(self, revision: SourceMappingRevision | None) -> dict[str, Any] | None:
        if revision is None:
            return None
        source_fields = self.sources.source_fields(revision.id)
        channels = self.sources.channel_mappings(revision.id)
        channel_fields = self.sources.channel_fields([channel.id for channel in channels])
        by_channel: dict[str, list[SourceChannelFieldMapping]] = defaultdict(list)
        for channel_field in channel_fields:
            by_channel[channel_field.channel_mapping_id].append(channel_field)
        rule_set = self.sources.worksheet_rule_set(revision.id)
        worksheet_rules: list[dict[str, Any]] = []
        if rule_set is not None:
            stored_rules = self.sources.worksheet_rules(rule_set.id)
            stored_fields = self.sources.worksheet_fields([rule.id for rule in stored_rules])
            stored_channels = self.sources.worksheet_channels([rule.id for rule in stored_rules])
            stored_channel_fields = self.sources.worksheet_channel_fields(
                [channel.id for channel in stored_channels]
            )
            fields_by_rule: dict[str, list[SourceWorksheetFieldMapping]] = defaultdict(list)
            channels_by_rule: dict[str, list[SourceWorksheetChannelMapping]] = defaultdict(list)
            fields_by_stored_channel: dict[
                str, list[SourceWorksheetChannelFieldMapping]
            ] = defaultdict(list)
            for stored_field in stored_fields:
                fields_by_rule[stored_field.worksheet_rule_id].append(stored_field)
            for stored_channel in stored_channels:
                channels_by_rule[stored_channel.worksheet_rule_id].append(stored_channel)
            for stored_channel_field in stored_channel_fields:
                fields_by_stored_channel[
                    stored_channel_field.worksheet_channel_mapping_id
                ].append(stored_channel_field)
            worksheet_rules = [
                {
                    "worksheetName": stored_rule.worksheet_name,
                    "enabled": stored_rule.enabled,
                    "dataStartRow": stored_rule.data_start_row,
                    "valuePolicy": stored_rule.value_policy_json,
                    "sourceFields": [
                        {
                            "field": field.field,
                            "referenceType": field.reference_type,
                            "referenceValue": field.reference_value,
                            "required": field.required,
                        }
                        for field in fields_by_rule[stored_rule.id]
                    ],
                    "channels": [
                        {
                            "channelId": channel.channel_id,
                            "worksheetName": channel.worksheet_name,
                            "enabled": channel.enabled,
                            "fields": [
                                {
                                    "field": field.field,
                                    "referenceType": field.reference_type,
                                    "referenceValue": field.reference_value,
                                }
                                for field in fields_by_stored_channel[channel.id]
                            ],
                        }
                        for channel in channels_by_rule[stored_rule.id]
                    ],
                }
                for stored_rule in stored_rules
            ]
        identity_validation = self._mapping_identity_validation_shape(revision)
        return {
            "id": revision.id,
            "sourceId": revision.source_id,
            "version": revision.version,
            "checksum": revision.checksum,
            "worksheetMode": revision.worksheet_mode,
            "worksheetName": revision.worksheet_name,
            "dataStartRow": revision.data_start_row,
            "valuePolicy": revision.value_policy_json,
            "identityAuthority": self._identity_authority_shape(revision),
            "identityPolicyVersion": int(revision.identity_policy_version or 1),
            "identityValidation": identity_validation,
            "mappingReadiness": self._mapping_readiness(identity_validation),
            "worksheetRuleMode": rule_set.mode if rule_set else "shared",
            "selectedWorksheetNames": (
                sorted(
                    rule["worksheetName"]
                    for rule in worksheet_rules
                    if rule["worksheetName"] != "*" and rule["enabled"]
                )
                if rule_set is not None and rule_set.mode == "shared"
                else []
            ),
            "duplicateProductPolicy": (
                rule_set.duplicate_product_policy if rule_set else "block"
            ),
            "worksheetRules": worksheet_rules,
            "sourceFields": [
                {
                    "field": source_field.field,
                    "referenceType": source_field.reference_type,
                    "referenceValue": source_field.reference_value,
                    "required": source_field.required,
                }
                for source_field in source_fields
            ],
            "channels": [
                {
                    "channelId": channel_mapping.channel_id,
                    "worksheetName": channel_mapping.worksheet_name,
                    "enabled": channel_mapping.enabled,
                    "fields": [
                        {
                            "field": field.field,
                            "referenceType": field.reference_type,
                            "referenceValue": field.reference_value,
                        }
                        for field in by_channel[channel_mapping.id]
                    ],
                }
                for channel_mapping in channels
            ],
            "createdAt": revision.created_at,
        }

    def _legacy_mapping_shape(self, source: SourceProfile) -> dict[str, Any] | None:
        """Expose the historical global mapping as WooCommerce-only compatibility data.

        It is never persisted into the revisioned Source mapping automatically. The
        user must review and save the prefilled mapping explicitly.
        """
        if source.source_kind != "external" or source.external_source_id != LEGACY_EXTERNAL_SOURCE_ID:
            return None
        raw = AppConfigService(self.db).get("nextcloud.source_mapping")
        if not raw:
            return None
        try:
            parsed = json.loads(raw)
        except (TypeError, ValueError):
            return None
        legacy = normalize_source_mapping(parsed)
        field_names = {"id": "external_id", "price": "price", "stock": "stock"}
        fields: list[dict[str, Any]] = []
        for legacy_field, target_field in field_names.items():
            item = legacy[legacy_field]
            value = str(item.get("column") or "").strip()
            fields.append(
                {
                    "field": target_field,
                    "referenceType": (
                        "disabled"
                        if not item.get("enabled")
                        else "column_letter"
                        if re.fullmatch(r"[A-Za-z]{1,3}", value)
                        else "header_name"
                    ),
                    "referenceValue": value or None,
                    "required": False,
                }
            )
        fields.append(
            {
                "field": "status",
                "referenceType": "disabled",
                "referenceValue": None,
                "required": False,
            }
        )
        return {
            "primaryChannelId": "woocommerce:primary",
            "fields": fields,
            "requiresConfirmation": True,
        }

    def _invalidate_source_reviews(self, source_id: str) -> None:
        """Invalidate prepared state tied to an older active Source mapping."""
        snapshots = (
            self.db.query(WorkspaceSnapshot)
            .filter(WorkspaceSnapshot.entry_point == "source")
            .all()
        )
        snapshot_ids = [
            item.id
            for item in snapshots
            if str((item.source_metadata_json or {}).get("source_id") or "") == source_id
        ]
        if not snapshot_ids:
            return
        now = utcnow()
        reviews = (
            self.db.query(Review)
            .filter(Review.snapshot_id.in_(snapshot_ids), Review.status != ReviewState.STALE)
            .all()
        )
        review_ids: list[str] = []
        for review in reviews:
            review.status = ReviewState.STALE
            review.invalidated_at = now
            review.stale_reason = "source_mapping_revision_changed"
            review.selection_checksum = None
            review_ids.append(review.id)
        if review_ids:
            for job in (
                self.db.query(ApplyJob)
                .filter(
                    ApplyJob.review_id.in_(review_ids),
                    ApplyJob.status == ApplyState.PENDING,
                )
                .all()
            ):
                job.status = ApplyState.STALE
                job.completed_at = now

    def _is_nextcloud_connector(self, connector_id: str | None) -> bool:
        if not connector_id:
            return False
        if connector_id == LEGACY_EXTERNAL_SOURCE_ID:
            return True
        connector = self.db.get(IntegrationConnectorInstance, connector_id)
        return bool(connector and connector.connector_type == "nextcloud")

    @staticmethod
    def _normalize_identity_authority(
        raw: dict[str, Any] | None,
        *,
        required: bool,
    ) -> dict[str, Any]:
        values = dict(raw or {})
        authority_type = str(values.get("type") or "unspecified").strip()
        system_identifier = str(
            values.get("system_identifier")
            or values.get("systemIdentifier")
            or ""
        ).strip()
        display_label = str(
            values.get("display_label") or values.get("displayLabel") or ""
        ).strip()
        if authority_type not in {
            "external_system",
            "internal",
            "custom",
            "unspecified",
        }:
            raise _unprocessable(
                "SOURCE_IDENTITY_AUTHORITY_INVALID",
                "Choose a supported identity authority type.",
            )
        if authority_type == "unspecified":
            if required:
                raise _unprocessable(
                    "SOURCE_IDENTITY_AUTHORITY_REQUIRED",
                    "Choose the system that owns the Source Product Key.",
                )
            return dict(UNSPECIFIED_IDENTITY_AUTHORITY)
        if not system_identifier or not re.fullmatch(
            r"[A-Za-z0-9][A-Za-z0-9._:-]{0,119}", system_identifier
        ):
            raise _unprocessable(
                "SOURCE_IDENTITY_AUTHORITY_INVALID",
                "Identity authority requires a stable system identifier.",
            )
        if len(display_label) > 160:
            raise _unprocessable(
                "SOURCE_IDENTITY_AUTHORITY_INVALID",
                "Identity authority display label is too long.",
            )
        return {
            "type": authority_type,
            "systemIdentifier": system_identifier.casefold(),
            "displayLabel": display_label or None,
        }

    @staticmethod
    def _identity_authority_shape(
        mapping: SourceMappingRevision,
    ) -> dict[str, Any]:
        raw = dict(mapping.identity_authority_json or {})
        authority_type = str(raw.get("type") or "unspecified")
        if authority_type == "unspecified":
            return dict(UNSPECIFIED_IDENTITY_AUTHORITY)
        return {
            "type": authority_type,
            "systemIdentifier": raw.get("systemIdentifier")
            or raw.get("system_identifier"),
            "displayLabel": raw.get("displayLabel") or raw.get("display_label"),
        }

    @staticmethod
    def _validate_worksheet(mode: str, name: str | None, data_start_row: int) -> None:
        if mode not in {"all", "selected"}:
            raise _unprocessable("WORKSHEET_MODE_INVALID", "Use all or selected worksheet mode.")
        if mode == "selected" and not str(name or "").strip():
            raise _unprocessable("WORKSHEET_REQUIRED", "Select a worksheet.")
        if data_start_row < 1 or data_start_row > 1_000_000:
            raise _unprocessable("DATA_START_ROW_INVALID", "Data start row is outside the valid range.")

    @staticmethod
    def _normalize_selected_worksheet_names(names: list[str]) -> list[str]:
        normalized: list[str] = []
        seen: set[str] = set()
        for raw_name in names:
            name = str(raw_name or "").strip()
            if not name or name == "*" or len(name) > 240 or name in seen:
                raise _unprocessable(
                    "WORKSHEET_SELECTION_INVALID",
                    "Selected worksheet names must be explicit, unique, and at most 240 characters.",
                )
            normalized.append(name)
            seen.add(name)
        return normalized

    def _normalize_field_mappings(
        self,
        mappings: list[dict[str, Any]],
        allowed_fields: set[str],
        *,
        required_fields: set[str] | None = None,
    ) -> list[dict[str, Any]]:
        seen: set[str] = set()
        normalized: list[dict[str, Any]] = []
        for raw in mappings:
            field = str(raw.get("field") or "").strip()
            reference_type = str(raw.get("reference_type") or raw.get("referenceType") or "disabled")
            reference_value = str(raw.get("reference_value") or raw.get("referenceValue") or "").strip() or None
            if field not in allowed_fields or field in seen:
                raise _unprocessable("FIELD_MAPPING_INVALID", "Field mappings must be unique and supported.")
            if reference_type not in REFERENCE_TYPES:
                raise _unprocessable("REFERENCE_TYPE_INVALID", "Unsupported column reference type.")
            if reference_type != "disabled" and not reference_value:
                raise _unprocessable("REFERENCE_VALUE_REQUIRED", "Configured fields require a column reference.")
            if reference_type == "column_letter" and reference_value and not re.fullmatch(r"[A-Za-z]{1,3}", reference_value):
                raise _unprocessable("COLUMN_LETTER_INVALID", "Column letter is invalid.")
            seen.add(field)
            normalized.append(
                {
                    "field": field,
                    "referenceType": reference_type,
                    "referenceValue": reference_value.upper() if reference_type == "column_letter" and reference_value else reference_value,
                    "required": bool(raw.get("required") or field in (required_fields or set())),
                }
            )
        for required in required_fields or set():
            configured = next((item for item in normalized if item["field"] == required), None)
            if configured is None or configured["referenceType"] == "disabled":
                raise _unprocessable("SOURCE_IDENTITY_REQUIRED", f"Source Product {required} must be mapped.")
        return sorted(normalized, key=lambda item: item["field"])

    def _normalize_channel_mappings(
        self,
        mappings: list[dict[str, Any]],
        *,
        require_enabled: bool = True,
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        enabled_count = 0
        for raw in mappings:
            channel_id = str(raw.get("channel_id") or raw.get("channelId") or "").strip()
            if not channel_id or channel_id in seen:
                raise _unprocessable("CHANNEL_MAPPING_INVALID", "Channel mappings must be unique.")
            channel = self.db.get(WorkspaceChannel, channel_id)
            enabled = bool(raw.get("enabled", True))
            if channel is None:
                raise _unprocessable("CHANNEL_UNKNOWN", "The Channel mapping identity is unknown.")
            if enabled and (not channel.enabled or channel.implementation_state != "implemented"):
                raise _unprocessable(
                    "CHANNEL_UNAVAILABLE", "Only enabled Channels with official connectors may be mapped."
                )
            fields = self._normalize_field_mappings(list(raw.get("fields") or []), CHANNEL_FIELDS)
            fields_by_name = {item["field"]: item for item in fields}
            required_fields = self._channel_mapping_required_fields(channel)
            external = fields_by_name.get("external_id")
            if enabled and "external_id" in required_fields and (
                external is None or external["referenceType"] == "disabled"
            ):
                raise _unprocessable(
                    "CHANNEL_EXTERNAL_ID_REQUIRED", "This enabled Channel requires a Product Identifier mapping."
                )
            if enabled and any(
                fields_by_name.get(field) is None
                or fields_by_name[field]["referenceType"] == "disabled"
                for field in set(required_fields) - {"external_id"}
            ):
                raise _unprocessable(
                    "CHANNEL_REQUIRED_FIELD_MISSING",
                    "This enabled Channel is missing a field required by its connector capability contract.",
                    {"requiredFields": sorted(required_fields)},
                )
            result.append(
                {
                    "channelId": channel_id,
                    "worksheetName": str(raw.get("worksheet_name") or raw.get("worksheetName") or "").strip() or None,
                    "enabled": enabled,
                    "fields": fields,
                }
            )
            enabled_count += int(enabled)
            seen.add(channel_id)
        if require_enabled and (not result or enabled_count == 0):
            raise _unprocessable("CHANNEL_MAPPING_REQUIRED", "Select at least one enabled Channel.")
        return sorted(result, key=lambda item: item["channelId"])

    def _normalize_worksheet_rules(
        self, rules: list[dict[str, Any]], *, require_source_key: bool = False
    ) -> list[dict[str, Any]]:
        if not rules:
            raise _unprocessable(
                "WORKSHEET_RULE_REQUIRED",
                "Configure at least one worksheet or use shared rules.",
            )
        seen: set[str] = set()
        normalized: list[dict[str, Any]] = []
        enabled_count = 0
        for raw in rules:
            worksheet_name = str(
                raw.get("worksheet_name") or raw.get("worksheetName") or ""
            ).strip()
            enabled = bool(raw.get("enabled", True))
            data_start_row = int(
                raw.get("data_start_row") or raw.get("dataStartRow") or 1
            )
            if not worksheet_name or worksheet_name == "*" or worksheet_name in seen:
                raise _unprocessable(
                    "WORKSHEET_RULE_INVALID",
                    "Worksheet rule names must be explicit and unique.",
                )
            self._validate_worksheet("selected", worksheet_name, data_start_row)
            source_fields = self._normalize_field_mappings(
                list(raw.get("source_fields") or raw.get("sourceFields") or []),
                SOURCE_FIELDS,
                required_fields=({"name", "source_key"} if require_source_key else {"name"}) if enabled else set(),
            )
            channels = self._normalize_channel_mappings(
                list(raw.get("channel_mappings") or raw.get("channels") or []),
                require_enabled=enabled,
            )
            normalized.append(
                {
                    "worksheetName": worksheet_name,
                    "enabled": enabled,
                    "dataStartRow": data_start_row,
                    "sourceFields": source_fields,
                    "channels": channels,
                    "valuePolicy": self._normalize_value_policy(
                        dict(raw.get("value_policy") or raw.get("valuePolicy") or {})
                    ),
                }
            )
            enabled_count += int(enabled)
            seen.add(worksheet_name)
        if enabled_count == 0:
            raise _unprocessable(
                "WORKSHEET_RULE_REQUIRED",
                "At least one worksheet must participate in Source processing.",
            )
        return normalized

    def _persist_worksheet_rule_set(
        self,
        *,
        revision: SourceMappingRevision,
        mode: str,
        duplicate_product_policy: str,
        rules: list[dict[str, Any]],
    ) -> None:
        rule_set = SourceWorksheetRuleSet(
            id=_id(),
            mapping_revision_id=revision.id,
            mode=mode,
            duplicate_product_policy=duplicate_product_policy,
            sealed=False,
        )
        self.db.add(rule_set)
        self.db.flush()
        for rule in rules:
            worksheet_rule = SourceWorksheetRule(
                id=_id(),
                rule_set_id=rule_set.id,
                worksheet_name=rule["worksheetName"],
                enabled=bool(rule["enabled"]),
                data_start_row=int(rule["dataStartRow"]),
                value_policy_json=dict(rule["valuePolicy"]),
            )
            self.db.add(worksheet_rule)
            self.db.flush()
            for field in rule["sourceFields"]:
                self.db.add(
                    SourceWorksheetFieldMapping(
                        id=_id(),
                        worksheet_rule_id=worksheet_rule.id,
                        field=field["field"],
                        reference_type=field["referenceType"],
                        reference_value=field["referenceValue"],
                        required=bool(field["required"]),
                    )
                )
            for channel in rule["channels"]:
                worksheet_channel = SourceWorksheetChannelMapping(
                    id=_id(),
                    worksheet_rule_id=worksheet_rule.id,
                    channel_id=channel["channelId"],
                    worksheet_name=channel.get("worksheetName"),
                    enabled=bool(channel["enabled"]),
                )
                self.db.add(worksheet_channel)
                self.db.flush()
                for field in channel["fields"]:
                    self.db.add(
                        SourceWorksheetChannelFieldMapping(
                            id=_id(),
                            worksheet_channel_mapping_id=worksheet_channel.id,
                            field=field["field"],
                            reference_type=field["referenceType"],
                            reference_value=field["referenceValue"],
                        )
                    )
        # Flush the complete aggregate while its construction window is open,
        # then perform the sole permitted parent update to seal it forever.
        self.db.flush()
        rule_set.sealed = True
        self.db.flush()

    @staticmethod
    def _normalize_value_policy(raw: dict[str, Any]) -> dict[str, Any]:
        allowed = {
            "blank": {"no_change", "blocked"},
            "x": {"unavailable", "no_change", "blocked"},
            "dash": {"no_change", "unavailable", "blocked"},
            "zero": {"explicit_zero", "no_change", "blocked"},
            "formula": {"calculated_value", "blocked"},
            "invalid": {"blocked"},
        }
        result: dict[str, Any] = dict(DEFAULT_VALUE_POLICY)
        for key, value in raw.items():
            if key == "channel_price_policies":
                if not isinstance(value, dict):
                    raise _unprocessable("VALUE_POLICY_INVALID", "Channel price policies must be an object.")
                normalized_channels: dict[str, dict[str, Any]] = {}
                for channel_id, policy in value.items():
                    if not isinstance(policy, dict):
                        raise _unprocessable("VALUE_POLICY_INVALID", "Each Channel price policy must be an object.")
                    fix = policy.get("fix_zero_decimal_prices")
                    applicability = policy.get("fix_zero_decimal_prices_applicability")
                    precision_version = policy.get("monetary_precision_contract_version")
                    if fix is not None and not isinstance(fix, bool):
                        raise _unprocessable("VALUE_POLICY_INVALID", "Fix zero-decimal prices must be Boolean.")
                    if applicability not in {None, "APPLICABLE", "NOT_APPLICABLE"}:
                        raise _unprocessable("VALUE_POLICY_INVALID", "Invalid zero-decimal applicability.")
                    if precision_version is not None and not isinstance(precision_version, str):
                        raise _unprocessable("VALUE_POLICY_INVALID", "Invalid monetary precision contract version.")
                    normalized_channels[str(channel_id)] = {
                        "fix_zero_decimal_prices": fix,
                        "fix_zero_decimal_prices_applicability": applicability,
                        "monetary_precision_contract_version": precision_version,
                    }
                result[key] = normalized_channels
                continue
            if key not in allowed or value not in allowed[key]:
                raise _unprocessable("VALUE_POLICY_INVALID", f"Invalid handling policy for {key}.")
            result[key] = value
        return result

    @staticmethod
    def _normalize_columns(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not columns or len(columns) > MAX_SHEET_COLUMNS:
            raise _unprocessable("SHEET_COLUMNS_INVALID", f"Use 1 to {MAX_SHEET_COLUMNS} columns.")
        keys: set[str] = set()
        positions: set[int] = set()
        normalized = []
        for index, item in enumerate(columns, start=1):
            key = str(item.get("column_key") or item.get("columnKey") or _id()).strip()
            position = int(item.get("position") or index)
            if not key or len(key) > 36 or key in keys or position in positions or position < 1:
                raise _unprocessable("SHEET_COLUMN_INVALID", "Column identities and positions must be unique.")
            keys.add(key)
            positions.add(position)
            normalized.append(
                {
                    "columnKey": key,
                    "name": _clean_name(item.get("name"), f"Column {column_name(position)}"),
                    "position": position,
                    "dataType": str(item.get("data_type") or item.get("dataType") or "text")[:30],
                }
            )
        return sorted(normalized, key=lambda item: item["position"])

    @staticmethod
    def _normalize_rows(
        rows: list[dict[str, Any]], columns: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        if len(rows) > MAX_SHEET_ROWS:
            raise _unprocessable("SHEET_ROW_LIMIT", f"A Sheet supports at most {MAX_SHEET_ROWS} rows.")
        column_keys = {item["columnKey"] for item in columns}
        keys: set[str] = set()
        positions: set[int] = set()
        normalized = []
        for index, item in enumerate(rows, start=1):
            key = str(item.get("row_key") or item.get("rowKey") or _id()).strip()
            position = int(item.get("position") or index)
            values = dict(item.get("values") or {})
            if key in keys or position in positions or position < 1:
                raise _unprocessable("SHEET_ROW_INVALID", "Row identities and positions must be unique.")
            unknown = set(values) - column_keys
            if unknown:
                raise _unprocessable("SHEET_CELL_COLUMN_INVALID", "A cell references an unknown column.")
            keys.add(key)
            positions.add(position)
            normalized.append(
                {"rowKey": key, "position": position, "values": values}
            )
        return sorted(normalized, key=lambda item: item["position"])

    @staticmethod
    def _calculate(
        columns: list[dict[str, Any]], rows: list[dict[str, Any]]
    ) -> dict[str, FormulaResult]:
        positions = {item["columnKey"]: item["position"] for item in columns}
        values = {
            f"{column_name(positions[column_key])}{row['position']}": None if value is None else str(value)
            for row in rows
            for column_key, value in row["values"].items()
        }
        return calculate_sheet(values)

    @staticmethod
    def _column_shape(column: SheetColumn) -> dict[str, Any]:
        return {
            "columnKey": column.column_key,
            "name": column.name,
            "position": column.position,
            "dataType": column.data_type,
        }

    @staticmethod
    def _decode_import(content_base64: str) -> bytes:
        try:
            content = base64.b64decode(content_base64, validate=True)
        except ValueError as exc:
            raise _unprocessable("IMPORT_ENCODING_INVALID", "Import content is not valid base64.") from exc
        if not content or len(content) > MAX_IMPORT_BYTES:
            raise _unprocessable("IMPORT_SIZE_INVALID", "Import file is empty or exceeds 20 MB.")
        return content

    @staticmethod
    def _read_import(filename: str, content: bytes) -> dict[str, list[list[Any]]]:
        lowered = filename.lower()
        if lowered.endswith(".csv"):
            decoded = None
            for encoding in ("utf-8-sig", "utf-8", "cp1256"):
                try:
                    decoded = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if decoded is None:
                raise _unprocessable("CSV_ENCODING_INVALID", "CSV must use a supported text encoding.")
            return {"Sheet1": [list(row) for row in csv.reader(io.StringIO(decoded))]}
        if lowered.endswith(".xlsx"):
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False, read_only=True)
            except Exception as exc:
                raise _unprocessable("XLSX_INVALID", "XLSX file could not be read.") from exc
            return {
                worksheet.title: [list(row) for row in worksheet.iter_rows(values_only=True)]
                for worksheet in workbook.worksheets
            }
        raise _unprocessable("IMPORT_FORMAT_UNSUPPORTED", "Use an XLSX or CSV file.")

    def _worksheet_rule_configs(
        self, mapping: SourceMappingRevision
    ) -> tuple[str, str, list[dict[str, Any]]]:
        """Load normalized rules, synthesizing a shared rule for FLOWHUB_018 data."""
        rule_set = self.sources.worksheet_rule_set(mapping.id)
        if rule_set is None:
            legacy_source_fields = self.sources.source_fields(mapping.id)
            legacy_channels = self.sources.channel_mappings(mapping.id)
            legacy_channel_fields = self.sources.channel_fields(
                [channel.id for channel in legacy_channels]
            )
            by_channel: dict[str, list[SourceChannelFieldMapping]] = defaultdict(list)
            for channel_field in legacy_channel_fields:
                by_channel[channel_field.channel_mapping_id].append(channel_field)
            return (
                "shared",
                "block",
                [
                    {
                        "worksheetName": "*",
                        "enabled": True,
                        "dataStartRow": mapping.data_start_row,
                        "valuePolicy": dict(mapping.value_policy_json),
                        "sourceFields": legacy_source_fields,
                        "channels": [
                            {
                                "channelId": item.channel_id,
                                "worksheetName": item.worksheet_name,
                                "enabled": item.enabled,
                                "fields": by_channel[item.id],
                            }
                            for item in legacy_channels
                        ],
                    }
                ],
            )
        rules = self.sources.worksheet_rules(rule_set.id)
        worksheet_source_fields = self.sources.worksheet_fields([rule.id for rule in rules])
        worksheet_channels = self.sources.worksheet_channels([rule.id for rule in rules])
        worksheet_channel_fields = self.sources.worksheet_channel_fields(
            [channel.id for channel in worksheet_channels]
        )
        source_fields_by_rule: dict[str, list[SourceWorksheetFieldMapping]] = defaultdict(list)
        channels_by_rule: dict[str, list[SourceWorksheetChannelMapping]] = defaultdict(list)
        channel_fields_by_mapping: dict[
            str, list[SourceWorksheetChannelFieldMapping]
        ] = defaultdict(list)
        for worksheet_source_field in worksheet_source_fields:
            source_fields_by_rule[worksheet_source_field.worksheet_rule_id].append(
                worksheet_source_field
            )
        for worksheet_channel in worksheet_channels:
            channels_by_rule[worksheet_channel.worksheet_rule_id].append(
                worksheet_channel
            )
        for worksheet_channel_field in worksheet_channel_fields:
            channel_fields_by_mapping[
                worksheet_channel_field.worksheet_channel_mapping_id
            ].append(worksheet_channel_field)
        return (
            rule_set.mode,
            rule_set.duplicate_product_policy,
            [
                {
                    "worksheetName": rule.worksheet_name,
                    "enabled": rule.enabled,
                    "dataStartRow": rule.data_start_row,
                    "valuePolicy": dict(rule.value_policy_json),
                    "sourceFields": source_fields_by_rule[rule.id],
                    "channels": [
                        {
                            "channelId": channel.channel_id,
                            "worksheetName": channel.worksheet_name,
                            "enabled": channel.enabled,
                            "fields": channel_fields_by_mapping[channel.id],
                        }
                        for channel in channels_by_rule[rule.id]
                    ],
                }
                for rule in rules
            ],
        )

    @staticmethod
    def _apply_cross_worksheet_duplicate_policy(
        records: list[dict[str, Any]], policy: str
    ) -> None:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            source = record.get("sourceProduct") or {}
            identity = _normalize_source_product_key(
                source.get("source_key")
                or ("" if record.get("sourceKeyRequired") else source.get("name"))
            )
            if identity:
                grouped[identity].append(record)
        for matches in grouped.values():
            if len(matches) < 2:
                continue
            authoritative_key = any(bool(item.get("sourceKeyRequired")) for item in matches)
            if not authoritative_key and len({str(item.get("worksheetName") or "") for item in matches}) < 2:
                continue
            if not authoritative_key and policy == "last_sheet_wins":
                for record in matches[:-1]:
                    record["recognized"] = False
                    record["channels"] = []
                    record["issues"].append({
                        "category": "duplicate_source_product_superseded", "severity": "warning", "channelId": None,
                        "message": "A later participating worksheet explicitly replaces this Source Product.",
                    })
                continue
            references = [f"{item.get('worksheetName')}!{item.get('rowNumber')}" for item in matches]
            key_value = str(matches[0].get("sourceProduct", {}).get("source_key") or "").strip()
            for record in matches:
                record["recognized"] = False
                record["channels"] = []
                record["issues"].append(
                    {
                        "category": "duplicate_source_product_key" if authoritative_key else "duplicate_source_product",
                        "severity": "blocked",
                        "channelId": None,
                        "message": "Source Product Key must be unique within this Source.",
                        "details": {
                            "conflictingRows": references,
                            # Source keys are intentionally chosen by the owner as
                            # the canonical identity and are already visible in
                            # this authenticated preview. Showing it here makes
                            # the duplicate repair actionable without exposing
                            # unrelated Source fields.
                            "keyValue": key_value[:240],
                        },
                    }
                )

    def _mapped_external_records(
        self,
        worksheets: dict[str, list[list[Any]]],
        mapping: SourceMappingRevision,
    ) -> list[dict[str, Any]]:
        """Resolve an acquired workbook once into independent Channel values."""
        if not worksheets:
            return []
        rule_mode, duplicate_policy, configured_rules = self._worksheet_rule_configs(mapping)
        rule_work: list[tuple[str, dict[str, Any]]] = []
        if rule_mode == "shared":
            explicit_rules = {
                item["worksheetName"]: item
                for item in configured_rules
                if item["worksheetName"] != "*"
            }
            if explicit_rules:
                missing = [
                    name
                    for name, rule in explicit_rules.items()
                    if rule["enabled"] and name not in worksheets
                ]
                if missing:
                    raise _unprocessable(
                        "WORKSHEET_NOT_FOUND",
                        "A selected worksheet is not present in the acquired workbook.",
                        {"worksheets": missing},
                    )
                rule_work = [
                    (name, explicit_rules[name])
                    for name in worksheets
                    if name in explicit_rules and explicit_rules[name]["enabled"]
                ]
            else:
                # FLOWHUB_018 and early FLOWHUB_019 shared mappings use a
                # wildcard rule plus the legacy all/single worksheet columns.
                worksheet_names = (
                    [str(mapping.worksheet_name or "")]
                    if mapping.worksheet_mode == "selected"
                    else list(worksheets)
                )
                missing = [name for name in worksheet_names if name not in worksheets]
                if missing:
                    raise _unprocessable(
                        "WORKSHEET_NOT_FOUND",
                        "The configured Source worksheet is not present in the acquired workbook.",
                        {"worksheets": missing},
                    )
                rule_work = [(name, configured_rules[0]) for name in worksheet_names]
        else:
            rules_by_name = {item["worksheetName"]: item for item in configured_rules}
            missing = [
                name
                for name, rule in rules_by_name.items()
                if rule["enabled"] and name not in worksheets
            ]
            if missing:
                raise _unprocessable(
                    "WORKSHEET_NOT_FOUND",
                    "A configured worksheet is not present in the acquired workbook.",
                    {"worksheets": missing},
                )
            rule_work = [
                (name, rules_by_name[name])
                for name in worksheets
                if name in rules_by_name and rules_by_name[name]["enabled"]
            ]
        channel_ids = {
            str(channel["channelId"])
            for _, rule in rule_work
            for channel in rule["channels"]
        }
        current_channels = {
            item.id: item
            for item in self.db.query(WorkspaceChannel)
            .filter(WorkspaceChannel.id.in_(channel_ids))
            .all()
        }
        header_cache: dict[tuple[str, int, str], int | None] = {}

        def column_index(
            worksheet: str,
            data_start_row: int,
            reference_type: str,
            reference_value: str | None,
        ) -> int | None:
            if reference_type == "column_letter":
                try:
                    return int(
                        openpyxl.utils.column_index_from_string(str(reference_value or ""))
                    ) - 1
                except ValueError:
                    return None
            if reference_type != "header_name":
                return None
            normalized = str(reference_value or "").strip().casefold()
            cache_key = (worksheet, data_start_row, normalized)
            if cache_key in header_cache:
                return header_cache[cache_key]
            found = None
            header_rows = worksheets.get(worksheet, [])[: max(data_start_row - 1, 0)]
            for header_row in reversed(header_rows):
                for index, value in enumerate(header_row):
                    if str(value or "").strip().casefold() == normalized:
                        found = index
                        break
                if found is not None:
                    break
            header_cache[cache_key] = found
            return found

        def read(
            worksheet: str,
            row_number: int,
            data_start_row: int,
            reference_type: str,
            reference_value: str | None,
        ) -> Any:
            index = column_index(
                worksheet, data_start_row, reference_type, reference_value
            )
            rows = worksheets.get(worksheet, [])
            if index is None or row_number < 1 or row_number > len(rows):
                return None
            row = rows[row_number - 1]
            return row[index] if index < len(row) else None

        records: list[dict[str, Any]] = []
        for worksheet_name, rule in rule_work:
            data_start_row = int(rule["dataStartRow"])
            policy = dict(DEFAULT_VALUE_POLICY) | dict(rule["valuePolicy"])
            for row_number in range(data_start_row, len(worksheets[worksheet_name]) + 1):
                row_issues: list[dict[str, str | None]] = []
                source_data = {
                    field.field: read(
                        worksheet_name,
                        row_number,
                        data_start_row,
                        field.reference_type,
                        field.reference_value,
                    )
                    for field in rule["sourceFields"]
                    if field.reference_type != "disabled"
                }
                source_key = str(source_data.get("source_key") or "").strip()
                require_source_key = any(
                    field.field == "source_key" and field.required
                    for field in rule["sourceFields"]
                )
                channel_data: list[dict[str, Any]] = []
                channel_value_present = False
                for channel in rule["channels"]:
                    channel_id = str(channel["channelId"])
                    current = current_channels.get(channel_id)
                    if (
                        not channel["enabled"]
                        or current is None
                        or not current.enabled
                        or current.implementation_state != "implemented"
                    ):
                        continue
                    channel_worksheet = channel.get("worksheetName") or worksheet_name
                    if channel_worksheet not in worksheets:
                        row_issues.append(
                            {
                                "category": "missing_channel_worksheet",
                                "severity": "blocked",
                                "channelId": channel_id,
                                "message": "The worksheet selected for this Channel is unavailable.",
                            }
                        )
                        continue
                    fields = {
                        item.field: read(
                            channel_worksheet,
                            row_number,
                            data_start_row,
                            item.reference_type,
                            item.reference_value,
                        )
                        for item in channel["fields"]
                        if item.reference_type != "disabled"
                    }
                    if any(value not in {None, ""} for value in fields.values()):
                        channel_value_present = True
                    # Channel identifiers are identity, never business-value
                    # sentinels.  x, dash, and zero reach the connector's
                    # identifier contract unchanged.
                    external_id = str(fields.get("external_id") or "").strip()
                    if external_id:
                        channel_data.append({"channelId": channel_id, "fields": fields})
                    elif any(value not in {None, ""} for value in fields.values()):
                        row_issues.append(
                            {
                                "category": "missing_mapping_identity",
                                "severity": "blocked",
                                "channelId": channel_id,
                                "message": "Channel values exist but External Listing ID is missing.",
                            }
                        )
                non_display_source = {
                    key: value for key, value in source_data.items() if key != "name"
                }
                row_has_product_data = (
                    any(value not in {None, ""} for value in non_display_source.values())
                    or channel_value_present
                )
                recognized = bool((source_key or not require_source_key) and channel_data)
                if require_source_key and not source_key and row_has_product_data:
                    row_issues.append(
                        {
                            "category": "missing_source_product_key",
                            "severity": "blocked",
                            "channelId": None,
                            "message": "Source Product Key is required.",
                            "details": {"row": f"{worksheet_name}!{row_number}"},
                        }
                    )
                records.append(
                    {
                        "rowKey": f"external:{worksheet_name}:{row_number}",
                        "rowNumber": row_number,
                        "worksheetName": worksheet_name,
                        "recognized": recognized,
                        "sourceKeyRequired": require_source_key,
                        "sourceProduct": source_data,
                        "channels": channel_data,
                        "valuePolicy": policy,
                        "issues": row_issues,
                    }
                )
        self._apply_cross_worksheet_duplicate_policy(records, duplicate_policy)
        return records

    def _mapped_sheet_records(
        self, revision: SheetRevision, mapping: SourceMappingRevision
    ) -> list[dict[str, Any]]:
        columns = self.sheets.columns(revision.id)
        rows = self.sheets.all_rows(revision.id)
        cells = self.sheets.cells(revision.id, [item.id for item in rows])
        column_by_key = {item.column_key: item for item in columns}
        by_name = {item.name.casefold(): item.column_key for item in columns}
        by_letter = {column_name(item.position): item.column_key for item in columns}
        by_row: dict[str, dict[str, SheetCell]] = defaultdict(dict)
        for cell in cells:
            by_row[cell.row_id][cell.column_key] = cell

        def mapped_key(reference_type: str, reference_value: str | None) -> str | None:
            if reference_type == "column_id":
                return reference_value if reference_value in column_by_key else None
            if reference_type == "column_letter":
                return by_letter.get(str(reference_value or "").upper())
            if reference_type == "header_name":
                return by_name.get(str(reference_value or "").casefold())
            return None

        rule_mode, _, configured_rules = self._worksheet_rule_configs(mapping)
        selected_rule: dict[str, Any] | None = None
        if rule_mode == "shared":
            internal_name = str(mapping.worksheet_name or "Sheet1")
            selected_rule = next(
                (
                    configured_rule
                    for configured_rule in configured_rules
                    if configured_rule["worksheetName"] in {"*", internal_name}
                    and configured_rule["enabled"]
                ),
                None,
            )
        else:
            internal_name = str(mapping.worksheet_name or "Sheet1")
            for configured_rule in configured_rules:
                if (
                    configured_rule["worksheetName"] == internal_name
                    and configured_rule["enabled"]
                ):
                    selected_rule = configured_rule
                    break
        if selected_rule is None:
            return []
        rule = selected_rule
        source_fields = rule["sourceFields"]
        channels = rule["channels"]
        channel_ids = [str(item["channelId"]) for item in channels]
        current_channels = {
            item.id: item
            for item in self.db.query(WorkspaceChannel)
            .filter(WorkspaceChannel.id.in_(channel_ids))
            .all()
        }
        records: list[dict[str, Any]] = []
        policy = dict(DEFAULT_VALUE_POLICY) | dict(rule["valuePolicy"])
        for row in rows:
            if row.position < int(rule["dataStartRow"]):
                continue
            values = by_row.get(row.id, {})
            row_issues: list[dict[str, str | None]] = []

            def read(
                reference_type: str,
                reference_value: str | None,
                row_values: dict[str, SheetCell] = values,
                issues: list[dict[str, str | None]] = row_issues,
            ) -> Any:
                key = mapped_key(reference_type, reference_value)
                cell = row_values.get(key or "")
                if cell and cell.calculation_error:
                    issues.append(
                        {
                            "category": "formula_error",
                            "severity": "blocked",
                            "channelId": None,
                            "message": f"Formula calculation failed: {cell.calculation_error}.",
                        }
                    )
                elif cell and cell.formula_expression and policy["formula"] == "blocked":
                    issues.append(
                        {
                            "category": "formula_blocked",
                            "severity": "blocked",
                            "channelId": None,
                            "message": "Formula values are blocked by the Source policy.",
                        }
                    )
                return cell.calculated_value if cell else None

            source_data = {
                field.field: read(field.reference_type, field.reference_value)
                for field in source_fields
                if field.reference_type != "disabled"
            }
            source_key = str(source_data.get("source_key") or "").strip()
            require_source_key = any(
                field.field == "source_key" and field.required for field in source_fields
            )
            channel_data = []
            channel_value_present = False
            for channel in channels:
                channel_id = str(channel["channelId"])
                current = current_channels.get(channel_id)
                if (
                    not channel["enabled"]
                    or current is None
                    or not current.enabled
                    or current.implementation_state != "implemented"
                ):
                    continue
                fields = {
                    item.field: read(item.reference_type, item.reference_value)
                    for item in channel["fields"]
                    if item.reference_type != "disabled"
                }
                if any(value not in {None, ""} for value in fields.values()):
                    channel_value_present = True
                # Channel identifiers are identity, never business-value
                # sentinels.  x, dash, and zero reach the connector's
                # identifier contract unchanged.
                external_id = str(fields.get("external_id") or "").strip()
                if external_id:
                    channel_data.append({"channelId": channel_id, "fields": fields})
                elif any(value not in {None, ""} for value in fields.values()):
                    row_issues.append(
                        {
                            "category": "missing_mapping_identity",
                            "severity": "blocked",
                            "channelId": channel_id,
                            "message": "Channel values exist but External Listing ID is missing.",
                        }
                    )
            non_display_source = {
                key: value for key, value in source_data.items() if key != "name"
            }
            row_has_product_data = (
                any(value not in {None, ""} for value in non_display_source.values())
                or channel_value_present
            )
            recognized = bool((source_key or not require_source_key) and channel_data)
            if require_source_key and not source_key and row_has_product_data:
                row_issues.append(
                    {
                        "category": "missing_source_product_key",
                        "severity": "blocked",
                        "channelId": None,
                        "message": "Source Product Key is required.",
                        "details": {"row": str(row.position)},
                    }
                )
            records.append(
                {
                    "rowKey": row.row_key,
                    "rowNumber": row.position,
                    "worksheetName": str(mapping.worksheet_name or "Sheet1"),
                    "recognized": recognized,
                    "sourceKeyRequired": require_source_key,
                    "sourceProduct": source_data,
                    "channels": channel_data,
                    "valuePolicy": policy,
                    "issues": row_issues,
                }
            )
        self._apply_cross_worksheet_duplicate_policy(records, "block")
        return records

    @staticmethod
    def _interpret_target(
        raw: Any, field: str, policy: dict[str, str]
    ) -> dict[str, str | None]:
        if raw is None or str(raw).strip() == "":
            if policy["blank"] == "blocked":
                return {
                    "target": None,
                    "issue": "BLANK_VALUE_BLOCKED",
                    "message": f"Blank {field} is blocked by the Source policy.",
                }
            return {"target": None, "issue": None, "message": None}
        text = str(raw).strip()
        lowered = text.casefold()
        if lowered == "x":
            behavior = policy["x"]
            if behavior == "blocked":
                return {
                    "target": None,
                    "issue": "X_MARKER_BLOCKED",
                    "message": f"The x marker is not valid for {field}.",
                }
            return {"target": None, "issue": None, "message": None}
        if lowered in {"-", "–", "—"}:
            behavior = policy["dash"]
            if behavior == "blocked":
                return {
                    "target": None,
                    "issue": "DASH_MARKER_BLOCKED",
                    "message": f"The dash marker is not valid for {field}.",
                }
            return {"target": None, "issue": None, "message": None}
        if field in {"price", "stock"}:
            try:
                number = Decimal(text.replace(",", ""))
            except InvalidOperation:
                return {
                    "target": None,
                    "issue": "INVALID_NUMERIC_VALUE",
                    "message": f"{field.title()} must be a valid numeric value.",
                }
            if number < 0:
                return {
                    "target": None,
                    "issue": "NEGATIVE_VALUE",
                    "message": f"{field.title()} cannot be negative.",
                }
            if number == 0 and policy["zero"] != "explicit_zero":
                if policy["zero"] == "blocked":
                    return {
                        "target": None,
                        "issue": "ZERO_VALUE_BLOCKED",
                        "message": f"Zero {field} is blocked by the Source policy.",
                    }
                return {"target": None, "issue": None, "message": None}
            text = format(number.normalize(), "f")
        return {"target": text, "issue": None, "message": None}

    @staticmethod
    def _classify_channel_targets(
        fields: dict[str, Any],
        policy: dict[str, Any],
        channel: WorkspaceChannel,
        cache: ChannelCache,
    ) -> tuple[dict[str, str], dict[str, Any], list[dict[str, str]]]:
        """Translate Source cells once into exact targets plus immutable evidence.

        The mapping builder previously reduced every sentinel and malformed value
        to ``None``.  That made it impossible to distinguish no instruction from
        the Owner-approved direct-price availability instruction.  This small
        adapter keeps the pure business classification in ``domain.py`` and
        leaves connector capability/write decisions to Unified Workspace.
        """

        capabilities = dict(channel.capabilities_json or {})
        # Cache currency/unit is pinned Source-to-Preview evidence when older
        # Channel capability payloads predate those declarations.  A missing
        # value on both surfaces remains a blocker; it is never inferred from
        # price magnitude.
        currency = str(capabilities.get("currency") or cache.price_currency or "").upper() or None
        unit = str(capabilities.get("unit") or cache.price_unit or "").upper() or None
        channel_policy = dict(policy.get("channel_price_policies") or {}).get(channel.id, {})
        is_rial_or_toman = currency == "IRR" and unit in {"RIAL", "TOMAN"}
        # Current architecture has a declared zero-decimal IRR unit contract.
        # Other currencies must carry an explicit precision in channel
        # capability evidence before direct Source Price can be classified.
        precision = 0 if is_rial_or_toman else capabilities.get("monetaryPrecision")
        price = (
            normalize_direct_price(
                fields.get("price"),
                currency=currency,
                unit=unit,
                monetary_precision=precision if isinstance(precision, int) else None,
                fix_zero_decimal_prices=channel_policy.get("fix_zero_decimal_prices"),
            )
            if "price" in fields
            else None
        )
        quantity = normalize_quantity(fields.get("stock"), mapped="stock" in fields)
        stock_status = normalize_stock_status(fields.get("status"), mapped="status" in fields)
        normalized = {"price": price, "stock": quantity, "status": stock_status}
        desired, blockers = resolve_availability(*normalized.values())
        targets: dict[str, str] = {}
        warnings: list[dict[str, str]] = []
        for field, result in normalized.items():
            if result.warning_code:
                warnings.append({"code": result.warning_code, "field": field})
        if blockers:
            return targets, {
                "version": "workspace-change-badges-v1",
                "fields": {
                    field: {
                        "instruction": result.instruction.value,
                        "rawLexeme": result.raw_lexeme,
                        "target": result.target,
                        "availabilitySignal": result.availability_signal.value if result.availability_signal else None,
                        "reason": result.reason_code,
                        "warning": result.warning_code,
                        "blocker": result.blocker_code,
                        "fixApplied": result.fix_applied,
                    }
                    for field, result in normalized.items()
                },
                "desiredStockStatus": None,
                "blockers": list(blockers),
                "warnings": warnings,
            }, warnings
        if price and price.instruction is SourceInstruction.SET and price.target is not None:
            targets["price"] = price.target
        if (
            quantity.instruction is SourceInstruction.SET
            and quantity.target is not None
            and not (desired is AvailabilitySignal.OUT_OF_STOCK and quantity.target != "0")
        ):
            # Positive quantity is intentionally suppressed by a winning OOS
            # instruction. Zero remains an explicit governed instruction.
            targets["stock"] = quantity.target
        current_status = str(cache.status or "").casefold()
        current_canonical = {
            "instock": AvailabilitySignal.IN_STOCK,
            "outofstock": AvailabilitySignal.OUT_OF_STOCK,
        }.get(current_status)
        if desired is not None and current_canonical is not None and desired is not current_canonical:
            targets["status"] = desired.value
        return targets, {
            "version": "workspace-change-badges-v1",
            "fields": {
                field: {
                    "instruction": result.instruction.value,
                    "rawLexeme": result.raw_lexeme,
                    "target": result.target,
                    "availabilitySignal": result.availability_signal.value if result.availability_signal else None,
                    "reason": result.reason_code,
                    "warning": result.warning_code,
                    "blocker": result.blocker_code,
                    "fixApplied": result.fix_applied,
                }
                for field, result in normalized.items()
            },
            "desiredStockStatus": desired.value if desired else None,
            "blockers": [],
            "warnings": warnings,
        }, warnings

    @staticmethod
    def _candidate_issue(
        record: dict[str, Any],
        channel_id: str | None,
        category: str,
        code: str,
        summary: str,
        action: str,
        details: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "sourceRowKey": record["rowKey"],
            "sourceRowNumber": record["rowNumber"],
            "worksheetName": record.get("worksheetName"),
            "channelId": channel_id,
            "sourceProductName": str(
                record.get("sourceProduct", {}).get("name")
                or record.get("sourceProduct", {}).get("source_key")
                or ""
            )[:240]
            or None,
            "mappingState": (
                "unmapped"
                if category == "missing_mapping"
                else "conflict"
                if category == "mapping_conflict"
                else "resolved"
            ),
            "category": category,
            "severity": "blocked",
            "code": code,
            "summary": summary,
            "recommendedAction": action,
            "technicalDetails": details,
        }

    @staticmethod
    def _preview_issue_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        grouped: dict[tuple[str, str, str | None], int] = defaultdict(int)
        for record in records:
            for issue in record["issues"]:
                grouped[(issue["category"], issue["severity"], issue.get("channelId"))] += 1
        return [
            {"category": key[0], "severity": key[1], "channelId": key[2], "count": count}
            for key, count in sorted(grouped.items())
        ]

    def _preview_business_summary(
        self,
        records: list[dict[str, Any]],
        mapping: SourceMappingRevision,
    ) -> dict[str, int | None]:
        """Return seller-facing counts without inventing Channel comparisons.

        A Source Product can produce several Listing rows, so product totals use
        the stable Source identity when present and fall back to the Source row
        identity.  Attention is intentionally row-scoped because one problematic
        row must remain independently actionable even when another row names the
        same product.
        """

        def product_identity(record: dict[str, Any]) -> str:
            source_product = dict(record.get("sourceProduct") or {})
            identity = str(source_product.get("source_key") or "").strip()
            if identity:
                return f"product:{identity.casefold()}"
            return f"row:{record.get('rowKey') or record.get('rowNumber') or ''}"

        recognized_products = {
            product_identity(record) for record in records if record.get("recognized")
        }
        products_with_issues = {
            product_identity(record) for record in records if record.get("issues")
        }
        attention_rows = {
            str(
                record.get("rowKey")
                or f"{record.get('worksheetName') or ''}:{record.get('rowNumber') or ''}"
            )
            for record in records
            if record.get("issues")
        }

        _, _, configured_rules = self._worksheet_rule_configs(mapping)
        configured_channel_ids = {
            str(channel["channelId"])
            for rule in configured_rules
            if bool(rule.get("enabled", True))
            for channel in rule["channels"]
            if bool(channel["enabled"])
        }
        available_channel_ids = {
            str(channel_id)
            for (channel_id,) in self.db.query(WorkspaceChannel.id)
            .filter(
                WorkspaceChannel.enabled.is_(True),
                WorkspaceChannel.implementation_state == "implemented",
            )
            .all()
        }
        ready_channel_ids = configured_channel_ids & available_channel_ids

        return {
            "productsFound": len(recognized_products),
            "productsReady": len(recognized_products - products_with_issues),
            # Preview has not compared targets with each Channel cache.  Returning
            # zero here would falsely claim that no change exists.
            "priceChanges": None,
            "stockChanges": None,
            "unchanged": None,
            "needsAttention": len(attention_rows),
            "channelsReady": len(ready_channel_ids),
            "channelsNotConfigured": len(available_channel_ids - ready_channel_ids),
        }
