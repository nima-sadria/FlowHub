"""FlowHub - XLSX spreadsheet parser (BU5).

Adapted from the legacy spreadsheet parser.
Preserves legacy spreadsheet behavior:
  - ALL worksheets are read; last sheet wins on duplicate product IDs.
  - Columns: A = Product Name, B = Product ID (int), C = Price.
  - Row 3 onward (rows 1-2 are headers).
  - Stops after 30 consecutive empty rows or row 1002 (max 1000 data rows).
  - Persian / Arabic-Indic digits are normalised to ASCII.
  - Out-of-stock markers are treated as price=None (not an error).
  - Unparseable prices set price_parse_error=True, price=None (not fatal).
  - Duplicate product IDs are recorded with prev/final sheet info.
  - Row colour: not available in read_only mode; always None.

Logging:
  - Per-sheet parse summary: sheet name, rows parsed, rows skipped.
  - Per-file summary: total unique products, duplicates.
  - Warnings for invalid prices (no secrets, no PII).
"""

from __future__ import annotations

import io
import logging
import re
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import openpyxl

logger = logging.getLogger(__name__)

# -- Constants -----------------------------------------------------------------

_OUT_OF_STOCK_MARKERS: frozenset[str] = frozenset({
    "0", "0.00", "-",
    # Persian
    "\u0646\u0627\u0645\u0648\u062c\u0648\u062f",
    "\u0646\u0627\u0645\u0648\u062c\u0648\u062f \u0634\u062f",
    "\u062a\u0645\u0627\u0633 \u0628\u06af\u06cc\u0631\u06cc\u062f",
    # English
    "out of stock", "oos", "n/a", "na",
    # Symbols
    "x", "\u274c", "X", "\u00d7",
})

# Persian (U+06F0-U+06F9) and Arabic-Indic (U+0660-U+0669) -> ASCII
_DIGIT_TRANSLATION = str.maketrans({
    "\u06f0": "0", "\u06f1": "1", "\u06f2": "2", "\u06f3": "3", "\u06f4": "4",
    "\u06f5": "5", "\u06f6": "6", "\u06f7": "7", "\u06f8": "8", "\u06f9": "9",
    "\u0660": "0", "\u0661": "1", "\u0662": "2", "\u0663": "3", "\u0664": "4",
    "\u0665": "5", "\u0666": "6", "\u0667": "7", "\u0668": "8", "\u0669": "9",
})

_COLUMN_REF_RE = re.compile(r"^[A-Za-z]{1,3}$")
_WHOLE_NUMBER_RE = re.compile(r"^[0-9]+$")

_INVALID_PRODUCT_ID = {
    "code": "INVALID_PRODUCT_ID",
    "message": "Product ID must be a positive whole number.",
}
_INVALID_STOCK = {
    "code": "INVALID_STOCK",
    "message": "Stock must be a non-negative whole number.",
}


# -- Internal helpers ----------------------------------------------------------

def _normalize_price_text(raw: str) -> str:
    """Translate Persian/Arabic digits, remove Arabic thousands sep and commas."""
    return (
        raw
        .translate(_DIGIT_TRANSLATION)
        .replace("\u066c", "")   # U+066C ARABIC THOUSANDS SEPARATOR
        .replace(",", "")
        .strip()
    )


def _parse_whole_number(raw: str, *, allow_zero: bool) -> int:
    """Parse a canonical integer without rounding, truncation, or exponent expansion."""
    normalized = raw.translate(_DIGIT_TRANSLATION).strip()
    if not _WHOLE_NUMBER_RE.fullmatch(normalized):
        raise ValueError("not a canonical whole number")
    value = int(normalized)
    if value < 0 or (value == 0 and not allow_zero):
        raise ValueError("whole number outside allowed range")
    return value


def _parse_sheet_rows(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[dict]:
    """Parse one worksheet.  Returns a list of parsed row dicts.

    Matches the legacy row parser: row 3+, 30-consecutive-empty stop,
    1000-row max, Persian digits, OOS markers, error flags.
    """
    items: list[dict] = []
    consecutive_empty = 0
    skipped_no_id = 0
    skipped_bad_id = 0

    for row in ws.iter_rows(min_row=3, max_row=1002, values_only=False):
        b_cell = row[1] if len(row) > 1 else None
        b_val = b_cell.value if b_cell is not None else None

        if b_val is None:
            consecutive_empty += 1
            if consecutive_empty >= 30:
                logger.debug(
                    "spreadsheet _parse_sheet_rows sheet=%r stopping - "
                    "30 consecutive empty rows in column B",
                    ws.title,
                )
                break
            continue

        consecutive_empty = 0

        pid_raw = str(b_val).strip()
        pid_normalized = _normalize_price_text(pid_raw)
        try:
            pid = int(float(pid_normalized))
        except (ValueError, TypeError):
            skipped_bad_id += 1
            logger.debug(
                "spreadsheet _parse_sheet_rows sheet=%r skipped - non-int product_id=%r",
                ws.title, b_val,
            )
            continue
        if pid <= 0:
            skipped_bad_id += 1
            continue

        a_cell = row[0] if len(row) > 0 else None
        a_val = a_cell.value if a_cell is not None else None
        name = str(a_val).strip() if a_val is not None else ""

        c_cell = row[2] if len(row) > 2 else None
        c_val = c_cell.value if c_cell is not None else None

        price: float | None = None
        price_parse_error = False
        price_str = ""
        warning: str | None = None

        if c_val is None:
            price_str = ""
        else:
            price_str = str(c_val).strip()
            normalized = _normalize_price_text(price_str)
            if normalized.lower() in _OUT_OF_STOCK_MARKERS:
                price = None  # intentional OOS - not an error
            else:
                try:
                    candidate = float(normalized)
                    if candidate < 0:
                        price_parse_error = True
                        warning = f"Negative price ignored: {price_str!r}"
                        logger.warning(
                            "spreadsheet _parse_sheet_rows sheet=%r product_id=%d "
                            "negative price %r - flagged invalid",
                            ws.title, pid, price_str,
                        )
                    else:
                        price = candidate
                except (ValueError, TypeError):
                    price_parse_error = True
                    warning = f"Unparseable price: {price_str!r}"
                    logger.warning(
                        "spreadsheet _parse_sheet_rows sheet=%r product_id=%d "
                        "non-numeric price %r - flagged invalid",
                        ws.title, pid, price_str,
                    )

        items.append({
            "product_id": pid,
            "name": name,
            "price": price,
            "price_str": price_str,
            "price_parse_error": price_parse_error,
            "row_color": None,  # not available in read_only mode
            "warning": warning,
        })

    logger.info(
        "spreadsheet _parse_sheet_rows sheet=%r parsed=%d "
        "skipped(no_id=%d bad_id=%d)",
        ws.title, len(items), skipped_no_id, skipped_bad_id,
    )
    return items


# -- Public API ----------------------------------------------------------------

def parse_price_list(
    wb: "openpyxl.Workbook",
) -> tuple[dict[int, dict], list[dict]]:
    """Parse ALL worksheets.  Last sheet wins on duplicate product IDs.

    Returns:
        entries:    dict[product_id -> row_dict]  (the final winning entry per ID)
        duplicates: list of dicts describing each override
                    {product_id, prev_sheet, final_sheet, prev_price, final_price}
    """
    sheet_names = wb.sheetnames
    logger.info(
        "spreadsheet parse_price_list sheets=%d names=%s",
        len(sheet_names), sheet_names,
    )

    entries: dict[int, dict] = {}
    duplicates: list[dict] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sheet_items = _parse_sheet_rows(ws)
        logger.info(
            "spreadsheet parse_price_list sheet=%r contributed=%d product(s)",
            sheet_name, len(sheet_items),
        )
        for row in sheet_items:
            pid = row["product_id"]
            row_with_sheet = {**row, "sheet_name": sheet_name}
            if pid in entries:
                prev = entries[pid]
                logger.debug(
                    "spreadsheet parse_price_list duplicate product_id=%d "
                    "prev_sheet=%r overridden_by=%r prev_price=%s new_price=%s",
                    pid, prev["sheet_name"], sheet_name, prev["price"], row["price"],
                )
                duplicates.append({
                    "product_id": pid,
                    "prev_sheet": prev["sheet_name"],
                    "final_sheet": sheet_name,
                    "prev_price": prev["price"],
                    "final_price": row["price"],
                })
            entries[pid] = row_with_sheet

    logger.info(
        "spreadsheet parse_price_list total_unique=%d duplicates=%d",
        len(entries), len(duplicates),
    )
    return entries, duplicates


def parse_source_price_rows(
    wb: "openpyxl.Workbook",
    *,
    mapping: dict[str, dict[str, object]] | None = None,
    worksheet_mode: str = "all",
    worksheet_name: str | None = None,
) -> tuple[list[dict], dict]:
    """Parse worksheets into normalized source-row candidates for Workspace.

    Default columns preserve the current WooPrice-compatible behavior:
    A = product name, B = WooCommerce product ID, C = proposed price.
    Column D is treated as an optional SKU unless D is explicitly mapped to
    enabled stock. Mapping values can be spreadsheet letters or header names.
    """
    import openpyxl.utils

    mapping = mapping or {
        "id": {"enabled": True, "column": "B"},
        "price": {"enabled": True, "column": "C"},
        "stock": {"enabled": False, "column": "D"},
    }
    sheet_names = _selected_sheet_names(wb, worksheet_mode, worksheet_name)
    rows: list[dict] = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        headers = _header_map(ws)
        id_column = _resolve_column_index(mapping.get("id", {}).get("column"), headers) if mapping.get("id", {}).get("enabled") else None
        price_column = _resolve_column_index(mapping.get("price", {}).get("column"), headers) if mapping.get("price", {}).get("enabled") else None
        stock_column = _resolve_column_index(mapping.get("stock", {}).get("column"), headers) if mapping.get("stock", {}).get("enabled") else None
        sku_column = _sku_column_index(headers, stock_column)
        consecutive_empty = 0
        for row in ws.iter_rows(min_row=3, max_row=1002, values_only=False):
            raw_values = _raw_values(row)
            name_cell = _cell_at(row, 1)
            product_id_cell = _cell_at(row, id_column)
            price_cell = _cell_at(row, price_column)
            stock_cell = _cell_at(row, stock_column)
            sku_cell = _cell_at(row, sku_column)

            product_id_raw = _cell_text(product_id_cell)
            sku = _cell_text(sku_cell)
            product_name = _cell_text(name_cell)
            raw_price = _cell_text(price_cell)
            raw_stock = _cell_text(stock_cell)
            if not product_name and not product_id_raw and not sku and not raw_price and not raw_stock:
                consecutive_empty += 1
                if consecutive_empty >= 30:
                    break
                continue
            consecutive_empty = 0

            row_errors: list[str] = []
            row_warnings: list[str] = []
            row_error_details: list[dict[str, str]] = []
            product_id: str | None = None
            product_id_error = False
            if mapping.get("id", {}).get("enabled"):
                if not product_id_raw:
                    product_id_error = True
                elif product_id_raw:
                    try:
                        product_id = str(_parse_whole_number(product_id_raw, allow_zero=False))
                    except ValueError:
                        product_id_error = True
                if product_id_error:
                    row_errors.append("invalid_product_id")
                    row_error_details.append(dict(_INVALID_PRODUCT_ID))

            price: float | None = None
            price_parse_error = False
            price_enabled = bool(mapping.get("price", {}).get("enabled"))
            if price_enabled:
                if raw_price:
                    normalized = _normalize_price_text(raw_price)
                    if normalized.lower() in _OUT_OF_STOCK_MARKERS:
                        price = None
                    else:
                        try:
                            price = float(normalized)
                        except (TypeError, ValueError):
                            price_parse_error = True
                else:
                    price_parse_error = True
                if price_parse_error:
                    row_errors.append("invalid_price")

            stock: int | None = None
            stock_parse_error = False
            stock_enabled = bool(mapping.get("stock", {}).get("enabled"))
            if stock_enabled:
                if raw_stock:
                    try:
                        stock = _parse_whole_number(raw_stock, allow_zero=True)
                    except ValueError:
                        stock_parse_error = True
                else:
                    stock_parse_error = True
                if stock_parse_error:
                    row_errors.append("invalid_stock")
                    row_error_details.append(dict(_INVALID_STOCK))

            row_number = getattr(product_id_cell or sku_cell or price_cell or stock_cell or name_cell, "row", None)
            rows.append({
                "source_id": "nextcloud:primary",
                "source_type": "nextcloud_spreadsheet",
                "worksheet": sheet_name,
                "sheet_name": sheet_name,
                "row_number": row_number,
                "product_id": product_id,
                "external_product_id": product_id,
                "raw_product_id": product_id_raw,
                "product_id_error": product_id_error,
                "sku": sku,
                "product_name": product_name,
                "proposed_price": price,
                "source_price": price,
                "raw_price": raw_price,
                "price_parse_error": price_parse_error,
                "price_enabled": price_enabled,
                "source_stock": stock,
                "raw_stock": raw_stock,
                "stock_parse_error": stock_parse_error,
                "stock_enabled": stock_enabled,
                "id_enabled": bool(mapping.get("id", {}).get("enabled")),
                "row_errors": row_errors,
                "row_error_details": row_error_details,
                "row_warnings": row_warnings,
                "raw_values": raw_values,
                "raw": {
                    "product_name": product_name,
                    "product_id": product_id_raw,
                    "price": raw_price,
                    "stock": raw_stock,
                    "sku": sku,
                    "values": raw_values,
                },
            })

    duplicate_product_ids = _duplicate_values(row["product_id"] for row in rows if row.get("product_id"))
    duplicate_skus = _duplicate_values(row["sku"].strip().lower() for row in rows if row.get("sku"))
    duplicate_product_id_set = set(duplicate_product_ids)
    duplicate_sku_set = set(duplicate_skus)
    for row in rows:
        duplicate_id = bool(row.get("product_id") and row["product_id"] in duplicate_product_id_set)
        duplicate_sku = bool(row.get("sku") and row["sku"].strip().lower() in duplicate_sku_set)
        row["duplicate_product_id"] = duplicate_id
        row["duplicate_sku"] = duplicate_sku
        if duplicate_id:
            row["row_errors"].append("duplicate_product_id")
            row["row_error_details"].append({
                "code": "DUPLICATE_PRODUCT_ID",
                "message": "Product ID appears more than once in the spreadsheet.",
            })
        if duplicate_sku:
            row["row_errors"].append("duplicate_sku")
            row["row_error_details"].append({
                "code": "DUPLICATE_SKU",
                "message": "SKU appears more than once in the spreadsheet.",
            })
    return rows, {
        "duplicate_product_ids": duplicate_product_ids,
        "duplicate_skus": duplicate_skus,
    }


def _selected_sheet_names(wb: "openpyxl.Workbook", mode: str, name: str | None) -> list[str]:
    if str(mode or "all").strip().lower() != "selected":
        return list(wb.sheetnames)
    requested = str(name or "").strip()
    if not requested:
        raise ValueError("Selected worksheet name is required.")
    if requested not in wb.sheetnames:
        raise ValueError(f"Selected worksheet not found: {requested}")
    return [requested]


def _header_map(ws: "openpyxl.worksheet.worksheet.Worksheet") -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
        value = _cell_text(cell).casefold()
        if value:
            headers.setdefault(value, cell.column)
    return headers


def _resolve_column_index(ref: object, headers: dict[str, int]) -> int | None:
    text = str(ref or "").strip()
    if not text:
        return None
    if _COLUMN_REF_RE.fullmatch(text):
        import openpyxl.utils

        return openpyxl.utils.column_index_from_string(text.upper())
    return headers.get(text.casefold())


def _sku_column_index(headers: dict[str, int], stock_column: int | None) -> int | None:
    for key in ("sku", "product sku"):
        if key in headers and headers[key] != stock_column:
            return headers[key]
    return None if stock_column == 4 else 4


def _cell_at(row: tuple, column_index: int | None):
    if column_index is None or column_index <= 0:
        return None
    offset = column_index - 1
    return row[offset] if offset < len(row) else None


def _cell_text(cell: object | None) -> str:
    value = getattr(cell, "value", None)
    return "" if value is None else str(value).strip()


def _raw_values(row: tuple) -> dict[str, str]:
    import openpyxl.utils

    values: dict[str, str] = {}
    for cell in row:
        value = _cell_text(cell)
        if value:
            values[openpyxl.utils.get_column_letter(cell.column)] = value
    return values


def load_workbook_bytes(data: bytes) -> "openpyxl.Workbook":
    """Load an openpyxl Workbook from raw bytes."""
    import openpyxl
    logger.info("spreadsheet load_workbook_bytes size=%d bytes", len(data))
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def _duplicate_values(values) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        item = str(value).strip()
        if not item:
            continue
        if item in seen:
            duplicates.add(item)
        seen.add(item)
    return sorted(duplicates)
