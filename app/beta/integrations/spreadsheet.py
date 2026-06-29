"""FlowHub Beta — XLSX spreadsheet parser (BU5).

Adapted from the production-proven WooPrice nextcloud.py parser.
Preserves all WooPrice behaviour:
  - ALL worksheets are read; last sheet wins on duplicate product IDs.
  - Columns: A = Product Name, B = Product ID (int), C = Price.
  - Row 3 onward (rows 1–2 are headers).
  - Stops after 30 consecutive empty rows or row 1002 (max 1000 data rows).
  - Persian / Arabic-Indic digits are normalised to ASCII.
  - Out-of-stock markers are treated as price=None (not an error).
  - Unparseable prices set price_parse_error=True, price=None (not fatal).
  - Duplicate product IDs are recorded with prev/final sheet info.
  - Row colour: not available in read_only mode; always None.

Logging:
  - Per-sheet parse summary: sheet name, rows parsed, rows skipped.
  - Per-file summary: total unique products, duplicates.
  - Warnings for invalid prices (no secrets, no PII).
"""

from __future__ import annotations

import io
import logging
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import openpyxl

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────

_OUT_OF_STOCK_MARKERS: frozenset[str] = frozenset({
    "0", "0.00", "-",
    # Persian
    "ناموجود",
    "ناموجود شد",
    "تماس بگیرید",
    # English
    "out of stock", "oos", "n/a", "na",
    # Symbols
    "x", "❌", "✗", "×",
})

# Persian (U+06F0–U+06F9) and Arabic-Indic (U+0660–U+0669) → ASCII
_DIGIT_TRANSLATION = str.maketrans({
    "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
    "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
    "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
    "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
})


# ── Internal helpers ──────────────────────────────────────────────────────────

def _normalize_price_text(raw: str) -> str:
    """Translate Persian/Arabic digits, remove Arabic thousands sep and commas."""
    return (
        raw
        .translate(_DIGIT_TRANSLATION)
        .replace("٬", "")   # U+066C ARABIC THOUSANDS SEPARATOR
        .replace(",", "")
        .strip()
    )


def _parse_sheet_rows(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[dict]:
    """Parse one worksheet.  Returns a list of parsed row dicts.

    Matches WooPrice _parse_sheet_rows: row 3+, 30-consecutive-empty stop,
    1000-row max, Persian digits, OOS markers, error flags.
    """
    items: list[dict] = []
    consecutive_empty = 0
    skipped_no_id = 0
    skipped_bad_id = 0

    for row in ws.iter_rows(min_row=3, max_row=1002, values_only=False):
        b_cell = row[1] if len(row) > 1 else None
        b_val = b_cell.value if b_cell is not None else None

        if b_val is None:
            consecutive_empty += 1
            if consecutive_empty >= 30:
                logger.debug(
                    "spreadsheet _parse_sheet_rows sheet=%r stopping — "
                    "30 consecutive empty rows in column B",
                    ws.title,
                )
                break
            continue

        consecutive_empty = 0

        pid_raw = str(b_val).strip()
        pid_normalized = _normalize_price_text(pid_raw)
        try:
            pid = int(float(pid_normalized))
        except (ValueError, TypeError):
            skipped_bad_id += 1
            logger.debug(
                "spreadsheet _parse_sheet_rows sheet=%r skipped — non-int product_id=%r",
                ws.title, b_val,
            )
            continue
        if pid <= 0:
            skipped_bad_id += 1
            continue

        a_cell = row[0] if len(row) > 0 else None
        a_val = a_cell.value if a_cell is not None else None
        name = str(a_val).strip() if a_val is not None else ""

        c_cell = row[2] if len(row) > 2 else None
        c_val = c_cell.value if c_cell is not None else None

        price: float | None = None
        price_parse_error = False
        price_str = ""
        warning: str | None = None

        if c_val is None:
            price_str = ""
        else:
            price_str = str(c_val).strip()
            normalized = _normalize_price_text(price_str)
            if normalized.lower() in _OUT_OF_STOCK_MARKERS:
                price = None  # intentional OOS — not an error
            else:
                try:
                    candidate = float(normalized)
                    if candidate < 0:
                        price_parse_error = True
                        warning = f"Negative price ignored: {price_str!r}"
                        logger.warning(
                            "spreadsheet _parse_sheet_rows sheet=%r product_id=%d "
                            "negative price %r — flagged invalid",
                            ws.title, pid, price_str,
                        )
                    else:
                        price = candidate
                except (ValueError, TypeError):
                    price_parse_error = True
                    warning = f"Unparseable price: {price_str!r}"
                    logger.warning(
                        "spreadsheet _parse_sheet_rows sheet=%r product_id=%d "
                        "non-numeric price %r — flagged invalid",
                        ws.title, pid, price_str,
                    )

        items.append({
            "product_id": pid,
            "name": name,
            "price": price,
            "price_str": price_str,
            "price_parse_error": price_parse_error,
            "row_color": None,  # not available in read_only mode
            "warning": warning,
        })

    logger.info(
        "spreadsheet _parse_sheet_rows sheet=%r parsed=%d "
        "skipped(no_id=%d bad_id=%d)",
        ws.title, len(items), skipped_no_id, skipped_bad_id,
    )
    return items


# ── Public API ────────────────────────────────────────────────────────────────

def parse_price_list(
    wb: "openpyxl.Workbook",
) -> tuple[dict[int, dict], list[dict]]:
    """Parse ALL worksheets.  Last sheet wins on duplicate product IDs.

    Returns:
        entries:    dict[product_id -> row_dict]  (the final winning entry per ID)
        duplicates: list of dicts describing each override
                    {product_id, prev_sheet, final_sheet, prev_price, final_price}
    """
    sheet_names = wb.sheetnames
    logger.info(
        "spreadsheet parse_price_list sheets=%d names=%s",
        len(sheet_names), sheet_names,
    )

    entries: dict[int, dict] = {}
    duplicates: list[dict] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sheet_items = _parse_sheet_rows(ws)
        logger.info(
            "spreadsheet parse_price_list sheet=%r contributed=%d product(s)",
            sheet_name, len(sheet_items),
        )
        for row in sheet_items:
            pid = row["product_id"]
            row_with_sheet = {**row, "sheet_name": sheet_name}
            if pid in entries:
                prev = entries[pid]
                logger.debug(
                    "spreadsheet parse_price_list duplicate product_id=%d "
                    "prev_sheet=%r overridden_by=%r prev_price=%s new_price=%s",
                    pid, prev["sheet_name"], sheet_name, prev["price"], row["price"],
                )
                duplicates.append({
                    "product_id": pid,
                    "prev_sheet": prev["sheet_name"],
                    "final_sheet": sheet_name,
                    "prev_price": prev["price"],
                    "final_price": row["price"],
                })
            entries[pid] = row_with_sheet

    logger.info(
        "spreadsheet parse_price_list total_unique=%d duplicates=%d",
        len(entries), len(duplicates),
    )
    return entries, duplicates


def load_workbook_bytes(data: bytes) -> "openpyxl.Workbook":
    """Load an openpyxl Workbook from raw bytes."""
    import openpyxl
    logger.info("spreadsheet load_workbook_bytes size=%d bytes", len(data))
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
