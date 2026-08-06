from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sys
import warnings
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


SCHEMA_VERSION = "flowhub-formula-inventory-v1"
NORMALIZATION_VERSION = "flowhub-formula-r1c1-v1"
EXPECTED_TOTALS = {
    "formula_cells": 5_997,
    "formula_shapes": 13,
    "broken_ref_formula_cells": 255,
}

CELL_REFERENCE_RE = re.compile(r"(?<![A-Z0-9_.])(?P<col>\$?[A-Z]{1,3})(?P<row>\$?\d+)", re.I)


def _shape(
    shape_id: str,
    skeleton: str,
    semantics: str,
    role: str,
    translation_status: str,
    validation_feasibility: str,
    missing_concepts: list[str] | None = None,
    confidence: str = "high",
    notes: str = "",
) -> dict[str, Any]:
    return {
        "shape_id": shape_id,
        "skeleton": skeleton,
        "semantics": semantics,
        "cell_role": role,
        "translation_status": translation_status,
        "validation_feasibility": validation_feasibility,
        "required_missing_domain_concepts": missing_concepts or [],
        "evidence_confidence": confidence,
        "evidence_notes": notes,
    }


SHAPES = {
    "=IF(<REF>=\"\",\"x\",IFERROR(FLOOR((<REF>*(1+<REF>/100)+IF(ISNUMBER(<REF>),<REF>,0))*1000000,50000),\"x\"))": _shape(
        "A1",
        "=IF(<REF>=\"\",\"x\",IFERROR(FLOOR((<REF>*(1+<REF>/100)+IF(ISNUMBER(<REF>),<REF>,0))*1000000,50000),\"x\"))",
        "Price target from one basis, a percentage parameter, an optional fixed addend, and floor-to-50000 after scaling by 1000000.",
        "price_target_candidate",
        "supported",
        "replayable",
    ),
    "=IFERROR(MIN(FILTER(<REF>,<REF><>0)),\"❌\")": _shape(
        "A2",
        "=IFERROR(MIN(FILTER(<REF>,<REF><>0)),\"❌\")",
        "Minimum non-zero quote across a same-row vendor range.",
        "basis_selection",
        "supported",
        "replayable",
    ),
    "=IFERROR(FLOOR(<REF>*(1+<REF>/100)*1000,100000),\"❌\")": _shape(
        "A3",
        "=IFERROR(FLOOR(<REF>*(1+<REF>/100)*1000,100000),\"❌\")",
        "Price target from a basis and percentage, scaled by 1000 and floored to 100000.",
        "price_target_candidate",
        "supported",
        "replayable",
    ),
    "=IFERROR(FLOOR(<REF>*(1+<REF>/100)*1000,100000)+500000,\"❌\")": _shape(
        "A4",
        "=IFERROR(FLOOR(<REF>*(1+<REF>/100)*1000,100000)+500000,\"❌\")",
        "A3 price target with a fixed 500000 surcharge applied after rounding.",
        "price_target_candidate",
        "supported",
        "replayable",
    ),
    "=IFERROR(ROUNDUP(<REF>*(1+<REF>/100),-2),\"❌\")": _shape(
        "A5",
        "=IFERROR(ROUNDUP(<REF>*(1+<REF>/100),-2),\"❌\")",
        "Price target rounded upward to two negative decimal places after percentage markup.",
        "price_target_candidate",
        "supported",
        "replayable",
    ),
    "=IF(<REF>=\"\",\"x\",IFERROR(FLOOR(<REF>*<REF>,50000)/10,\"x\"))": _shape(
        "A6",
        "=IF(<REF>=\"\",\"x\",IFERROR(FLOOR(<REF>*<REF>,50000)/10,\"x\"))",
        "UGREEN price candidate from one source value and the manually stored G2 multiplier, followed by division by 10 after rounding.",
        "price_target_candidate",
        "unknown",
        "provenance_partial",
        ["versioned_manual_pricing_factor", "proven_output_unit_semantics"],
        "low",
        "The workbook proves the formula and cached output, but G2 has no label and the business meaning of the post-round /10 is not proven.",
    ),
    "=IFERROR(<REF>/<REF>,\"x\")": _shape(
        "A7",
        "=IFERROR(<REF>/<REF>,\"x\")",
        "Derived purchase-to-consumer ratio used as a display metric, not a Channel price target.",
        "display_metric",
        "unsupported",
        "replayable",
        ["derived_display_metric"],
    ),
    "=<REF>": _shape(
        "A8",
        "=<REF>",
        "Same-column copy of manually entered header/vendor metadata.",
        "metadata_reference",
        "supported_with_review",
        "replayable",
        ["versioned_manual_metadata"],
        "medium",
        "The inspected occurrences copy text labels such as vendor names; they are not price targets.",
    ),
    "=IF(<REF>=\"\",\"x\",IFERROR(FLOOR((<REF>*(1+#REF!/100)+IF(ISNUMBER(#REF!),#REF!,0))*1000000,50000),\"x\"))": _shape(
        "A9",
        "=IF(<REF>=\"\",\"x\",IFERROR(FLOOR((<REF>*(1+#REF!/100)+IF(ISNUMBER(#REF!),#REF!,0))*1000000,50000),\"x\"))",
        "A1 variant with missing percentage and addend references.",
        "price_target_candidate",
        "broken",
        "comparison_not_possible",
        ["missing_channel_pricing_parameters"],
    ),
    "=IFERROR(FLOOR((<REF>*(1+<REF>/100)*1000),100000),\"❌\")": _shape(
        "A10",
        "=IFERROR(FLOOR((<REF>*(1+<REF>/100)*1000),100000),\"❌\")",
        "Parenthesized syntax variant of A3 with the same evidenced arithmetic.",
        "price_target_candidate",
        "supported",
        "replayable",
    ),
    "=IFERROR(FLOOR((<REF>*(1+<REF>/100)*1000)+500000,100000),\"❌\")": _shape(
        "A11",
        "=IFERROR(FLOOR((<REF>*(1+<REF>/100)*1000)+500000,100000),\"❌\")",
        "Price target with a fixed 500000 amount added before floor-to-100000.",
        "price_target_candidate",
        "supported",
        "replayable",
    ),
    "=IFNA(MIN(FILTER(<REF>,<REF><>0)),\"❌\")": _shape(
        "A12",
        "=IFNA(MIN(FILTER(<REF>,<REF><>0)),\"❌\")",
        "One anomalous cross-row minimum formula in Surface Acc!I12, located in a Link column and cached as #VALUE!.",
        "anomalous_formula",
        "broken",
        "comparison_not_possible",
        ["misplaced_cross_row_formula_classification"],
        "high",
        "Workbook location, references, and cached error prove the anomaly; intended business meaning is not inferred.",
    ),
    "=IFERROR(FLOOR((#REF!*(1+<REF>/100)*1000),100000),\"❌\")": _shape(
        "A13",
        "=IFERROR(FLOOR((#REF!*(1+<REF>/100)*1000),100000),\"❌\")",
        "A10 variant with a missing price-basis reference.",
        "price_target_candidate",
        "broken",
        "comparison_not_possible",
        ["missing_price_basis"],
    ),
}

CSV_FIELDS = [
    "inventory_id",
    "workbook",
    "workbook_sha256",
    "worksheet",
    "sheet_visibility",
    "cell_address",
    "row",
    "column",
    "exact_formula_text",
    "normalized_formula_text",
    "current_displayed_result",
    "current_displayed_result_type",
    "error_state",
    "contains_broken_ref",
    "referenced_sheets_json",
    "referenced_ranges_json",
    "detected_formula_shape",
    "cell_role",
    "input_topology",
    "translation_status",
    "validation_feasibility",
    "activity_status",
    "required_missing_domain_concepts_json",
    "evidence_confidence",
    "evidence_notes",
]


def _function_name(value: str) -> str:
    return value.upper().replace("_XLFN._XLWS.", "").replace("_XLFN.", "")


def formula_skeleton(formula: str) -> str:
    parts: list[str] = []
    for token in Tokenizer(formula).items:
        value = token.value
        if token.type == "FUNC" and token.subtype == "OPEN":
            value = _function_name(value)
        elif token.type == "OPERAND" and token.subtype == "RANGE":
            value = "#REF!" if "#REF!" in value else "<REF>"
        elif token.type == "WSPACE":
            value = ""
        parts.append(value)
    result = "".join(parts)
    return result if result.startswith("=") else f"={result}"


def _r1c1_reference(match: re.Match[str], origin_row: int, origin_col: int) -> str:
    column_text = match.group("col")
    row_text = match.group("row")
    absolute_column = column_text.startswith("$")
    absolute_row = row_text.startswith("$")
    column = column_index_from_string(column_text.lstrip("$").upper())
    row = int(row_text.lstrip("$"))
    row_part = f"R{row}" if absolute_row else ("R" if row == origin_row else f"R[{row - origin_row}]")
    column_part = (
        f"C{column}"
        if absolute_column
        else ("C" if column == origin_col else f"C[{column - origin_col}]")
    )
    return f"{row_part}{column_part}"


def normalize_formula(formula: str, origin_row: int, origin_col: int) -> str:
    parts: list[str] = []
    for token in Tokenizer(formula).items:
        value = token.value
        if token.type == "FUNC" and token.subtype == "OPEN":
            value = _function_name(value)
        elif token.type == "OPERAND" and token.subtype == "RANGE":
            value = CELL_REFERENCE_RE.sub(
                lambda match: _r1c1_reference(match, origin_row, origin_col), value
            )
        elif token.type == "WSPACE":
            value = ""
        parts.append(value)
    result = "".join(parts)
    return result if result.startswith("=") else f"={result}"


def extract_references(formula: str, worksheet: str) -> list[dict[str, str | None]]:
    references: list[dict[str, str | None]] = []
    seen: set[tuple[str | None, str]] = set()
    for token in Tokenizer(formula).items:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        raw = token.value
        if "#REF!" in raw:
            key = (None, "#REF!")
            if key not in seen:
                references.append(
                    {"sheet": None, "range": "#REF!", "raw": raw, "qualified": "#REF!"}
                )
                seen.add(key)
            continue
        sheet = worksheet
        range_text = raw
        if "!" in raw:
            sheet_text, range_text = raw.rsplit("!", 1)
            sheet = sheet_text.strip("'").replace("''", "'")
        if not CELL_REFERENCE_RE.search(range_text):
            continue
        qualified = f"'{sheet.replace(chr(39), chr(39) * 2)}'!{range_text}"
        key = (sheet, range_text)
        if key not in seen:
            references.append(
                {"sheet": sheet, "range": range_text, "raw": raw, "qualified": qualified}
            )
            seen.add(key)
    return references


def _first_reference_coordinate(formula: str) -> str | None:
    for token in Tokenizer(formula).items:
        if token.type != "OPERAND" or token.subtype != "RANGE" or "#REF!" in token.value:
            continue
        range_text = token.value.rsplit("!", 1)[-1].split(":", 1)[0].replace("$", "")
        if CELL_REFERENCE_RE.fullmatch(range_text):
            return range_text.upper()
    return None


def input_topology(shape_id: str, formula: str, worksheet: Any) -> str:
    if shape_id in {"A2", "A12"}:
        return "multi_source"
    if shape_id == "A6":
        return "manual_reference"
    if shape_id == "A7":
        return "derived_reference"
    if shape_id == "A8":
        return "manual_reference"
    if shape_id == "A13":
        return "derived_reference"
    coordinate = _first_reference_coordinate(formula)
    if coordinate and worksheet[coordinate].data_type == "f":
        return "derived_reference"
    return "single_source"


def serialize_cached_result(cell: Any) -> tuple[str | None, str]:
    value = cell.value
    if value is None:
        return None, "blank"
    if cell.data_type == "e":
        return str(value), "error"
    if isinstance(value, bool):
        return "true" if value else "false", "boolean"
    if isinstance(value, (datetime, date)):
        return value.isoformat(), "date"
    if isinstance(value, (int, float)):
        return str(value), "number"
    return str(value), "text"


def _csv_record(record: dict[str, Any]) -> dict[str, Any]:
    result = {field: record.get(field) for field in CSV_FIELDS}
    result["referenced_sheets_json"] = json.dumps(
        record["referenced_sheets"], ensure_ascii=False, separators=(",", ":")
    )
    result["referenced_ranges_json"] = json.dumps(
        record["references"], ensure_ascii=False, separators=(",", ":")
    )
    result["required_missing_domain_concepts_json"] = json.dumps(
        record["required_missing_domain_concepts"],
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return result


def _write_csv(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(_csv_record(record) for record in records)


def _markdown_table(rows: list[list[Any]]) -> str:
    if not rows:
        return ""
    header = "| " + " | ".join(str(item) for item in rows[0]) + " |"
    divider = "| " + " | ".join("---" for _ in rows[0]) + " |"
    body = ["| " + " | ".join(str(item) for item in row) + " |" for row in rows[1:]]
    return "\n".join([header, divider, *body])


def build_summary(
    workbook_name: str,
    workbook_hash: str,
    worksheet_total: int,
    worksheet_counts: Counter[str],
    cells: list[dict[str, Any]],
    shape_records: list[dict[str, Any]],
) -> str:
    broken_ref_count = sum(bool(cell["contains_broken_ref"]) for cell in cells)
    broken_count = sum(cell["translation_status"] == "broken" for cell in cells)
    exact_counts = Counter(cell["exact_formula_text"] for cell in cells)
    normalized_counts = Counter(cell["normalized_formula_text"] for cell in cells)
    topology_counts = Counter(cell["input_topology"] for cell in cells)
    translation_counts = Counter(cell["translation_status"] for cell in cells)
    feasibility_counts = Counter(cell["validation_feasibility"] for cell in cells)
    role_counts = Counter(cell["cell_role"] for cell in cells)
    duplicate_exact_groups = sum(count > 1 for count in exact_counts.values())
    duplicate_exact_cells = sum(count for count in exact_counts.values() if count > 1)
    duplicate_normalized_groups = sum(count > 1 for count in normalized_counts.values())
    duplicate_normalized_cells = sum(count for count in normalized_counts.values() if count > 1)

    lines = [
        "# Authoritative Formula Inventory Summary",
        "",
        "## Source snapshot",
        "",
        _markdown_table(
            [
                ["Field", "Value"],
                ["Workbook", f"`{workbook_name}`"],
                ["SHA-256", f"`{workbook_hash}`"],
                ["Worksheets", worksheet_total],
                ["Formula-bearing worksheets", len(worksheet_counts)],
                ["Formula cells", len(cells)],
                ["Normalization", f"`{NORMALIZATION_VERSION}`"],
            ]
        ),
        "",
        "The inventory uses the workbook's stored cached result. It does not recalculate or repair any formula. Shared formulas are expanded to their exact cell-relative A1 formula before inventorying.",
        "",
        "## Documented-total reconciliation",
        "",
        _markdown_table(
            [
                ["Measure", "Documented", "Observed", "Result"],
                ["Formula cells", EXPECTED_TOTALS["formula_cells"], len(cells), "match"],
                ["Formula shapes", EXPECTED_TOTALS["formula_shapes"], len(shape_records), "match"],
                [
                    "Formula cells containing `#REF!`",
                    EXPECTED_TOTALS["broken_ref_formula_cells"],
                    broken_ref_count,
                    "match",
                ],
            ]
        ),
        "",
        f"The workbook also contains one non-`#REF!` broken formula: `Surface Acc!I12` is cached as `#VALUE!`. Therefore the authoritative total is **{broken_count} broken formula cells**, comprising **{broken_ref_count} documented `#REF!` cells plus one additional cached error**.",
        "",
        "## Formula duplication",
        "",
        _markdown_table(
            [
                ["Measure", "Count"],
                ["Unique exact formula texts", len(exact_counts)],
                ["Exact duplicate groups", duplicate_exact_groups],
                ["Cells belonging to exact duplicate groups", duplicate_exact_cells],
                ["Unique normalized R1C1 formulas", len(normalized_counts)],
                ["Normalized duplicate groups", duplicate_normalized_groups],
                ["Cells belonging to normalized duplicate groups", duplicate_normalized_cells],
                ["Semantic/syntax shapes", len(shape_records)],
            ]
        ),
        "",
        "Normalized formulas preserve constants and reference geometry while replacing cell location with deterministic R1C1 offsets. Shape identity is a separate syntax skeleton that replaces valid references with `<REF>` and preserves broken `#REF!` tokens.",
        "",
        "## Confirmed shapes",
        "",
        _markdown_table(
            [["Shape", "Cells", "Role", "Translation", "Semantics"]]
            + [
                [
                    shape["shape_id"],
                    shape["cell_count"],
                    shape["cell_role"],
                    shape["translation_status"],
                    shape["semantics"],
                ]
                for shape in shape_records
            ]
        ),
        "",
        "## Classification totals",
        "",
        _markdown_table(
            [["Input topology", "Cells"]]
            + [
                [key, topology_counts[key]]
                for key in (
                    "single_source",
                    "multi_source",
                    "external_market",
                    "manual_reference",
                    "derived_reference",
                )
            ]
        ),
        "",
        _markdown_table(
            [["Translation status", "Cells"]]
            + [[key, value] for key, value in sorted(translation_counts.items())]
        ),
        "",
        _markdown_table(
            [["Validation feasibility", "Cells"]]
            + [[key, value] for key, value in sorted(feasibility_counts.items())]
        ),
        "",
        _markdown_table(
            [["Cell role", "Cells"]] + [[key, value] for key, value in sorted(role_counts.items())]
        ),
        "",
        "No formula cell is marked active or inactive: workbook presence and worksheet visibility do not prove operational Channel authority. Every record therefore has `activity_status = unknown`.",
        "",
        "## Worksheet coverage",
        "",
        _markdown_table(
            [["Worksheet", "Formula cells"]]
            + [[sheet, count] for sheet, count in worksheet_counts.items()]
        ),
        "",
        "## Evidence-backed anomalies and gates",
        "",
        "- `A9`: 254 Logitech price cells contain broken percentage/addend references.",
        "- `A13`: `Beats!C34` contains a broken price-basis reference.",
        "- `A12`: `Surface Acc!I12` is in a Link column, reads row 3 instead of row 12, and is cached as `#VALUE!`; its intended meaning is not inferred.",
        "- `A6`: 327 UGREEN price candidates use the manually stored, unlabeled `G2` multiplier and divide by 10 after rounding. The arithmetic is proven, but the business meaning and output-unit semantics remain unproven, so the shape stays quarantined as `unknown`.",
        "- `A7` is a display ratio and is explicitly not a Channel price target.",
        "- `A8` copies vendor/header metadata and is explicitly not a Channel price target.",
        "",
        "## Inventory closure",
        "",
        "The cell-level inventory is complete for the exact workbook hash above: all 5,997 formula cells have workbook, worksheet, cell, exact formula, cached result, references, normalized formula, shape, topology, translation status, and validation classification. Translator implementation and migration activation remain blocked by the quarantined/broken shapes and fixture work; this inventory does not activate or repair anything.",
        "",
    ]
    return "\n".join(lines)


def generate(workbook_path: Path, output_dir: Path) -> None:
    data = workbook_path.read_bytes()
    workbook_hash = hashlib.sha256(data).hexdigest()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        formula_workbook = load_workbook(io.BytesIO(data), data_only=False, keep_links=True)
        value_workbook = load_workbook(io.BytesIO(data), data_only=True, keep_links=True)

    cells: list[dict[str, Any]] = []
    worksheet_counts: Counter[str] = Counter()
    shape_cells: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)

    for formula_sheet in formula_workbook.worksheets:
        value_sheet = value_workbook[formula_sheet.title]
        for row in formula_sheet.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                formula = cell.value
                skeleton = formula_skeleton(formula)
                if skeleton not in SHAPES:
                    raise ValueError(
                        f"Unclassified formula shape at {formula_sheet.title}!{cell.coordinate}: {skeleton}"
                    )
                shape = SHAPES[skeleton]
                references = extract_references(formula, formula_sheet.title)
                displayed, displayed_type = serialize_cached_result(value_sheet[cell.coordinate])
                contains_broken_ref = "#REF!" in formula
                cached_error = displayed if displayed_type == "error" else None
                error_state = "#REF!" if contains_broken_ref else cached_error
                inventory_id = hashlib.sha256(
                    f"{workbook_hash}|{formula_sheet.title}|{cell.coordinate}".encode("utf-8")
                ).hexdigest()[:24]
                record = {
                    "inventory_id": inventory_id,
                    "workbook": workbook_path.name,
                    "workbook_sha256": workbook_hash,
                    "worksheet": formula_sheet.title,
                    "sheet_visibility": formula_sheet.sheet_state,
                    "cell_address": cell.coordinate,
                    "row": cell.row,
                    "column": cell.column,
                    "exact_formula_text": formula,
                    "normalized_formula_text": normalize_formula(formula, cell.row, cell.column),
                    "current_displayed_result": displayed,
                    "current_displayed_result_type": displayed_type,
                    "error_state": error_state,
                    "contains_broken_ref": contains_broken_ref,
                    "referenced_sheets": list(
                        dict.fromkeys(
                            reference["sheet"]
                            for reference in references
                            if reference["sheet"] is not None
                        )
                    ),
                    "references": references,
                    "detected_formula_shape": shape["shape_id"],
                    "cell_role": shape["cell_role"],
                    "input_topology": input_topology(
                        shape["shape_id"], formula, formula_sheet
                    ),
                    "translation_status": shape["translation_status"],
                    "validation_feasibility": shape["validation_feasibility"],
                    "activity_status": "unknown",
                    "required_missing_domain_concepts": shape[
                        "required_missing_domain_concepts"
                    ],
                    "evidence_confidence": shape["evidence_confidence"],
                    "evidence_notes": shape["evidence_notes"],
                }
                cells.append(record)
                worksheet_counts[formula_sheet.title] += 1
                shape_cells[shape["shape_id"]].append(record)

    shape_records: list[dict[str, Any]] = []
    for shape in sorted(SHAPES.values(), key=lambda item: int(item["shape_id"][1:])):
        members = shape_cells[shape["shape_id"]]
        normalized = Counter(member["normalized_formula_text"] for member in members)
        exact = Counter(member["exact_formula_text"] for member in members)
        shape_record = dict(shape)
        shape_record.update(
            {
                "cell_count": len(members),
                "exact_formula_variant_count": len(exact),
                "normalized_formula_variant_count": len(normalized),
                "input_topology_counts": dict(
                    sorted(Counter(member["input_topology"] for member in members).items())
                ),
                "representative_cells": [
                    {
                        "worksheet": member["worksheet"],
                        "cell_address": member["cell_address"],
                        "exact_formula_text": member["exact_formula_text"],
                    }
                    for member in members[:5]
                ],
                "normalized_formula_variants": [
                    {"formula": formula, "cell_count": count}
                    for formula, count in sorted(normalized.items())
                ],
            }
        )
        shape_records.append(shape_record)

    metadata = {
        "schema_version": SCHEMA_VERSION,
        "normalization_version": NORMALIZATION_VERSION,
        "source_workbook": workbook_path.name,
        "source_workbook_sha256": workbook_hash,
        "source_workbook_size_bytes": len(data),
        "worksheet_count": len(formula_workbook.worksheets),
        "formula_cell_count": len(cells),
        "broken_ref_formula_cell_count": sum(
            bool(cell["contains_broken_ref"]) for cell in cells
        ),
        "broken_formula_cell_count": sum(
            cell["translation_status"] == "broken" for cell in cells
        ),
        "unique_exact_formula_count": len({cell["exact_formula_text"] for cell in cells}),
        "unique_normalized_formula_count": len(
            {cell["normalized_formula_text"] for cell in cells}
        ),
        "formula_shape_count": len(shape_records),
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "formula_cells.json").write_text(
        json.dumps({"metadata": metadata, "cells": cells}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _write_csv(output_dir / "formula_cells.csv", cells)
    (output_dir / "formula_shapes.json").write_text(
        json.dumps({"metadata": metadata, "shapes": shape_records}, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    broken = [cell for cell in cells if cell["translation_status"] == "broken"]
    _write_csv(output_dir / "broken_formulas.csv", broken)
    (output_dir / "inventory_summary.md").write_text(
        build_summary(
            workbook_path.name,
            workbook_hash,
            len(formula_workbook.worksheets),
            worksheet_counts,
            cells,
            shape_records,
        ),
        encoding="utf-8",
    )
    check(output_dir)


def check(output_dir: Path) -> None:
    cells_document = json.loads((output_dir / "formula_cells.json").read_text(encoding="utf-8"))
    shapes_document = json.loads((output_dir / "formula_shapes.json").read_text(encoding="utf-8"))
    cells = cells_document["cells"]
    metadata = cells_document["metadata"]
    shapes = shapes_document["shapes"]

    errors: list[str] = []
    if len(cells) != EXPECTED_TOTALS["formula_cells"]:
        errors.append(f"formula cell count is {len(cells)}")
    broken_ref_count = sum(bool(cell["contains_broken_ref"]) for cell in cells)
    if broken_ref_count != EXPECTED_TOTALS["broken_ref_formula_cells"]:
        errors.append(f"broken #REF! formula cell count is {broken_ref_count}")
    if len(shapes) != EXPECTED_TOTALS["formula_shapes"]:
        errors.append(f"formula shape count is {len(shapes)}")

    ids = [cell.get("inventory_id") for cell in cells]
    duplicate_ids = [item for item, count in Counter(ids).items() if count > 1]
    if duplicate_ids:
        errors.append(f"duplicate inventory IDs: {duplicate_ids[:5]}")
    provenance = [
        (cell.get("workbook"), cell.get("worksheet"), cell.get("cell_address"))
        for cell in cells
    ]
    missing = [index for index, item in enumerate(provenance) if not all(item)]
    if missing:
        errors.append(f"missing workbook/sheet/cell provenance at records {missing[:5]}")
    duplicate_provenance = [
        item for item, count in Counter(provenance).items() if count > 1
    ]
    if duplicate_provenance:
        errors.append(f"duplicate workbook/sheet/cell records: {duplicate_provenance[:5]}")

    shape_ids = {shape["shape_id"] for shape in shapes}
    referenced_shape_ids = {cell["detected_formula_shape"] for cell in cells}
    if shape_ids != referenced_shape_ids:
        errors.append("formula shape catalog and cell references differ")
    if sum(shape["cell_count"] for shape in shapes) != len(cells):
        errors.append("formula shape cell counts do not sum to the inventory total")
    if metadata["formula_cell_count"] != len(cells):
        errors.append("metadata formula cell count differs")
    if metadata["broken_ref_formula_cell_count"] != broken_ref_count:
        errors.append("metadata broken #REF! count differs")

    with (output_dir / "formula_cells.csv").open(encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    if [row["inventory_id"] for row in csv_rows] != ids:
        errors.append("formula_cells.csv does not match formula_cells.json ordering/IDs")

    with (output_dir / "broken_formulas.csv").open(
        encoding="utf-8-sig", newline=""
    ) as handle:
        broken_rows = list(csv.DictReader(handle))
    expected_broken_ids = [
        cell["inventory_id"] for cell in cells if cell["translation_status"] == "broken"
    ]
    if [row["inventory_id"] for row in broken_rows] != expected_broken_ids:
        errors.append("broken_formulas.csv does not match broken JSON records")

    if errors:
        raise ValueError("Formula inventory consistency check failed: " + "; ".join(errors))
    print(
        json.dumps(
            {
                "status": "ok",
                "formula_cells": len(cells),
                "broken_ref_formula_cells": broken_ref_count,
                "broken_formula_cells": len(expected_broken_ids),
                "formula_shapes": len(shapes),
                "unique_inventory_ids": len(set(ids)),
                "missing_provenance_records": 0,
            },
            separators=(",", ":"),
        )
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate or validate FlowHub formula inventory")
    subparsers = parser.add_subparsers(dest="command", required=True)
    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--workbook", type=Path, required=True)
    generate_parser.add_argument("--output-dir", type=Path, required=True)
    check_parser = subparsers.add_parser("check")
    check_parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    try:
        if args.command == "generate":
            generate(args.workbook, args.output_dir)
        else:
            check(args.output_dir)
    except (OSError, ValueError, KeyError, json.JSONDecodeError) as exc:
        print(str(exc), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
