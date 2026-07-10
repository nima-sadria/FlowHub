from __future__ import annotations

import os
import uuid
from datetime import datetime
from io import BytesIO

import pytest

os.environ.setdefault("FLOWHUB_DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("FLOWHUB_JWT_SECRET", "test-workspace-workflow-jwt-secret-32bytes!")

from app.flowhub.auth import models as _auth_models  # noqa: F401
from app.flowhub.data_layer import models as _data_layer_models  # noqa: F401
from app.flowhub.integration_platform import models as _integration_platform_models  # noqa: F401
from app.flowhub.setup import models as _setup_models  # noqa: F401
from app.flowhub.write_pipeline import models as _write_pipeline_models  # noqa: F401


@pytest.fixture()
def db_engine():
    from sqlalchemy import create_engine
    from sqlalchemy.pool import StaticPool

    from app.flowhub.database import FlowHubBase, _get_engine

    _get_engine.cache_clear()
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    FlowHubBase.metadata.create_all(engine)
    yield engine
    FlowHubBase.metadata.drop_all(engine)
    engine.dispose()
    _get_engine.cache_clear()


@pytest.fixture()
def db(db_engine):
    from sqlalchemy.orm import sessionmaker

    Session = sessionmaker(bind=db_engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture()
def client(db_engine):
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import sessionmaker

    from app.flowhub.app import app
    from app.flowhub.database import get_db

    Session = sessionmaker(bind=db_engine)

    def _override_get_db():
        s = Session()
        try:
            yield s
        finally:
            s.close()

    app.dependency_overrides[get_db] = _override_get_db
    with TestClient(app, raise_server_exceptions=True) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture()
def auth_headers(client, db):
    from app.flowhub.auth.jwt_service import create_access_token
    from app.flowhub.auth.models import FlowHubUser
    from app.flowhub.auth.password import hash_password

    username = f"workspaceadmin_{uuid.uuid4().hex}"
    user = FlowHubUser(username=username, hashed_password=hash_password("password123"), role="admin")
    db.add(user)
    db.commit()
    db.refresh(user)
    token = create_access_token(user.id, user.username, user.role)
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture()
def configured_db(db):
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(db).set_many(
        {
            "woocommerce.url": "https://store.example.test",
            "woocommerce.key": "ck_test",
            "woocommerce.secret": "cs_test",
            "nextcloud.url": "https://cloud.example.test",
            "nextcloud.username": "user",
            "nextcloud.password": "pass",
            "nextcloud.spreadsheet_path": "/prices.xlsx",
            "server.currency": "EUR",
            "setup.completed": "true",
        }
    )
    return db


def test_nextcloud_spreadsheet_import_success_generates_preview_and_dry_run(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(configured_db, "101", "Test Product", "SKU-101", "100.00")

    async def fake_download(self, path):
        assert path == "/prices.xlsx"
        return _xlsx([["Test Product", 101, "110.00", "SKU-101"]]), {"etag": "etag-1", "last_modified": "now"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    preview = client.post("/api/v2/workspace/preview", headers=auth_headers)
    assert preview.status_code == 200
    data = preview.json()
    from app.flowhub.data_layer.models import DlWorkspacePreview

    stored_preview = configured_db.get(DlWorkspacePreview, data["id"])
    assert stored_preview is not None
    assert stored_preview.preview_hash and len(stored_preview.preview_hash) == 64
    assert stored_preview.rows_json[0]["id"] == data["rows"][0]["id"]
    from app.flowhub.integration_platform.models import IntegrationConnectorEvent

    preview_event = (
        configured_db.query(IntegrationConnectorEvent)
        .filter(IntegrationConnectorEvent.event_name == "preview_created")
        .one()
    )
    assert preview_event.metadata_json["preview_id"] == data["id"]
    assert preview_event.metadata_json["preview_hash"] == stored_preview.preview_hash
    assert data["external_call_performed"] is True
    assert data["summary"]["total_rows"] == 1
    assert data["summary"]["valid_changes"] == 1
    row = data["rows"][0]
    assert row["source"]["sourceId"] == "nextcloud:primary"
    assert row["source"]["sourceFilePath"] == "/prices.xlsx"
    assert row["source"]["worksheet"] == "Sheet1"
    assert row["source"]["rowNumber"] == 3
    assert row["source"]["productId"] == "101"
    assert row["source"]["sku"] == "SKU-101"
    assert row["source"]["productName"] == "Test Product"
    assert row["matchedProduct"]["name"] == "Test Product"
    assert row["currentPrice"] == 100.0
    assert row["proposedPrice"] == 110.0
    assert row["difference"] == 10.0
    assert row["changePct"] == 10.0
    assert row["status"] == "valid_change"
    assert row["warnings"] == []
    assert row["errors"] == []
    assert row["eligible_for_dry_run"] is True
    assert data["changes"][0]["source"]["worksheet"] == "Sheet1"
    assert data["changes"][0]["source"]["rowNumber"] == 3

    dry_run = client.post(
        "/api/v2/write-pipeline/dry-run",
        headers=auth_headers,
        json={
            "previewId": data["id"],
            "selectedRowIds": [row["id"]],
        },
    )
    assert dry_run.status_code == 201
    batch = dry_run.json()
    assert batch["status"] == "dry_run_ready"
    assert batch["sourcePreviewId"] == data["id"]
    assert batch["items"][0]["productId"] == "101"

    execute = client.post(f"/api/v2/write-pipeline/batches/{batch['id']}/execute", headers=auth_headers)
    assert execute.status_code == 409
    assert "approved Dry Run" in execute.text


def test_workspace_preview_succeeds_after_woocommerce_channel_cache_refresh(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient

    configured = client.put(
        "/api/v2/commerce/channels/woocommerce:primary/settings",
        headers=auth_headers,
        json={
            "display_name": "WooCommerce",
            "enabled": True,
            "settings": {"url": "https://store.example.test"},
            "secrets": {"key": "ck_test", "secret": "cs_test"},
        },
    )
    assert configured.status_code == 200

    async def fake_list_products(_creds, *, page=1, **_kwargs):
        if page > 1:
            return [], 1, 1
        return [
            {
                "id": 101,
                "name": "Test Product",
                "type": "simple",
                "sku": "SKU-101",
                "regular_price": "100.00",
                "sale_price": "",
                "price": "100.00",
                "stock_quantity": 4,
                "stock_status": "instock",
                "manage_stock": True,
                "backorders": "no",
                "categories": [],
                "images": [],
                "status": "publish",
                "date_modified_gmt": "2026-07-10T10:00:00",
            }
        ], 1, 1

    async def no_variations(_creds, _product_id, **_kwargs):
        return []

    async def fake_download(self, path):
        assert path == "/prices.xlsx"
        return _xlsx([["Test Product", 101, "110.00", "SKU-101"]]), {"etag": "etag-after-refresh"}

    monkeypatch.setattr("app.connectors.read.woocommerce.list_products_paged", fake_list_products)
    monkeypatch.setattr("app.connectors.read.woocommerce.list_variations", no_variations)
    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    refresh = client.post(
        "/api/v2/commerce/channels/woocommerce:primary/refresh-cache",
        headers=auth_headers,
    )
    assert refresh.status_code == 200
    assert refresh.json()["ok"] is True
    assert refresh.json()["cache_rows_upserted"] == 1

    preview = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert preview.status_code == 200
    row = preview.json()["rows"][0]
    assert row["source"]["productId"] == "101"
    assert row["currentPrice"] == 100.0
    assert row["proposedPrice"] == 110.0
    assert row["eligible_for_dry_run"] is True


def test_spreadsheet_path_selected_from_source_settings_feeds_preview_workflow(
    client, auth_headers, db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(db).set_many(
        {
            "woocommerce.url": "https://store.example.test",
            "woocommerce.key": "ck_test",
            "woocommerce.secret": "cs_test",
            "server.currency": "EUR",
            "setup.completed": "true",
        }
    )
    _cache_product(db, "101", "Test Product", "SKU-101", "100.00")

    saved = client.put(
        "/api/v2/commerce/sources/nextcloud:primary/settings",
        headers=auth_headers,
        json={
            "enabled": True,
            "settings": {
                "url": "https://softpple.business",
                "username": "woo",
                "spreadsheet_path": "/Selected/Prices.xlsx",
            },
            "secrets": {"password": "app-password-secret"},
        },
    )
    assert saved.status_code == 200
    assert "app-password-secret" not in saved.text

    async def fake_download(self, path):
        assert path == "/Selected/Prices.xlsx"
        return _xlsx([["Test Product", 101, "110.00", "SKU-101"]]), {"etag": "etag-1", "last_modified": "now"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    preview = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert preview.status_code == 200
    row = preview.json()["rows"][0]
    assert row["source"]["sourceFilePath"] == "/Selected/Prices.xlsx"
    assert row["eligible_for_dry_run"] is True


def test_variation_row_matched_by_variation_id_is_eligible_for_dry_run(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(configured_db, "100", "Parent Hoodie", "PARENT-100", "0.00", product_type="variable")
    _cache_product(
        configured_db,
        "201",
        "Parent Hoodie - Blue / XL",
        "VAR-201",
        "120.00",
        product_type="variation",
        parent_id="100",
        raw_data={"attributes": [{"name": "Color", "option": "Blue"}, {"name": "Size", "option": "XL"}]},
    )

    async def fake_download(self, path):
        return _xlsx([["Parent Hoodie - Blue / XL", 201, "132.00", "VAR-201"]]), {"etag": "etag-var"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    row = data["rows"][0]
    assert row["status"] == "valid_change"
    assert row["eligible_for_dry_run"] is True
    assert row["matchedProduct"]["itemType"] == "variation"
    assert row["matchedProduct"]["variationId"] == "201"
    assert row["matchedProduct"]["parentProductId"] == "100"
    assert row["matchedProduct"]["parentProductName"] == "Parent Hoodie"
    assert row["matchedProduct"]["variationAttributes"] == [
        {"name": "Color", "value": "Blue"},
        {"name": "Size", "value": "XL"},
    ]
    change = data["changes"][0]
    assert change["itemType"] == "variation"
    assert change["productId"] == "201"
    assert change["variationId"] == "201"
    assert change["parentProductId"] == "100"

    dry_run = client.post(
        "/api/v2/write-pipeline/dry-run",
        headers=auth_headers,
        json={
            "previewId": data["id"],
            "selectedRowIds": [row["id"]],
        },
    )
    assert dry_run.status_code == 201
    item = dry_run.json()["items"][0]
    assert item["itemType"] == "variation"
    assert item["variationId"] == "201"
    assert item["parentProductId"] == "100"
    assert item["variationAttributes"][0] == {"name": "Color", "value": "Blue"}


def test_variation_row_matched_by_sku_if_safe(client, auth_headers, configured_db, monkeypatch):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(configured_db).set_many(
        {
            "nextcloud.source_mapping": (
                '{"id":{"enabled":false,"column":"B"},'
                '"price":{"enabled":true,"column":"C"},'
                '"stock":{"enabled":false,"column":"D"}}'
            )
        },
        updated_by="test",
    )

    _cache_product(configured_db, "100", "Parent Hoodie", "PARENT-100", "0.00", product_type="variable")
    _cache_product(
        configured_db,
        "202",
        "Parent Hoodie - Red / L",
        "VAR-202",
        "100.00",
        product_type="variation",
        parent_id="100",
    )

    async def fake_download(self, path):
        return _xlsx([["Parent Hoodie - Red / L", "", "115.00", "VAR-202"]]), {"etag": "etag-var-sku"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["matchedProduct"]["variationId"] == "202"
    assert row["eligible_for_dry_run"] is True


def test_variation_row_missing_parent_id_fails_closed(client, auth_headers, configured_db, monkeypatch):
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(
        configured_db,
        "203",
        "Orphan Variation",
        "VAR-203",
        "100.00",
        product_type="variation",
        parent_id=None,
    )

    async def fake_download(self, path):
        return _xlsx([["Orphan Variation", 203, "110.00", "VAR-203"]]), {"etag": "etag-orphan"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["status"] == "error"
    assert "missing_variation_parent_id" in row["errors"]
    assert row["eligible_for_dry_run"] is False


def test_simple_woocommerce_price_workflow_end_to_end_with_mocked_adapter(
    client, auth_headers, configured_db, monkeypatch
):
    from app.connectors.destinations.woocommerce.write_adapter import WooCommercePriceWriteAdapter
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(configured_db, "101", "Test Product", "SKU-101", "100.00")

    async def fake_download(self, path):
        return _xlsx([["Test Product", 101, "110.00", "SKU-101"]]), {"etag": "etag-e2e", "last_modified": "now"}

    async def fake_execute(_adapter, item, _context):
        return {
            "provider": "woocommerce",
            "product_id": item.channel_product_id,
            "regular_price": f"{item.proposed_price:.2f}",
            "stock_update": False,
        }

    async def fake_verify(_adapter, item, _context):
        return {
            "provider": "woocommerce",
            "verified": True,
            "observed_price": item.proposed_price,
            "expected_price": item.proposed_price,
            "verification_error": None,
        }

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)
    monkeypatch.setattr(WooCommercePriceWriteAdapter, "execute_item", fake_execute)
    monkeypatch.setattr(WooCommercePriceWriteAdapter, "verify_item", fake_verify)

    preview = client.post("/api/v2/workspace/preview", headers=auth_headers).json()
    dry_run = client.post(
        "/api/v2/write-pipeline/dry-run",
        headers=auth_headers,
        json={
            "previewId": preview["id"],
            "selectedRowIds": [preview["rows"][0]["id"]],
        },
    ).json()
    approved = client.post(
        f"/api/v2/write-pipeline/batches/{dry_run['id']}/approve",
        headers=auth_headers,
        json={"reason": "owner approved"},
    )
    assert approved.status_code == 200

    enabled = client.put(
        "/api/v2/commerce/channels/woocommerce:primary/settings",
        headers=auth_headers,
        json={"access_mode": "write_enabled"},
    )
    assert enabled.status_code == 200

    result = client.post(f"/api/v2/write-pipeline/batches/{dry_run['id']}/execute", headers=auth_headers)

    assert result.status_code == 200
    data = result.json()
    assert data["status"] == "applied"
    assert data["items"][0]["providerResult"]["stock_update"] is False
    assert data["items"][0]["verification"]["verified"] is True
    assert data["resultSummary"]["success_count"] == 1
    assert data["resultSummary"]["verified_count"] == 1
    events = client.get(f"/api/v2/write-pipeline/batches/{dry_run['id']}/events", headers=auth_headers).json()
    assert {item["eventType"] for item in events} >= {
        "dry_run_created_from_preview",
        "approved",
        "item_applied",
        "execution_finished",
    }


def test_workspace_preview_classifies_validation_errors_warnings_and_unchanged(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(configured_db, "101", "Duplicate Product", "SKU-101", "100.00")
    _cache_product(configured_db, "102", "Cache Name", "SKU-102", "100.00")
    _cache_product(configured_db, "103", "Same Price", "SKU-103", "100.00")
    _cache_product(configured_db, "104", "Valid Product", "SKU-104", "100.00")
    _cache_product(configured_db, "105", "Large Warning", "SKU-105", "100.00")
    _cache_product(configured_db, "106", "Invalid Price", "SKU-106", "100.00")
    _cache_product(configured_db, "107", "Variation Product", "SKU-107", "100.00", product_type="variation")
    _cache_product(configured_db, "108", "SKU Match", "DUP-SKU", "100.00")
    _cache_product(configured_db, "109", "Huge Change", "SKU-109", "100.00")

    async def fake_download(self, path):
        return _xlsx(
            [
                ["Duplicate Product", 101, "110.00", "SKU-101"],
                ["Duplicate Product", 101, "111.00", "SKU-101B"],
                ["Wrong Name", 102, "120.00", "SKU-102"],
                ["Same Price", 103, "100.00", "SKU-103"],
                ["Valid Product", 104, "112.00", "SKU-104"],
                ["Large Warning", 105, "140.00", "SKU-105"],
                ["Invalid Price", 106, "abc", "SKU-106"],
                ["Missing Product", 999, "120.00", "SKU-999"],
                ["Variation Product", 107, "120.00", "SKU-107"],
                ["SKU Match", "", "120.00", "DUP-SKU"],
                ["SKU Match", "", "121.00", "DUP-SKU"],
                ["Missing Identifier", "", "120.00", ""],
                ["Huge Change", 109, "170.00", "SKU-109"],
            ]
        ), {"etag": "etag-2"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)
    assert response.status_code == 200
    data = response.json()
    rows = data["rows"]
    by_row = {item["source"]["rowNumber"]: item for item in rows}

    assert "duplicate_product_id" in by_row[3]["errors"]
    assert "product_name_mismatch" in by_row[5]["errors"]
    assert by_row[6]["status"] == "unchanged"
    assert by_row[7]["eligible_for_dry_run"] is True
    assert "large_price_change" in by_row[8]["warnings"]
    assert "invalid_or_missing_price" in by_row[9]["errors"]
    assert "missing_product" in by_row[10]["errors"]
    assert by_row[11]["status"] == "valid_change"
    assert by_row[11]["matchedProduct"]["itemType"] == "variation"
    assert by_row[11]["eligible_for_dry_run"] is True
    assert "duplicate_sku" in by_row[12]["errors"]
    assert "duplicate_sku" in by_row[13]["errors"]
    assert "missing_product_identifier" in by_row[14]["errors"]
    assert "large_price_change_blocked" in by_row[15]["errors"]
    assert data["summary"]["error_rows"] == 9
    assert data["summary"]["unchanged_rows"] == 1
    assert data["summary"]["large_changes"] == 2

    change_ids = {item["productId"] for item in data["changes"]}
    assert change_ids == {"104", "105", "107"}
    assert all("stock" not in item for item in data["changes"])


def test_rows_with_errors_are_rejected_by_write_pipeline(client, auth_headers):
    response = client.post(
        "/api/v2/write-pipeline/dry-run",
        headers=auth_headers,
        json={
            "previewId": "wp_bad",
            "selectedRowIds": ["wp_bad:Sheet1:3"],
            "changes": [
                {
                    "productId": "101",
                    "productName": "Bad",
                    "sku": "SKU-101",
                    "currentPrice": 100.0,
                    "proposedPrice": 110.0,
                    "currency": "EUR",
                    "eligible_for_dry_run": False,
                    "validationStatus": "error",
                    "source": {
                        "previewId": "wp_bad",
                        "sourceId": "nextcloud:primary",
                        "sourceType": "nextcloud_spreadsheet",
                        "sourceSnapshotId": 1,
                        "sourceSnapshotVersion": 1,
                        "sourceFilePath": "/prices.xlsx",
                        "worksheet": "Sheet1",
                        "rowNumber": 3,
                    },
                }
            ],
        },
    )
    assert response.status_code == 422
    assert "DRY_RUN_REQUEST_FIELDS_INVALID" in response.text


def test_preview_fails_safely_when_cache_is_empty(client, auth_headers, configured_db, monkeypatch):
    from app.flowhub.integrations.nextcloud import NextcloudClient

    async def fake_download(self, path):
        raise AssertionError("empty product cache must fail before source download")

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)
    response = client.post("/api/v2/workspace/preview", headers=auth_headers)
    assert response.status_code == 409
    assert "product cache is empty" in response.text


@pytest.mark.parametrize("refresh_status", ["failed", "partial_failed"])
def test_preview_blocks_incomplete_cache_refresh(client, auth_headers, configured_db, monkeypatch, refresh_status):
    from app.flowhub.data_layer.models import DlRefreshJob
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(configured_db, "101", "Blocked Product", "SKU-101", "100.00")
    configured_db.add(
        DlRefreshJob(
            job_type="manual",
            entity_type="products",
            connector_id="woocommerce:primary",
            status=refresh_status,
            triggered_by="tester",
            created_at=datetime.utcnow(),
            completed_at=datetime.utcnow(),
            meta={"products_stored": 1},
        )
    )
    configured_db.commit()

    async def fake_download(self, path):
        raise AssertionError("incomplete cache refresh must fail before source download")

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)
    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 409
    assert "CACHE_REFRESH_INCOMPLETE" in response.text
    assert refresh_status in response.text
    assert "Refresh product cache again in Commerce Hub -> Channels." in response.text


def test_preview_allows_completed_with_warnings_cache_refresh(client, auth_headers, configured_db, monkeypatch):
    from app.flowhub.data_layer.models import DlRefreshJob
    from app.flowhub.integrations.nextcloud import NextcloudClient

    _cache_product(configured_db, "101", "Test Product", "SKU-101", "100.00")
    configured_db.add(
        DlRefreshJob(
            job_type="manual",
            entity_type="products",
            connector_id="woocommerce:primary",
            status="completed_with_warnings",
            triggered_by="tester",
            created_at=datetime.utcnow(),
            completed_at=datetime.utcnow(),
            meta={"products_stored": 1},
        )
    )
    configured_db.commit()

    async def fake_download(self, path):
        assert path == "/prices.xlsx"
        return _xlsx([["Test Product", 101, "110.00", "SKU-101"]]), {"etag": "etag-warning"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)
    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    assert data["duplicateWarnings"]
    assert "completed with warnings" in data["duplicateWarnings"][0]


def test_preview_fails_safely_when_channel_credentials_are_missing(client, auth_headers, db, monkeypatch):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(db).set_many(
        {
            "nextcloud.url": "https://cloud.example.test",
            "nextcloud.username": "user",
            "nextcloud.password": "pass",
            "nextcloud.spreadsheet_path": "/prices.xlsx",
        }
    )
    _cache_product(db, "101", "Test Product", "SKU-101", "100.00")

    async def fake_download(self, path):
        raise AssertionError("missing channel credentials must fail before source download")

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)
    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 422
    assert "woocommerce.url" in response.text


def test_spreadsheet_normalizes_persian_digits_and_commas():
    from app.flowhub.integrations.spreadsheet import load_workbook_bytes, parse_source_price_rows

    wb = load_workbook_bytes(_xlsx([["Persian Product", "۱۰۱", "۱,۲۳۴.۵۰", "SKU-FA"]]))
    rows, duplicates = parse_source_price_rows(wb)
    assert rows[0]["product_id"] == "101"
    assert rows[0]["proposed_price"] == 1234.5
    assert duplicates["duplicate_product_ids"] == []


@pytest.mark.parametrize("raw_id", [101, "101", " 101 ", "\u06f1\u06f0\u06f1", "\u0661\u0660\u0661"])
def test_spreadsheet_accepts_only_positive_whole_product_ids(raw_id):
    from app.flowhub.integrations.spreadsheet import load_workbook_bytes, parse_source_price_rows

    rows, _ = parse_source_price_rows(load_workbook_bytes(_xlsx([["Product", raw_id, "110.00", "SKU-101"]])))

    assert rows[0]["product_id"] == "101"
    assert "invalid_product_id" not in rows[0]["row_errors"]


@pytest.mark.parametrize("raw_id", [101.9, "101.9", "1e3", "-10", "0", "101abc", True, ""])
def test_spreadsheet_rejects_noncanonical_product_ids(raw_id):
    from app.flowhub.integrations.spreadsheet import load_workbook_bytes, parse_source_price_rows

    rows, _ = parse_source_price_rows(load_workbook_bytes(_xlsx([["Product", raw_id, "110.00", "SKU-101"]])))

    assert rows[0]["product_id"] is None
    assert "invalid_product_id" in rows[0]["row_errors"]
    assert rows[0]["row_error_details"] == [{
        "code": "INVALID_PRODUCT_ID",
        "message": "Product ID must be a positive whole number.",
    }]


@pytest.mark.parametrize(
    ("raw_stock", "accepted"),
    [(0, True), ("0", True), (8, True), (8.5, False), ("8.5", False), ("1e3", False), (-1, False)],
)
def test_spreadsheet_stock_requires_non_negative_whole_number(raw_stock, accepted):
    from app.flowhub.integrations.spreadsheet import load_workbook_bytes, parse_source_price_rows

    mapping = {
        "id": {"enabled": True, "column": "B"},
        "price": {"enabled": True, "column": "C"},
        "stock": {"enabled": True, "column": "D"},
    }
    rows, _ = parse_source_price_rows(
        load_workbook_bytes(_xlsx([["Product", 101, "110.00", raw_stock]])),
        mapping=mapping,
    )

    if accepted:
        assert rows[0]["source_stock"] == int(raw_stock)
        assert "invalid_stock" not in rows[0]["row_errors"]
    else:
        assert rows[0]["source_stock"] is None
        assert "invalid_stock" in rows[0]["row_errors"]
        assert rows[0]["row_error_details"][-1]["code"] == "INVALID_STOCK"


def test_duplicate_product_ids_and_skus_are_detected_across_worksheets():
    import openpyxl

    from app.flowhub.integrations.spreadsheet import parse_source_price_rows

    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "First"
    second = workbook.create_sheet("Second")
    for sheet, row in (
        (first, ["First Product", 101, "110", "DUP-SKU"]),
        (second, ["Second Product", 101, "120", "DUP-SKU"]),
    ):
        sheet.append(["Name", "Product ID", "Price", "SKU"])
        sheet.append(["", "", "", ""])
        sheet.append(row)

    rows, duplicates = parse_source_price_rows(workbook)

    assert duplicates == {"duplicate_product_ids": ["101"], "duplicate_skus": ["dup-sku"]}
    assert all("duplicate_product_id" in row["row_errors"] for row in rows)
    assert all("duplicate_sku" in row["row_errors"] for row in rows)


def test_workspace_preview_uses_configured_mapping_and_reads_stock(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(configured_db).set_many(
        {
            "nextcloud.source_mapping": (
                '{"id":{"enabled":true,"column":"E"},'
                '"price":{"enabled":true,"column":"D"},'
                '"stock":{"enabled":true,"column":"B"}}'
            )
        },
        updated_by="test",
    )
    _cache_product(configured_db, "101", "Mapped Product", "SKU-101", "100.00", stock_qty=8)

    async def fake_download(self, path):
        return _xlsx_custom(
            headers=["Name", "Stock", "SKU", "Price", "Product ID"],
            rows=[["Mapped Product", "12", "SKU-101", "125.00", "101"]],
        ), {"etag": "etag-mapped"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    row = data["rows"][0]
    assert row["source"]["rawValues"] == {
        "A": "Mapped Product",
        "B": "12",
        "C": "SKU-101",
        "D": "125.00",
        "E": "101",
    }
    assert row["source"]["sourceStock"] == 12
    assert row["currentStock"] == 8
    assert row["sourceStock"] == 12
    assert row["stockDifference"] == 4
    assert row["changeType"] == "price_and_stock_changed"
    assert row["eligible_for_dry_run"] is True
    assert data["summary"]["changed_stock"] == 1
    assert "stock" not in data["changes"][0]
    assert "stockQuantity" not in data["changes"][0]


def test_price_disabled_generates_no_price_changes(client, auth_headers, configured_db, monkeypatch):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(configured_db).set_many(
        {
            "nextcloud.source_mapping": (
                '{"id":{"enabled":true,"column":"B"},'
                '"price":{"enabled":false,"column":"C"},'
                '"stock":{"enabled":true,"column":"D"}}'
            )
        },
        updated_by="test",
    )
    _cache_product(configured_db, "101", "Stock Only", "SKU-101", "100.00", stock_qty=5)

    async def fake_download(self, path):
        return _xlsx([["Stock Only", 101, "999.00", "8"]]), {"etag": "etag-price-disabled"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    row = data["rows"][0]
    assert row["status"] == "stock_changed"
    assert row["proposedPrice"] is None
    assert row["sourceStock"] == 8
    assert row["eligible_for_dry_run"] is False
    assert data["changes"] == []
    assert data["summary"]["estimated_woocommerce_updates"] == 0


def test_stock_only_row_is_visible_but_excluded_from_write_changes(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    AppConfigService(configured_db).set_many(
        {
            "nextcloud.source_mapping": (
                '{"id":{"enabled":true,"column":"B"},'
                '"price":{"enabled":true,"column":"C"},'
                '"stock":{"enabled":true,"column":"D"}}'
            )
        },
        updated_by="test",
    )
    _cache_product(configured_db, "101", "Stock Only", "SKU-101", "100.00", stock_qty=5)

    async def fake_download(self, path):
        return _xlsx([["Stock Only", 101, "100.00", "8"]]), {"etag": "etag-stock-only"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)

    assert response.status_code == 200
    data = response.json()
    row = data["rows"][0]
    assert row["status"] == "stock_changed"
    assert row["sourceStock"] == 8
    assert row["stockDifference"] == 3
    assert row["eligible_for_dry_run"] is False
    assert data["changes"] == []


def test_selected_worksheet_is_read_and_missing_worksheet_fails(
    client, auth_headers, configured_db, monkeypatch
):
    from app.flowhub.integrations.nextcloud import NextcloudClient
    from app.flowhub.setup.service import AppConfigService

    _cache_product(configured_db, "101", "Selected Product", "SKU-101", "100.00")

    async def fake_download(self, path):
        return _xlsx_multi_sheet(), {"etag": "etag-sheets"}

    monkeypatch.setattr(NextcloudClient, "download_file", fake_download)
    AppConfigService(configured_db).set_many(
        {"nextcloud.worksheet_mode": "selected", "nextcloud.worksheet_name": "Prices"},
        updated_by="test",
    )

    response = client.post("/api/v2/workspace/preview", headers=auth_headers)
    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["source"]["worksheet"] == "Prices"
    assert row["source"]["productName"] == "Selected Product"

    AppConfigService(configured_db).set_many({"nextcloud.worksheet_name": "Missing"}, updated_by="test")
    missing = client.post("/api/v2/workspace/preview", headers=auth_headers)
    assert missing.status_code == 422
    assert "Selected worksheet not found" in missing.text


def _cache_product(
    db,
    product_id: str,
    name: str,
    sku: str,
    price: str,
    *,
    product_type: str = "simple",
    parent_id: str | None = "100",
    raw_data: dict | None = None,
    stock_qty: int | None = None,
    stock_status: str | None = "instock",
    manage_stock: bool | None = True,
) -> None:
    from app.flowhub.data_layer.models import DlProductCache

    db.add(
        DlProductCache(
            connector_id="woocommerce:primary",
            product_id=product_id,
            external_id=int(product_id),
            sku=sku,
            name=name,
            product_type=product_type,
            parent_id=parent_id if product_type == "variation" else None,
            regular_price=price,
            price=price,
            stock_qty=stock_qty,
            stock_status=stock_status,
            manage_stock=manage_stock,
            categories=[{"name": "Default"}],
            images=[{"src": f"https://example.test/{product_id}.jpg"}],
            freshness="fresh",
            exists=True,
            raw_data=raw_data or {},
            last_fetched_at=datetime.utcnow(),
        )
    )
    db.commit()


def _xlsx(rows: list[list[object]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Product ID", "Price", "SKU"])
    ws.append(["", "", "", ""])
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _xlsx_custom(headers: list[str], rows: list[list[object]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    ws.append(["" for _ in headers])
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _xlsx_multi_sheet() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ignore"
    ws.append(["Name", "Product ID", "Price", "SKU"])
    ws.append(["", "", "", ""])
    ws.append(["Wrong Sheet", 999, "130.00", "SKU-999"])
    prices = wb.create_sheet("Prices")
    prices.append(["Name", "Product ID", "Price", "SKU"])
    prices.append(["", "", "", ""])
    prices.append(["Selected Product", 101, "120.00", "SKU-101"])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
