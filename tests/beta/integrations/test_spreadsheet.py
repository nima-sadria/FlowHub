"""Tests for the BU5 XLSX spreadsheet parser.

All tests use in-memory openpyxl workbooks — no file I/O needed.
"""

from __future__ import annotations

import openpyxl

from app.beta.integrations.spreadsheet import parse_price_list, _normalize_price_text


# ── _normalize_price_text ─────────────────────────────────────────────────────

class TestNormalizePriceText:
    def test_ascii_digits_unchanged(self):
        assert _normalize_price_text("1234.50") == "1234.50"

    def test_persian_digits(self):
        assert _normalize_price_text("۱۲۳") == "123"

    def test_arabic_indic_digits(self):
        assert _normalize_price_text("٤٥٦") == "456"

    def test_arabic_thousands_separator_removed(self):
        # U+066C ARABIC THOUSANDS SEPARATOR
        assert _normalize_price_text("1٬234") == "1234"

    def test_comma_removed(self):
        assert _normalize_price_text("1,234.00") == "1234.00"

    def test_strips_whitespace(self):
        assert _normalize_price_text("  99.9  ") == "99.9"


# ── parse_price_list ──────────────────────────────────────────────────────────

def _make_wb(*sheets: list[tuple]) -> openpyxl.Workbook:
    """Helper: build a workbook from a list of (sheet_name, rows) pairs.

    Each row is (name, product_id, price). Rows start at row 3 (rows 1–2 = header).
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    if default:
        wb.remove(default)

    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=sheet_name)
        # rows 1–2 are header rows (left blank)
        ws.append(["Product Name", "Product ID", "Price"])  # row 1
        ws.append([])                                         # row 2
        for row in rows:
            ws.append(list(row))

    return wb


class TestParsePriceList:
    def test_basic_parse(self):
        wb = _make_wb(("Sheet1", [
            ("Widget A", 101, 9.99),
            ("Widget B", 102, 14.50),
        ]))
        entries, duplicates = parse_price_list(wb)
        assert 101 in entries
        assert entries[101]["price"] == 9.99
        assert entries[101]["name"] == "Widget A"
        assert 102 in entries
        assert not duplicates

    def test_last_sheet_wins_on_duplicate(self):
        wb = _make_wb(
            ("Sheet1", [("Product", 200, 10.00)]),
            ("Sheet2", [("Product", 200, 20.00)]),
        )
        entries, duplicates = parse_price_list(wb)
        assert entries[200]["price"] == 20.00
        assert entries[200]["sheet_name"] == "Sheet2"
        assert len(duplicates) == 1
        dup = duplicates[0]
        assert dup["product_id"] == 200
        assert dup["prev_sheet"] == "Sheet1"
        assert dup["final_sheet"] == "Sheet2"
        assert dup["prev_price"] == 10.00
        assert dup["final_price"] == 20.00

    def test_out_of_stock_markers_give_none_price(self):
        wb = _make_wb(("Sheet1", [
            ("OOS", 300, "ناموجود"),
            ("OOS2", 301, "out of stock"),
            ("OOS3", 302, "-"),
            ("OOS4", 303, "0"),
        ]))
        entries, _ = parse_price_list(wb)
        for pid in (300, 301, 302, 303):
            assert entries[pid]["price"] is None
            assert not entries[pid]["price_parse_error"]

    def test_unparseable_price_sets_error_flag(self):
        wb = _make_wb(("Sheet1", [("Bad", 400, "not-a-price")]))
        entries, _ = parse_price_list(wb)
        assert entries[400]["price"] is None
        assert entries[400]["price_parse_error"] is True
        assert entries[400]["warning"] is not None

    def test_persian_digits_in_price(self):
        wb = _make_wb(("Sheet1", [("Widget", 500, "۱۰۰.۵")]))
        entries, _ = parse_price_list(wb)
        assert entries[500]["price"] == pytest.approx(100.5)

    def test_skips_rows_with_no_product_id(self):
        wb = _make_wb(("Sheet1", [
            ("With ID", 601, 5.00),
            ("No ID", None, 5.00),
            ("With ID 2", 602, 7.00),
        ]))
        entries, _ = parse_price_list(wb)
        assert 601 in entries
        assert 602 in entries
        assert len(entries) == 2

    def test_skips_non_integer_product_id(self):
        wb = _make_wb(("Sheet1", [("Bad ID", "abc", 5.00)]))
        entries, _ = parse_price_list(wb)
        assert len(entries) == 0

    def test_negative_price_is_error(self):
        wb = _make_wb(("Sheet1", [("Neg", 700, -5.00)]))
        entries, _ = parse_price_list(wb)
        assert entries[700]["price"] is None
        assert entries[700]["price_parse_error"] is True

    def test_multiple_sheets_all_read(self):
        wb = _make_wb(
            ("Sheet1", [("A", 1, 1.0), ("B", 2, 2.0)]),
            ("Sheet2", [("C", 3, 3.0)]),
            ("Sheet3", [("D", 4, 4.0)]),
        )
        entries, _ = parse_price_list(wb)
        assert set(entries.keys()) == {1, 2, 3, 4}

    def test_sheet_name_stored_in_entry(self):
        wb = _make_wb(("MyPrices", [("Widget", 800, 50.0)]))
        entries, _ = parse_price_list(wb)
        assert entries[800]["sheet_name"] == "MyPrices"

    def test_empty_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Empty"
        entries, duplicates = parse_price_list(wb)
        assert entries == {}
        assert duplicates == []


import pytest  # noqa: E402 (imported after test body for clarity)
